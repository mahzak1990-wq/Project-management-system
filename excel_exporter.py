import pandas as pd
from datetime import datetime, date
import io
import hashlib
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from data_manager import DataManager
from evm_calculator import EVMCalculator

class ExcelExporter:
    def __init__(self, data_manager: DataManager):
        self.data_manager = data_manager
        self.evm_calculator = EVMCalculator(data_manager)
        self.max_columns = 730  # Optimized for 2 years of daily tracking (365*2)
    
    def generate_project_cash_flow_report(self, project_name: str, start_date: date, end_date: date) -> Optional[pd.DataFrame]:
        """Generate cash flow report for a specific project"""
        try:
            cash_flow_data = self.data_manager.get_cash_flow_data(project_name, start_date, end_date)
            
            if cash_flow_data.empty:
                return None
            
            # Process the data for reporting
            cash_flow_data['entry_date'] = pd.to_datetime(cash_flow_data['entry_date'])
            cash_flow_data = cash_flow_data.sort_values('entry_date')
            
            # Calculate cumulative values
            cash_flow_data['cumulative_planned'] = cash_flow_data['planned_cost'].cumsum()
            cash_flow_data['cumulative_actual'] = cash_flow_data['actual_cost'].cumsum()
            cash_flow_data['variance'] = cash_flow_data['actual_cost'] - cash_flow_data['planned_cost']
            cash_flow_data['cumulative_variance'] = cash_flow_data['variance'].cumsum()
            
            # Format for display
            report_data = cash_flow_data[['entry_date', 'planned_cost', 'actual_cost', 'variance', 
                                        'cumulative_planned', 'cumulative_actual', 'cumulative_variance']]
            
            return report_data
        except Exception as e:
            print(f"Error generating project cash flow report: {e}")
            return None
    
    def generate_portfolio_cash_flow_report(self, start_date: date, end_date: date) -> Optional[pd.DataFrame]:
        """Generate cash flow report for the entire portfolio"""
        try:
            cash_flow_data = self.data_manager.get_cash_flow_data(None, start_date, end_date)
            
            if cash_flow_data.empty:
                return None
            
            # Group by date and sum across all projects
            cash_flow_data['entry_date'] = pd.to_datetime(cash_flow_data['entry_date'])
            grouped_data = cash_flow_data.groupby('entry_date').agg({
                'planned_cost': 'sum',
                'actual_cost': 'sum'
            }).reset_index()
            
            grouped_data = grouped_data.sort_values('entry_date')
            
            # Calculate cumulative values and variance
            grouped_data['cumulative_planned'] = grouped_data['planned_cost'].cumsum()
            grouped_data['cumulative_actual'] = grouped_data['actual_cost'].cumsum()
            grouped_data['variance'] = grouped_data['actual_cost'] - grouped_data['planned_cost']
            grouped_data['cumulative_variance'] = grouped_data['variance'].cumsum()
            
            return grouped_data
        except Exception as e:
            print(f"Error generating portfolio cash flow report: {e}")
            return None
    
    def export_cash_flow_to_excel(self, data: pd.DataFrame, project_name: str, start_date: date, end_date: date) -> Optional[bytes]:
        """Export cash flow data to Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "تقرير التدفق النقدي"
            
            # Set Arabic text alignment
            arabic_alignment = Alignment(horizontal='right', vertical='center')
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            # Add title
            ws['A1'] = f"تقرير التدفق النقدي - {project_name if project_name != 'جميع المشاريع' else 'المحفظة الكاملة'}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = arabic_alignment
            ws.merge_cells('A1:G1')
            
            # Add date range
            ws['A2'] = f"من {start_date} إلى {end_date}"
            ws['A2'].alignment = arabic_alignment
            ws.merge_cells('A2:G2')
            
            # Add headers
            headers = ['التاريخ', 'التكلفة المخططة', 'التكلفة الفعلية', 'الانحراف', 
                      'التكلفة المخططة التراكمية', 'التكلفة الفعلية التراكمية', 'الانحراف التراكمي']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = arabic_alignment
            
            # Add data
            for row_idx, (_, row) in enumerate(data.iterrows(), 5):
                ws.cell(row=row_idx, column=1, value=row['entry_date'].strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=2, value=row['planned_cost'])
                ws.cell(row=row_idx, column=3, value=row['actual_cost'])
                ws.cell(row=row_idx, column=4, value=row['variance'])
                ws.cell(row=row_idx, column=5, value=row['cumulative_planned'])
                ws.cell(row=row_idx, column=6, value=row['cumulative_actual'])
                ws.cell(row=row_idx, column=7, value=row['cumulative_variance'])
                
                # Apply formatting
                for col in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.alignment = arabic_alignment
                    if col > 1:  # Format currency columns
                        cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add worksheets data table to the export
            self._add_worksheets_data_table(wb)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None
    
    def export_portfolio_kpi_to_excel(self, kpi_data: Dict) -> Optional[bytes]:
        """Export portfolio KPI data to Excel format"""
        try:
            wb = Workbook()
            
            # Overview sheet
            ws_overview = wb.active
            ws_overview.title = "ملخص المحفظة"
            
            # Formatting
            arabic_alignment = Alignment(horizontal='right', vertical='center')
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            
            # Title
            ws_overview['A1'] = "تقرير مؤشرات أداء المحفظة"
            ws_overview['A1'].font = title_font
            ws_overview['A1'].alignment = arabic_alignment
            ws_overview.merge_cells('A1:D1')
            
            # Date
            ws_overview['A2'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"
            ws_overview['A2'].alignment = arabic_alignment
            ws_overview.merge_cells('A2:D2')
            
            # KPI Summary
            kpi_data_rows = [
                ['عدد المشاريع الإجمالي', kpi_data.get('total_projects', 0)],
                ['القيمة المخططة الإجمالية (PV)', kpi_data.get('total_pv', 0)],
                ['القيمة المكتسبة الإجمالية (EV)', kpi_data.get('total_ev', 0)],
                ['التكلفة الفعلية الإجمالية (AC)', kpi_data.get('total_ac', 0)],
                ['مؤشر أداء التكلفة المتوسط (CPI)', kpi_data.get('avg_cpi', 0)],
                ['مؤشر أداء الجدولة المتوسط (SPI)', kpi_data.get('avg_spi', 0)],
                ['انحراف التكلفة الإجمالي (CV)', kpi_data.get('total_cv', 0)],
                ['انحراف الجدولة الإجمالي (SV)', kpi_data.get('total_sv', 0)]
            ]
            
            # Add KPI data
            start_row = 4
            for idx, (label, value) in enumerate(kpi_data_rows):
                row = start_row + idx
                ws_overview.cell(row=row, column=1, value=label).alignment = arabic_alignment
                ws_overview.cell(row=row, column=1).font = header_font
                
                if isinstance(value, (int, float)):
                    if 'مؤشر' in label or 'CPI' in label or 'SPI' in label:
                        ws_overview.cell(row=row, column=2, value=f"{value:.3f}")
                    else:
                        ws_overview.cell(row=row, column=2, value=value)
                        ws_overview.cell(row=row, column=2).number_format = '#,##0.00'
                else:
                    ws_overview.cell(row=row, column=2, value=value)
                
                ws_overview.cell(row=row, column=2).alignment = arabic_alignment
            
            # Status distribution
            status_counts = kpi_data.get('status_counts', {})
            if status_counts:
                status_start_row = start_row + len(kpi_data_rows) + 2
                ws_overview.cell(row=status_start_row, column=1, value="توزيع حالات المشاريع").font = title_font
                ws_overview.cell(row=status_start_row, column=1).alignment = arabic_alignment
                
                for idx, (status, count) in enumerate(status_counts.items()):
                    row = status_start_row + idx + 1
                    ws_overview.cell(row=row, column=1, value=status).alignment = arabic_alignment
                    ws_overview.cell(row=row, column=2, value=count).alignment = arabic_alignment
            
            # Project details sheet
            if 'project_details' in kpi_data and kpi_data['project_details']:
                ws_projects = wb.create_sheet("تفاصيل المشاريع")
                
                # Headers
                project_headers = ['اسم المشروع', 'القيمة المخططة (PV)', 'القيمة المكتسبة (EV)', 
                                 'التكلفة الفعلية (AC)', 'مؤشر CPI', 'مؤشر SPI', 'الحالة']
                
                for col, header in enumerate(project_headers, 1):
                    cell = ws_projects.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = arabic_alignment
                
                # Project data
                for row_idx, project in enumerate(kpi_data['project_details'], 2):
                    ws_projects.cell(row=row_idx, column=1, value=project['project_name']).alignment = arabic_alignment
                    ws_projects.cell(row=row_idx, column=2, value=project['pv']).number_format = '#,##0.00'
                    ws_projects.cell(row=row_idx, column=3, value=project['ev']).number_format = '#,##0.00'
                    ws_projects.cell(row=row_idx, column=4, value=project['ac']).number_format = '#,##0.00'
                    ws_projects.cell(row=row_idx, column=5, value=f"{project['cpi']:.3f}")
                    ws_projects.cell(row=row_idx, column=6, value=f"{project['spi']:.3f}")
                    ws_projects.cell(row=row_idx, column=7, value=project['status']).alignment = arabic_alignment
            
            # Auto-adjust column widths for all sheets
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add worksheets data table to the export
            self._add_worksheets_data_table(wb)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        except Exception as e:
            print(f"Error exporting KPI to Excel: {e}")
            return None
    
    def _add_worksheets_data_table(self, workbook) -> None:
        """Add a worksheet containing all worksheet data with proper formatting"""
        try:
            # Get the latest original Excel file
            original_file_data = self.data_manager.get_latest_original_excel_file()
            
            if not original_file_data or not original_file_data.get('file_content'):
                return  # No original file to reference
            
            import openpyxl
            from io import BytesIO
            
            # Read the original Excel file
            file_content = original_file_data['file_content']
            original_wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            
            if not original_wb.worksheets:
                return  # No worksheets to process
            
            # Create new worksheet in the export workbook
            ws_data = workbook.create_sheet("جدول أوراق العمل الكاملة")
            
            # Formatting
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            date_fill = PatternFill(start_color="e8f5e8", end_color="e8f5e8", fill_type="solid")
            number_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")
            center_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            
            # Border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add title and file info
            ws_data['A1'] = "جدول أوراق العمل الكاملة - Complete Worksheets Data"
            ws_data['A1'].font = Font(bold=True, size=14)
            ws_data['A1'].alignment = right_alignment
            ws_data.merge_cells('A1:AX1')  # Merge across many columns
            
            ws_data['A2'] = f"الملف الأصلي: {original_file_data.get('file_name', 'غير محدد')} - تاريخ الاستيراد: {original_file_data.get('imported_date', 'غير محدد')}"
            ws_data['A2'].font = Font(size=10, color="7f8c8d")
            ws_data['A2'].alignment = right_alignment
            ws_data.merge_cells('A2:AX2')
            
            # Get first worksheet from original file
            original_ws = original_wb.worksheets[0]
            
            # Read data from original worksheet
            data = []
            for row in original_ws.iter_rows(values_only=True):
                if row and any(cell is not None for cell in row):  # Skip empty rows
                    data.append(row)
            
            if not data:
                return  # No data to process
            
            # Get maximum columns with data (limit to 50 for export)
            max_cols = min(50, max(len(row) for row in data if row))
            
            # Header row with column letters
            header_row = 4  # Start after title and info
            ws_data.cell(row=header_row, column=1, value="رقم الصف").font = header_font
            ws_data.cell(row=header_row, column=1).fill = header_fill
            ws_data.cell(row=header_row, column=1).alignment = right_alignment
            ws_data.cell(row=header_row, column=1).border = thin_border
            
            for col_idx in range(max_cols):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                cell = ws_data.cell(row=header_row, column=col_idx + 2, value=col_letter)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Data rows (limit to first 30 rows for export)
            for row_idx, row_data in enumerate(data[:30]):
                if row_idx == 0:
                    continue  # Skip first row (usually headers)
                
                excel_row = header_row + row_idx
                actual_row_number = row_idx + 1
                
                # Row number column
                row_cell = ws_data.cell(row=excel_row, column=1, value=actual_row_number)
                row_cell.font = Font(bold=True, color="34495e")
                row_cell.alignment = center_alignment
                row_cell.border = thin_border
                
                # Determine if this is a date row (21-24)
                is_date_row = 21 <= actual_row_number <= 24
                
                # Fill row data
                for col_idx in range(max_cols):
                    cell_value = row_data[col_idx] if col_idx < len(row_data) else None
                    cell = ws_data.cell(row=excel_row, column=col_idx + 2)
                    
                    if cell_value is not None:
                        if is_date_row:
                            # Format as date for rows 21-24
                            try:
                                if isinstance(cell_value, (int, float)):
                                    # Excel date number
                                    date_val = pd.to_datetime('1899-12-30') + pd.Timedelta(days=int(cell_value))
                                    cell.value = date_val.date()
                                    cell.number_format = 'DD-MM-YYYY'
                                elif isinstance(cell_value, str):
                                    # Try to parse as date
                                    date_val = pd.to_datetime(cell_value, errors='coerce')
                                    if pd.notna(date_val):
                                        cell.value = date_val.date()
                                        cell.number_format = 'DD-MM-YYYY'
                                    else:
                                        cell.value = str(cell_value)
                                else:
                                    cell.value = cell_value
                                    if hasattr(cell_value, 'date'):
                                        cell.number_format = 'DD-MM-YYYY'
                            except:
                                cell.value = str(cell_value) if cell_value else ""
                            
                            cell.fill = date_fill
                            
                        else:
                            # Format as number for other rows
                            try:
                                if isinstance(cell_value, (int, float)):
                                    cell.value = cell_value
                                    if float(cell_value).is_integer():
                                        cell.number_format = '0'
                                    else:
                                        cell.number_format = '0.00'
                                else:
                                    # Try to convert to number
                                    try:
                                        num_val = float(str(cell_value).replace(',', ''))
                                        cell.value = num_val
                                        if num_val.is_integer():
                                            cell.number_format = '0'
                                        else:
                                            cell.number_format = '0.00'
                                    except:
                                        cell.value = str(cell_value) if cell_value else ""
                            except:
                                cell.value = str(cell_value) if cell_value else ""
                            
                            cell.fill = number_fill
                    else:
                        cell.value = ""
                        cell.fill = number_fill if not is_date_row else date_fill
                    
                    cell.alignment = center_alignment
                    cell.border = thin_border
            
            # Add legend at the bottom
            legend_row = header_row + min(30, len(data)) + 2
            ws_data.cell(row=legend_row, column=1, value="ملاحظات:").font = Font(bold=True)
            ws_data.cell(row=legend_row, column=1).alignment = right_alignment
            
            ws_data.cell(row=legend_row + 1, column=1, value="• الصفوف 21-24: منسقة كتواريخ")
            ws_data.cell(row=legend_row + 1, column=1).alignment = right_alignment
            ws_data.cell(row=legend_row + 1, column=1).fill = date_fill
            
            ws_data.cell(row=legend_row + 2, column=1, value="• الصفوف الأخرى: منسقة كأرقام")
            ws_data.cell(row=legend_row + 2, column=1).alignment = right_alignment
            ws_data.cell(row=legend_row + 2, column=1).fill = number_fill
            
            # Auto-adjust column widths
            for col_idx in range(1, max_cols + 2):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if col_idx == 1:
                    ws_data.column_dimensions[col_letter].width = 12  # Row number column
                else:
                    ws_data.column_dimensions[col_letter].width = 10  # Data columns
                    
        except Exception as e:
            print(f"Error adding worksheets data table: {e}")
            # Don't fail the entire export if this fails
            pass

    def _add_new_budget_table_to_existing_sheets(self, workbook) -> None:
        """Add new budget table below existing data in each worksheet"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            print(f"DEBUG - Starting to add budget table to {len(workbook.worksheets)} worksheets")
            # Process each worksheet in the workbook
            for ws_idx, ws in enumerate(workbook.worksheets):
                print(f"DEBUG - Processing worksheet {ws_idx + 1}: {ws.title}")
                # Find the last row with data
                last_row = 1
                for row in range(1, ws.max_row + 1):
                    if any(ws.cell(row, col).value is not None for col in range(1, ws.max_column + 1)):
                        last_row = row
                
                # Start adding the table 3 rows below the last data
                start_row = last_row + 3
                print(f"DEBUG - Adding budget table to worksheet '{ws.title}' starting at row {start_row}")
                
                # Formatting
                header_font = Font(bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                data_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # Border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Find how many columns the upper table has (to match the span)
                upper_table_max_col = 1
                for col in range(1, ws.max_column + 1):
                    if any(ws.cell(row, col).value is not None for row in range(1, last_row + 1)):
                        upper_table_max_col = col
                
                print(f"DEBUG - Upper table spans {upper_table_max_col} columns")
                
                # Define row headers (what used to be column headers)
                row_headers = [
                    "Date",
                    "Budgeted Labor Units Weekly", 
                    "Budgeted Nonlabor Units Weekly",
                    "Date",
                    "Budgeted Labor Units Monthly",
                    "Budgeted Nonlabor Units Monthly"
                ]
                
                # Create the table with row headers in first column and data spanning to match upper table
                for row_idx, header in enumerate(row_headers):
                    current_row = start_row + row_idx
                    
                    # Add row header in first column
                    header_cell = ws.cell(row=current_row, column=1, value=header)
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = center_alignment
                    header_cell.border = thin_border
                    
                    # Set column width for header column
                    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 8, 30)
                    
                    # Fill data cells across all columns to match upper table span
                    for col_idx in range(2, upper_table_max_col + 1):
                        data_cell = ws.cell(row=current_row, column=col_idx)
                        data_cell.fill = data_fill
                        data_cell.alignment = center_alignment
                        data_cell.border = thin_border
                        
                        # Set format based on row type
                        if "Date" in header:
                            data_cell.number_format = 'DD-MM-YYYY'
                        else:
                            data_cell.number_format = '0.00'
                
                print(f"DEBUG - Successfully added budget table to worksheet '{ws.title}'")
                
        except Exception as e:
            print(f"Error adding budget table to existing sheets: {e}")
            # Don't fail the entire export if this fails
            pass

    def _remove_third_table_from_all_sheets(self, workbook) -> None:
        """حذف الجدول الثالث من جميع أوراق العمل (الصفوف 25-30)"""
        try:
            print(f"DEBUG - حذف الجدول الثالث من {len(workbook.worksheets)} ورقة عمل")
            
            # معالجة كل ورقة عمل في الكتاب
            for ws_idx, ws in enumerate(workbook.worksheets):
                print(f"DEBUG - معالجة ورقة العمل {ws_idx + 1}: {ws.title}")
                
                # حذف الصفوف 25-30 (الجدول الثالث)
                # حذف من الأسفل إلى الأعلى لتجنب مشاكل الفهرسة
                for row_num in range(30, 24, -1):  # من 30 إلى 25
                    try:
                        ws.delete_rows(row_num)
                        print(f"DEBUG - تم حذف الصف {row_num} من ورقة '{ws.title}'")
                    except Exception as e:
                        print(f"DEBUG - خطأ في حذف الصف {row_num}: {e}")
                        continue
                
                print(f"DEBUG - تم حذف الجدول الثالث من ورقة '{ws.title}' بنجاح")
                
        except Exception as e:
            print(f"خطأ في حذف الجدول الثالث من أوراق العمل: {e}")
            # عدم إفشال التصدير بالكامل إذا فشل هذا
            pass

    def export_project_template(self, existing_projects: List = None) -> Optional[bytes]:
        """Export project template Excel file using the original imported file with added worksheet data table"""
        
        # Always use the original Excel file to preserve exact formatting and data
        original_file = self.data_manager.get_latest_original_excel_file()
        if original_file and original_file.get('file_content'):
            print("DEBUG - Using original Excel file for export and adding worksheet data table below existing data")
            
            try:
                import openpyxl
                from io import BytesIO
                import pandas as pd
                
                # Load the original workbook
                file_content = original_file['file_content']
                wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
                
                # الجدول الثالث محذوف حسب طلب المستخدم (الصفوف 25-30)
                # self._add_new_budget_table_to_existing_sheets(wb)
                
                # حذف الجدول الثالث من جميع أوراق العمل (الصفوف 25-30)
                self._remove_third_table_from_all_sheets(wb)
                
                # Save the modified workbook
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                return excel_buffer.getvalue()
                
            except Exception as e:
                print(f"DEBUG - Error adding worksheet data table: {e}")
                # If error adding table, return original file
                return original_file['file_content']
        else:
            print("DEBUG - No original file found. Import an Excel file first to enable template export.")
            # Return None to indicate no template available - user needs to import first
            return None
    
    def _generate_new_template(self, existing_projects: List = None) -> Optional[bytes]:
        """Generate new project template Excel file matching the required table format with multiple sheets"""
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Ensure we always have at least 40 sheets
            if not existing_projects:
                existing_projects = [None] * 40  # Create 40 empty project templates
            else:
                # If we have existing projects, ensure we have at least 40 sheets
                while len(existing_projects) < 40:
                    existing_projects.append(None)
            
            # Remove the default sheet
            wb.remove(wb.active)
            
            # Create a sheet for each project
            for project_idx, project in enumerate(existing_projects):
                # Create sheet name in format [number].[project_code] but limit to 31 characters for Excel
                if project and project.get('project_id'):
                    base_name = f"{project_idx + 1}.{project['project_id']}"
                    # Truncate if too long, keeping the format but ensuring under 31 chars
                    if len(base_name) > 31:
                        project_code = project['project_id'][:26-(len(str(project_idx + 1)))]  # Leave space for number and dot
                        sheet_name = f"{project_idx + 1}.{project_code}"
                    else:
                        sheet_name = base_name
                else:
                    sheet_name = f"{project_idx + 1}.P{project_idx + 1:03d}"
                
                # Final safety check for 31 character limit
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                    
                ws = wb.create_sheet(title=sheet_name)
                
                # Formatting
                header_font = Font(bold=True, size=12, color="FFFFFF")
                field_font = Font(bold=True, size=10)
                data_font = Font(size=10)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                field_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                center_alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                
                # Border
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Main Header
                ws.merge_cells('A1:O1')
                header_cell = ws['A1']
                header_cell.value = "Project Management Data Template - Abdullah Al-Saeed Engineering Consulting Company"
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = center_alignment
                
                # Project information fields (row 3)
                ws['A3'].value = "Project Name:"
                ws['A3'].font = field_font
                ws['A3'].alignment = left_alignment
                
                ws['D3'].value = "Project ID:"
                ws['D3'].font = field_font
                ws['D3'].alignment = left_alignment
                
                ws['G3'].value = "Start Date:"
                ws['G3'].font = field_font
                ws['G3'].alignment = left_alignment
                
                ws['J3'].value = "End Date:"
                ws['J3'].font = field_font
                ws['J3'].alignment = left_alignment
                
                ws['M3'].value = "Planned Total Cost:"
                ws['M3'].font = field_font
                ws['M3'].alignment = left_alignment
                
                # Additional fields (row 4)
                ws['A4'].value = "Contractor:"
                ws['A4'].font = field_font
                ws['A4'].alignment = left_alignment
                
                ws['D4'].value = "Project Manager (Water Administration):"
                ws['D4'].font = field_font
                ws['D4'].alignment = left_alignment
                
                # Fill in project data if available
                if project:
                    ws['B3'].value = project.get('project_name', '')
                    ws['E3'].value = project.get('project_id', f"P{project_idx + 1:03d}")
                    ws['H3'].value = project.get('start_date', '')
                    ws['K3'].value = project.get('end_date', '')
                    ws['N3'].value = project.get('total_budget', '')
                    # Additional data fields
                    ws['B4'].value = project.get('contractor_name', '')
                    ws['E4'].value = project.get('project_manager', '')
                else:
                    ws['B3'].value = f"[Enter Project Name]"
                    ws['E3'].value = f"P{project_idx + 1:03d}"
                    ws['B4'].value = "[Enter Contractor Name]"
                    ws['E4'].value = "[Enter Project Manager]"
                
                # Set up the table structure with row headers (moved down to row 6)
                row_headers = [
                    'Dates',
                    'Planned Total Cost',
                    'Cum Budgeted Total Cost', 
                    'Planned % daily',
                    'Cum % daily',
                    'The elapsed period %',
                    'The elapsed period',
                    'Actual'
                ]
                
                # Create header row for the table (row 6)
                ws['A6'].value = ''  # Empty top-left cell
                ws['A6'].font = field_font
                ws['A6'].fill = field_fill
                ws['A6'].border = thin_border
                
                # Add column headers (B6 to BXL6) - 2000 columns for dates
                for col in range(2, 2002):  # Columns B to BXL (2000 columns)
                    col_letter = get_column_letter(col)
                    cell = ws[f'{col_letter}6']
                    cell.font = field_font
                    cell.fill = field_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                # Add row headers and create the grid structure (starting from row 7)
                for i, header in enumerate(row_headers):
                    row_num = 7 + i
                    cell = ws[f'A{row_num}']
                    cell.value = header
                    cell.font = field_font
                    cell.fill = field_fill
                    cell.alignment = left_alignment
                    cell.border = thin_border
                    
                    # Create bordered cells for data entry - 2000 columns
                    for col in range(2, 2002):  # Columns B to BXL (2000 columns)
                        col_letter = get_column_letter(col)
                        data_cell = ws[f'{col_letter}{row_num}']
                        data_cell.border = thin_border
                        data_cell.alignment = center_alignment
                        
                        # Special formatting for the Dates row
                        if header == 'Dates':
                            data_cell.font = field_font
                
                # Fill in existing project data if available
                if project:
                    progress_data = self.data_manager.get_progress_data(project['project_name'])
                    if not progress_data.empty:
                        # Sort by date and limit to available columns
                        progress_data = progress_data.sort_values('entry_date')
                        
                        for col_idx, (_, row_data) in enumerate(progress_data.iterrows()):
                            if col_idx >= 2000:  # Limit to available columns (B to BXL = 2000 columns)
                                break
                                
                            col_letter = get_column_letter(col_idx + 2)  # Start from column B
                            
                            # Fill data in each row
                            # ws[f'{col_letter}6'].value = row_data['entry_date']  # Dates - تم حذف هذا السطر لعدم الحاجة للتواريخ الوهمية
                            
                            # Extract data from notes field if available
                            notes = row_data.get('notes', '')
                            
                            # Helper function to extract value from notes
                            def extract_from_notes(notes_str, row_key):
                                try:
                                    if f'{row_key}:' in notes_str:
                                        value_str = notes_str.split(f'{row_key}:')[1].split('|')[0]
                                        return float(value_str) if value_str and value_str != 'None' else 0
                                    return 0
                                except:
                                    return 0
                            
                            # Extract values from notes
                            r7_value = extract_from_notes(notes, 'R7')  # Planned Total Cost
                            r8_value = extract_from_notes(notes, 'R8')  # Cum Budgeted Total Cost
                            r9_value = extract_from_notes(notes, 'R9')  # Planned % daily
                            r10_value = extract_from_notes(notes, 'R10')  # Cum % daily
                            r11_value = extract_from_notes(notes, 'R11')  # The elapsed period %
                            r12_value = extract_from_notes(notes, 'R12')  # The elapsed period
                            r13_value = extract_from_notes(notes, 'R13')  # Actual
                            
                            # Fill rows with proper formatting
                            # Row 7: Planned Total Cost (number format)
                            cell_7 = ws[f'{col_letter}7']
                            cell_7.value = r7_value
                            cell_7.number_format = '#,##0.00'
                            
                            # Row 8: Cum Budgeted Total Cost (number format)
                            cell_8 = ws[f'{col_letter}8']
                            cell_8.value = r8_value
                            cell_8.number_format = '#,##0.00'
                            
                            # Row 9: Planned % daily (number format as requested)
                            cell_9 = ws[f'{col_letter}9']
                            cell_9.value = r9_value
                            cell_9.number_format = '#,##0.00'
                            
                            # Row 10: Cum % daily (percentage format)
                            cell_10 = ws[f'{col_letter}10']
                            cell_10.value = r10_value / 100 if r10_value > 1 else r10_value
                            cell_10.number_format = '0.00%'
                            
                            # Row 11: The elapsed period % (percentage format)
                            cell_11 = ws[f'{col_letter}11']
                            cell_11.value = r11_value / 100 if r11_value > 1 else r11_value
                            cell_11.number_format = '0.00%'
                            
                            # Row 12: The elapsed period (number format)
                            cell_12 = ws[f'{col_letter}12']
                            cell_12.value = r12_value
                            cell_12.number_format = '#,##0.00'
                            
                            # Row 13: Actual (number format)
                            cell_13 = ws[f'{col_letter}13']
                            cell_13.value = r13_value
                            cell_13.number_format = '#,##0.00'
                
                # Adjust column widths to match the image layout
                ws.column_dimensions['A'].width = 25  # Row headers column
                for col in range(2, 2002):  # 2000 data columns
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = 12  # Data columns
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        except Exception as e:
            print(f"Error exporting project template: {e}")
            return None
    
    def import_project_template(self, uploaded_file) -> Dict:
        """Import project data from uploaded Excel template with table format"""
        error_details = []  # Track specific errors
        warnings = []  # Track warnings
        
        try:
            # Validate file format first
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                return {
                    'success': False,
                    'imported_count': 0,
                    'updated_count': 0,
                    'message': 'نوع الملف غير صحيح. يرجى تحميل ملف Excel (.xlsx أو .xls)',
                    'error_details': ['نوع الملف غير مدعوم - يجب أن يكون ملف Excel']
                }
            
            # Read the Excel file with multiple sheets
            try:
                wb = pd.ExcelFile(uploaded_file)
            except Exception as e:
                return {
                    'success': False,
                    'imported_count': 0,
                    'updated_count': 0,
                    'message': 'خطأ في قراءة ملف Excel. تأكد من أن الملف غير تالف',
                    'error_details': [f'فشل في قراءة الملف: {str(e)}']
                }
            
            success_count = 0
            updated_count = 0
            imported_projects = []
            
            # Validate that we have sheets to process
            if not wb.sheet_names:
                return {
                    'success': False,
                    'imported_count': 0,
                    'updated_count': 0,
                    'message': 'الملف فارغ أو لا يحتوي على أوراق عمل صحيحة',
                    'error_details': ['لم يتم العثور على أوراق عمل في الملف']
                }
            
            # Save the original Excel file before processing
            try:
                # Reset file pointer to beginning
                uploaded_file.seek(0)
                file_content = uploaded_file.read()
                file_hash = hashlib.md5(file_content).hexdigest()
                
                # Reset file pointer again for processing
                uploaded_file.seek(0)
                
                print(f"DEBUG - Saving original Excel file: {uploaded_file.name}, Hash: {file_hash}")
            except Exception as e:
                print(f"DEBUG - Error preparing file for save: {e}")
                file_content = None
                file_hash = None
            
            # Clear all existing projects before importing new ones
            print("DEBUG - Clearing all existing projects before import")
            try:
                self.data_manager.clear_all_data()
                # Also clear original Excel files
                self.data_manager.clear_original_excel_files()
            except Exception as e:
                error_details.append(f'فشل في مسح البيانات السابقة: {str(e)}')
            
            # Process each sheet in order to maintain sheet order
            for sheet_index, sheet_name in enumerate(wb.sheet_names):
                wb_data = pd.read_excel(uploaded_file, sheet_name=sheet_name, engine='openpyxl', header=None)
                
                # Extract project information from row 3 (index 2)
                if len(wb_data) > 2:
                    # Get project name from B3
                    project_name_cell = str(wb_data.iloc[2, 1]).strip() if not pd.isna(wb_data.iloc[2, 1]) else ""
                    if ":" in project_name_cell:
                        project_name = project_name_cell.split(":", 1)[1].strip()
                    else:
                        project_name = project_name_cell
                
                    # Skip empty or placeholder project names
                    if not project_name or project_name == "[Enter Project Name]":
                        warnings.append(f'تم تجاهل الورقة "{sheet_name}" - لا يوجد اسم مشروع صحيح')
                        print(f"DEBUG - Skipping sheet {sheet_name} - no valid project name found: '{project_name}'")
                        continue  # Skip this sheet and continue with next
                    
                    # Get project ID from E3 cell (column 4, row 3)
                    raw_project_id = str(wb_data.iloc[2, 4]).strip() if len(wb_data.columns) > 4 and not pd.isna(wb_data.iloc[2, 4]) else f"P001"
                    # Extract only the project code part (before any " - " or additional text)
                    if " - " in raw_project_id:
                        project_id = raw_project_id.split(" - ")[0].strip()
                    else:
                        project_id = raw_project_id
                    start_date_str = str(wb_data.iloc[2, 7]).strip() if len(wb_data.columns) > 7 and not pd.isna(wb_data.iloc[2, 7]) else ""
                    end_date_str = str(wb_data.iloc[2, 10]).strip() if len(wb_data.columns) > 10 and not pd.isna(wb_data.iloc[2, 10]) else ""
                    total_budget_str = str(wb_data.iloc[2, 13]).strip() if len(wb_data.columns) > 13 and not pd.isna(wb_data.iloc[2, 13]) else "0"
                    
                    # Get additional fields (row 4) - contractor and project manager
                    contractor_name = str(wb_data.iloc[3, 1]).strip() if len(wb_data) > 3 and not pd.isna(wb_data.iloc[3, 1]) else ""
                    project_manager = str(wb_data.iloc[3, 4]).strip() if len(wb_data) > 3 and not pd.isna(wb_data.iloc[3, 4]) else ""
                    
                    # Parse dates and budget with error handling
                    try:
                        start_date = pd.to_datetime(start_date_str).date() if start_date_str and start_date_str.lower() != 'nan' else datetime.now().date()
                    except Exception as e:
                        warnings.append(f'تاريخ البداية غير صحيح في المشروع {project_name}: {start_date_str}')
                        start_date = datetime.now().date()
                    
                    try:
                        end_date = pd.to_datetime(end_date_str).date() if end_date_str and end_date_str.lower() != 'nan' else datetime.now().date()
                    except Exception as e:
                        warnings.append(f'تاريخ الانتهاء غير صحيح في المشروع {project_name}: {end_date_str}')
                        end_date = datetime.now().date()
                    
                    try:
                        clean_budget = total_budget_str.replace('.', '').replace(',', '').replace(' ', '') if isinstance(total_budget_str, str) else str(total_budget_str)
                        total_budget = float(clean_budget) if clean_budget and clean_budget.replace('-', '').isdigit() else 0
                    except Exception as e:
                        warnings.append(f'الميزانية غير صحيحة في المشروع {project_name}: {total_budget_str}')
                        total_budget = 0
                
                    # Create project data dictionary
                    project_data = {
                        'project_name': project_name,
                        'project_id': project_id,
                        'executing_company': 'Default Company',
                        'consulting_company': 'Abdullah Al-Saeed Engineering Consulting Company',
                        'start_date': start_date,
                        'end_date': end_date,
                        'total_budget': total_budget,
                        'project_location': '',
                        'project_type': 'Construction Project',
                        'project_description': f'Project ID: {project_id}',
                        'display_order': sheet_index,  # Set order based on sheet position
                        'contractor_name': contractor_name,
                        'project_manager': project_manager,
                        'created_date': datetime.now()
                    }
                
                    # Add new project (since we cleared all existing projects)
                    if self.data_manager.add_project(project_data):
                        success_count += 1
                
                        imported_projects.append({'project_name': project_name, 'project_id': project_id, 'status': 'new', 'start_date': start_date, 'end_date': end_date, 'total_budget': total_budget})
                        current_project_name = project_name
                        print(f"DEBUG - Successfully added project {project_name}")
                    else:
                        print(f"DEBUG - Failed to add project {project_name}")
                        continue  # Skip this sheet and continue with next
                    
                    # Now import progress data from the table (moved down to row 7 due to additional fields)
                    # Look for the dates row (row 7, index 6) and get the column data 
                    if len(wb_data) > 6:
                        for col_idx in range(1, min(2001, len(wb_data.columns))):  # Columns B to BXL (2000 columns)
                            # Get date from row 7 (Dates row) - force date format
                            date_value = wb_data.iloc[6, col_idx] if not pd.isna(wb_data.iloc[6, col_idx]) else None
                            
                            # DEBUG: Show raw date value
                            if date_value:
                                print(f"DEBUG - Raw date value from Excel: {date_value}, type: {type(date_value)}")
                            
                            if date_value:
                                try:
                                    # Force date format with multiple formats support
                                    if isinstance(date_value, str):
                                        # Try different date formats
                                        for date_format in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
                                            try:
                                                entry_date = pd.to_datetime(date_value, format=date_format)
                                                break
                                            except:
                                                continue
                                        else:
                                            # If no format worked, let pandas infer
                                            entry_date = pd.to_datetime(date_value, errors="coerce")
                                    elif isinstance(date_value, (int, float)):
                                        # Handle Excel date numbers
                                        try:
                                            # Excel stores dates as numbers since 1900-01-01
                                            entry_date = pd.to_datetime('1899-12-30') + pd.Timedelta(days=int(date_value))
                                        except:
                                            entry_date = pd.to_datetime(date_value, errors="coerce")
                                    else:
                                        entry_date = pd.to_datetime(date_value)
                                    
                                    if pd.isna(entry_date):
                                        print(f"DEBUG - Could not convert date value: {date_value}")
                                        continue
                                        
                                    entry_date = entry_date.date()
                                    print(f"DEBUG - Converted date: {entry_date} from {date_value}")
                                    
                                    # DEBUG: Print Excel data being read
                                    print(f"DEBUG - Processing date: {entry_date}, Column: {col_idx}")
                                    print(f"DEBUG - Excel sheet rows available: {len(wb_data)}")
                                    
                                    # Extract data from exact Excel rows as specified in mapping (adjusted for new row structure):
                                    # Row 7: Date headers (already processed above)
                                    # Row 8: Planned Total Cost (for interval flows)
                                    # Row 9: Cumulative Budgeted Cost (for cumulative flows)
                                    # Row 10: Planned % Daily (for interval completion percentage)
                                    # Row 11: Cumulative % Daily (for cumulative completion percentage)
                                    # Row 12: Elapsed % 
                                    # Row 13: Elapsed Period
                                    # Row 14: Actual Cost
                                    
                                    # Helper function to safely extract and convert values
                                    def safe_extract_value(row_idx, col_idx, default=0):
                                        if len(wb_data) > row_idx:
                                            try:
                                                raw_value = wb_data.iloc[row_idx, col_idx]
                                                
                                                # Special debugging for resource rows 17-22
                                                if row_idx >= 16 and row_idx <= 21:  # rows 17-22 (0-indexed)
                                                    print(f"🔍 RESOURCE DEBUG - Row {row_idx+1} (R{row_idx+1}), Col {col_idx}: '{raw_value}' (type: {type(raw_value)})")
                                                    
                                                    # Check first few columns to see if data exists elsewhere
                                                    if col_idx > 5:  # Only check if we're not in early columns
                                                        for check_col in range(0, min(10, wb_data.shape[1])):
                                                            check_val = wb_data.iloc[row_idx, check_col] if wb_data.shape[1] > check_col else None
                                                            if not pd.isna(check_val) and check_val != '' and check_val is not None:
                                                                print(f"🔍 FOUND DATA - Row {row_idx+1}, Col {check_col}: '{check_val}' (type: {type(check_val)})")
                                                                break
                                                    
                                                print(f"DEBUG - Raw value at row {row_idx+1}, col {col_idx}: '{raw_value}' (type: {type(raw_value)})")
                                                
                                                if pd.isna(raw_value) or raw_value == '' or raw_value is None:
                                                    if row_idx >= 16 and row_idx <= 21:
                                                        print(f"🔍 RESOURCE DEBUG - Row {row_idx+1} is NaN/empty, returning default: {default}")
                                                    print(f"DEBUG - Value is NaN/empty, returning default: {default}")
                                                    return default
                                                
                                                # Handle percentage strings
                                                if isinstance(raw_value, str):
                                                    # Clean the string first
                                                    clean_value = str(raw_value).strip()
                                                    
                                                    # Remove thousands separators and handle Arabic/English numbers
                                                    clean_value = clean_value.replace(',', '').replace('٬', '')
                                                    
                                                    if '%' in clean_value:
                                                        try:
                                                            result = float(clean_value.replace('%', '').strip())
                                                            print(f"DEBUG - Converted percentage '{raw_value}' to {result}")
                                                            return result
                                                        except:
                                                            print(f"DEBUG - Failed to convert percentage '{raw_value}', returning {default}")
                                                            return default
                                                    
                                                    # Handle numeric strings
                                                    try:
                                                        result = float(clean_value)
                                                        print(f"DEBUG - Converted string '{raw_value}' to {result}")
                                                        return result
                                                    except:
                                                        print(f"DEBUG - Failed to convert string '{raw_value}', returning {default}")
                                                        return default
                                                
                                                # Handle datetime values (for date rows 17 and 20)
                                                if isinstance(raw_value, (datetime, pd.Timestamp)):
                                                    print(f"DEBUG - Found datetime value: {raw_value}")
                                                    return raw_value  # Return datetime as-is
                                                
                                                # Handle numeric values
                                                try:
                                                    result = float(raw_value)
                                                    print(f"DEBUG - Numeric value: {result}")
                                                    return result
                                                except:
                                                    print(f"DEBUG - Failed to convert '{raw_value}' to float, returning {default}")
                                                    return default
                                            except Exception as e:
                                                print(f"DEBUG - Error extracting value at row {row_idx+1}, col {col_idx}: {e}")
                                                return default
                                        else:
                                            if row_idx >= 16 and row_idx <= 21:
                                                print(f"🔍 RESOURCE DEBUG - Row {row_idx+1} NOT AVAILABLE! Only {len(wb_data)} rows in Excel data")
                                            print(f"DEBUG - Row {row_idx+1} not available in data (only {len(wb_data)} rows)")
                                            return default
                                    
                                    planned_total_cost = safe_extract_value(7, col_idx)  # Row 8
                                    cumulative_budgeted_cost = safe_extract_value(8, col_idx)  # Row 9
                                    planned_daily_percent = safe_extract_value(9, col_idx)  # Row 10
                                    cumulative_daily_percent = safe_extract_value(10, col_idx)  # Row 11
                                    elapsed_percent = safe_extract_value(11, col_idx)  # Row 12
                                    elapsed_period = safe_extract_value(12, col_idx)  # Row 13
                                    actual_cost = safe_extract_value(13, col_idx)  # Row 14
                                    
                                    # Smart resource data extraction with header validation and multiple column search
                                    def validate_header_alignment():
                                        """Check if headers in rows 17-22 match expected values"""
                                        expected_headers = {
                                            16: ["Date", "تاريخ"],  # Row 17
                                            17: ["Budgeted Labor Units Weekly", "العمالة الأسبوعية"],  # Row 18
                                            18: ["Budgeted Nonlabor Units Weekly", "المعدات الأسبوعية"],  # Row 19
                                            19: ["Date", "تاريخ"],  # Row 20
                                            20: ["Budgeted Labor Units Monthly", "العمالة الشهرية"],  # Row 21
                                            21: ["Budgeted Nonlabor Units Monthly", "المعدات الشهرية"]   # Row 22
                                        }
                                        
                                        header_issues = []
                                        for row_idx, expected_values in expected_headers.items():
                                            if len(wb_data) > row_idx:
                                                actual_header = str(wb_data.iloc[row_idx, 0]).strip()
                                                if not any(expected in actual_header for expected in expected_values):
                                                    header_issues.append(f"Row {row_idx+1}: Expected {expected_values}, got '{actual_header}'")
                                        
                                        if header_issues:
                                            print(f"⚠️ HEADER ALIGNMENT ISSUES: {'; '.join(header_issues)}")
                                        return len(header_issues) == 0
                                    
                                    def find_resource_data_smart(row_idx, data_type="numeric"):
                                        """Smart search for resource data starting from column B with validation"""
                                        # Start from column B (index 1) and search through available columns
                                        start_col = 1  # Column B
                                        max_search_cols = min(wb_data.shape[1], col_idx + 20)  # Extended search range
                                        
                                        for check_col in range(start_col, max_search_cols):
                                            if len(wb_data) > row_idx and wb_data.shape[1] > check_col:
                                                val = safe_extract_value(row_idx, check_col)
                                                
                                                # For date fields, check if it's a valid date
                                                if data_type == "date":
                                                    # Handle datetime objects directly
                                                    if isinstance(val, (datetime, pd.Timestamp)):
                                                        print(f"🔍 FOUND DATE DATA - Row {row_idx+1}, Col {check_col}: {val} (datetime object)")
                                                        return val
                                                    # Handle Excel date serials
                                                    elif isinstance(val, (int, float)) and val > 40000:  # Excel date serial
                                                        print(f"🔍 FOUND DATE DATA - Row {row_idx+1}, Col {check_col}: {val} (Excel serial)")
                                                        return val
                                                    # Try to parse string dates
                                                    elif val and val != 0:
                                                        try:
                                                            parsed_date = pd.to_datetime(val, errors='coerce')
                                                            if parsed_date is not pd.NaT:
                                                                print(f"🔍 FOUND DATE DATA - Row {row_idx+1}, Col {check_col}: {val} (parsed)")
                                                                return parsed_date
                                                        except:
                                                            continue
                                                
                                                # For numeric fields, check if it's a valid number > 0
                                                elif data_type == "numeric" and val != 0 and not pd.isna(val):
                                                    try:
                                                        numeric_val = float(val)
                                                        if numeric_val > 0:  # Only positive values
                                                            print(f"🔍 FOUND RESOURCE DATA - Row {row_idx+1}, Col {check_col}: {numeric_val}")
                                                            return numeric_val
                                                    except:
                                                        continue
                                        
                                        # No valid data found - return None instead of 0 to indicate missing data
                                        print(f"❌ NO DATA FOUND - Row {row_idx+1} ({data_type}) in columns {start_col}-{max_search_cols-1}")
                                        return None
                                    
                                    # Validate header alignment first
                                    headers_valid = validate_header_alignment()
                                    if not headers_valid:
                                        print("⚠️ HEADER VALIDATION FAILED - Resource data may not be reliable")
                                    
                                    # Extract resource data with smart search (NO DEFAULT VALUES)
                                    weekly_date = find_resource_data_smart(16, "date")         # Row 17
                                    weekly_manpower = find_resource_data_smart(17, "numeric")  # Row 18
                                    weekly_equipment = find_resource_data_smart(18, "numeric") # Row 19
                                    monthly_date = find_resource_data_smart(19, "date")        # Row 20
                                    monthly_manpower = find_resource_data_smart(20, "numeric") # Row 21
                                    monthly_equipment = find_resource_data_smart(21, "numeric")# Row 22
                                    
                                    # Convert None to 0 only for storage, but keep track of missing data
                                    # Special handling for dates - convert datetime to string for storage
                                    if isinstance(weekly_date, (datetime, pd.Timestamp)):
                                        weekly_date = weekly_date.strftime('%Y-%m-%d')
                                    elif weekly_date is None:
                                        weekly_date = 0
                                    
                                    if isinstance(monthly_date, (datetime, pd.Timestamp)):
                                        monthly_date = monthly_date.strftime('%Y-%m-%d')
                                    elif monthly_date is None:
                                        monthly_date = 0
                                    
                                    weekly_manpower = weekly_manpower if weekly_manpower is not None else 0
                                    weekly_equipment = weekly_equipment if weekly_equipment is not None else 0
                                    monthly_manpower = monthly_manpower if monthly_manpower is not None else 0
                                    monthly_equipment = monthly_equipment if monthly_equipment is not None else 0
                                    
                                    # DEBUG: Print extracted resource values
                                    print(f"DEBUG - Resource extraction for date {entry_date}:")
                                    print(f"  R17 (weekly_date): {weekly_date}")
                                    print(f"  R18 (weekly_manpower): {weekly_manpower}")  
                                    print(f"  R19 (weekly_equipment): {weekly_equipment}")
                                    print(f"  R20 (monthly_date): {monthly_date}")
                                    print(f"  R21 (monthly_manpower): {monthly_manpower}")
                                    print(f"  R22 (monthly_equipment): {monthly_equipment}")
                                    
                                    
                                    # DEBUG: Print extracted values with better formatting
                                    print(f"DEBUG - Extracted values for {entry_date}:")
                                    print(f"  Row 8 (Planned Total): {planned_total_cost} (type: {type(planned_total_cost)})")
                                    print(f"  Row 9 (Cumulative Budgeted): {cumulative_budgeted_cost} (type: {type(cumulative_budgeted_cost)})")
                                    print(f"  Row 9 (Daily %): {planned_daily_percent}")
                                    print(f"  Row 10 (Cum %): {cumulative_daily_percent}")
                                    print(f"  Row 13 (Actual): {actual_cost}")
                                    
                                    # Check if any values are non-zero
                                    has_data = any([
                                        planned_total_cost and planned_total_cost != 0,
                                        cumulative_budgeted_cost and cumulative_budgeted_cost != 0,
                                        actual_cost and actual_cost != 0
                                    ])
                                    print(f"DEBUG - Has non-zero data: {has_data}")
                                    
                                    # Clean percentage values
                                    if isinstance(planned_daily_percent, str) and '%' in str(planned_daily_percent):
                                        planned_daily_percent = float(str(planned_daily_percent).replace('%', ''))
                                    if isinstance(cumulative_daily_percent, str) and '%' in str(cumulative_daily_percent):
                                        cumulative_daily_percent = float(str(cumulative_daily_percent).replace('%', ''))
                                    if isinstance(elapsed_percent, str) and '%' in str(elapsed_percent):
                                        elapsed_percent = float(str(elapsed_percent).replace('%', ''))
                                    
                                    # Create progress data based on exact Excel mapping specification
                                    # Store data according to cumulative vs interval mapping rules
                                    # IMPORTANT: Database field mapping:
                                    # - planned_cost: Row 8 (Cumulative Budgeted Cost) for cumulative calculations
                                    # - actual_cost: Row 7 (Planned Total Cost) for interval calculations  
                                    # - Row 13 (Actual Cost) stored in notes for future use
                                    
                                    progress_data = {
                                        'project_name': current_project_name,
                                        'entry_date': entry_date,
                                        'planned_completion': float(cumulative_daily_percent),  # Row 10: Cumulative % Daily
                                        'planned_cost': float(cumulative_budgeted_cost),  # Row 8: Cumulative Budgeted Cost (for cumulative flows)
                                        'actual_completion': float(elapsed_percent),  # Row 11: Elapsed %
                                        'actual_cost': float(planned_total_cost),  # Row 7: Planned Total Cost (for interval flows)
                                        'notes': f'R7:{planned_total_cost}|R8:{cumulative_budgeted_cost}|R9:{planned_daily_percent}|R10:{cumulative_daily_percent}|R11:{elapsed_percent}|R12:{elapsed_period}|R13:{actual_cost}|R17:{weekly_date}|R18:{weekly_manpower}|R19:{weekly_equipment}|R20:{monthly_date}|R21:{monthly_manpower}|R22:{monthly_equipment}'
                                    }
                                    
                                    
                                    # Only skip if ALL values are exactly 0 or None
                                    # Allow saving even if some values are 0 (for testing and debugging)
                                    all_zero = (
                                        (planned_total_cost == 0 or planned_total_cost is None) and
                                        (cumulative_budgeted_cost == 0 or cumulative_budgeted_cost is None) and
                                        (actual_cost == 0 or actual_cost is None) and
                                        (planned_daily_percent == 0 or planned_daily_percent is None) and
                                        (cumulative_daily_percent == 0 or cumulative_daily_percent is None)
                                    )
                                    
                                    if all_zero:
                                        print(f"DEBUG - Skipping {entry_date} - all values are 0 or None, not saving empty data")
                                        continue  # Skip this entry entirely instead of saving with fake values
                                    else:
                                        print(f"DEBUG - Saving data for {entry_date} - has non-zero values")
                                    
                                    # DEBUG: Print final progress data being saved
                                    print(f"DEBUG - Saving progress data: {progress_data}")
                                    
                                    result = self.data_manager.add_progress_data(progress_data)
                                    print(f"DEBUG - Save result: {result}")
                                    
                                except Exception as e:
                                    continue  # Skip invalid date entries
                
            # Final validation and return
            if success_count == 0 and not error_details:
                error_details.append('لم يتم العثور على مشاريع صحيحة للاستيراد في الملف')
                error_details.append('تأكد من وجود أسماء المشاريع في الخلية B3 من كل ورقة عمل')
                error_details.append('تأكد من أن الملف يحتوي على البيانات المطلوبة في المواضع الصحيحة')
            
            result = {
                'success': success_count > 0,
                'imported_count': success_count,
                'updated_count': 0,  # Always 0 since we clear all data first
                'projects': imported_projects,
                'error_details': error_details,
                'warnings': warnings
            }
            
            if success_count > 0:
                result['message'] = f'تم استيراد {success_count} مشروع بنجاح (تم استبدال جميع المشاريع السابقة)'
                
                # Save the original Excel file after successful import
                if file_content and file_hash:
                    try:
                        save_result = self.data_manager.save_original_excel_file(
                            uploaded_file.name, file_content, imported_projects, file_hash
                        )
                        if save_result:
                            print("DEBUG - Successfully saved original Excel file")
                        else:
                            print("DEBUG - Failed to save original Excel file")
                            warnings.append('تم الاستيراد بنجاح لكن لم يتم حفظ الملف الأصلي')
                    except Exception as e:
                        print(f"DEBUG - Error saving original file: {e}")
                        warnings.append(f'تم الاستيراد بنجاح لكن فشل في حفظ الملف الأصلي: {str(e)}')
            else:
                result['message'] = 'فشل في استيراد المشاريع - راجع التفاصيل أدناه'
            
            # Update warnings in result
            result['warnings'] = warnings
            return result
            
        except Exception as e:
            error_details.append(f'خطأ عام في معالجة الملف: {str(e)}')
            return {
                'success': False,
                'imported_count': 0,
                'updated_count': 0,
                'message': 'خطأ غير متوقع في معالجة الملف',
                'error_details': error_details,
                'warnings': warnings
            }

    def export_project_detailed_report(self, project_name: str) -> Optional[bytes]:
        """Export detailed project report to Excel"""
        try:
            wb = Workbook()
            
            # Project info sheet
            ws_info = wb.active
            ws_info.title = "معلومات المشروع"
            
            project_info = self.data_manager.get_project_info(project_name)
            if not project_info:
                return None
            
            # Formatting
            arabic_alignment = Alignment(horizontal='right', vertical='center')
            header_font = Font(bold=True, size=12)
            title_font = Font(bold=True, size=14)
            
            # Add project information
            info_data = [
                ['اسم المشروع', project_info['project_name']],
                ['الشركة المنفذة', project_info['executing_company']],
                ['الشركة الاستشارية', project_info['consulting_company']],
                ['تاريخ البدء', project_info['start_date']],
                ['تاريخ الانتهاء', project_info['end_date']],
                ['الميزانية الإجمالية', project_info['total_budget']],
                ['موقع المشروع', project_info['project_location']],
                ['نوع المشروع', project_info['project_type']]
            ]
            
            for idx, (label, value) in enumerate(info_data, 1):
                ws_info.cell(row=idx, column=1, value=label).font = header_font
                ws_info.cell(row=idx, column=1).alignment = arabic_alignment
                ws_info.cell(row=idx, column=2, value=value).alignment = arabic_alignment
                
                if 'الميزانية' in label and isinstance(value, (int, float)):
                    ws_info.cell(row=idx, column=2).number_format = '#,##0.00'
            
            # Progress data sheet
            progress_data = self.data_manager.get_progress_data(project_name)
            if not progress_data.empty:
                ws_progress = wb.create_sheet("بيانات التقدم")
                
                # Add headers
                progress_headers = ['تاريخ الإدخال', 'نسبة الإنجاز المخطط (%)', 'التكلفة المخططة',
                                  'نسبة الإنجاز الفعلي (%)', 'التكلفة الفعلية', 'ملاحظات']
                
                for col, header in enumerate(progress_headers, 1):
                    cell = ws_progress.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = arabic_alignment
                
                # Add progress data
                for row_idx, (_, row) in enumerate(progress_data.iterrows(), 2):
                    ws_progress.cell(row=row_idx, column=1, value=row['entry_date']).alignment = arabic_alignment
                    ws_progress.cell(row=row_idx, column=2, value=row['planned_completion']).alignment = arabic_alignment
                    ws_progress.cell(row=row_idx, column=3, value=row['planned_cost']).number_format = '#,##0.00'
                    ws_progress.cell(row=row_idx, column=4, value=row['actual_completion']).alignment = arabic_alignment
                    ws_progress.cell(row=row_idx, column=5, value=row['actual_cost']).number_format = '#,##0.00'
                    ws_progress.cell(row=row_idx, column=6, value=row['notes']).alignment = arabic_alignment
            
            # KPI sheet
            project_kpi = self.evm_calculator.calculate_project_kpi(project_name)
            if project_kpi:
                ws_kpi = wb.create_sheet("مؤشرات الأداء")
                
                kpi_data = [
                    ['القيمة المخططة (PV)', project_kpi['pv']],
                    ['القيمة المكتسبة (EV)', project_kpi['ev']],
                    ['التكلفة الفعلية (AC)', project_kpi['ac']],
                    ['مؤشر أداء التكلفة (CPI)', project_kpi['cpi']],
                    ['مؤشر أداء الجدولة (SPI)', project_kpi['spi']],
                    ['انحراف التكلفة (CV)', project_kpi['cv']],
                    ['انحراف الجدولة (SV)', project_kpi['sv']],
                    ['حالة المشروع', project_kpi['status']]
                ]
                
                for idx, (label, value) in enumerate(kpi_data, 1):
                    ws_kpi.cell(row=idx, column=1, value=label).font = header_font
                    ws_kpi.cell(row=idx, column=1).alignment = arabic_alignment
                    
                    if isinstance(value, (int, float)):
                        if 'مؤشر' in label:
                            ws_kpi.cell(row=idx, column=2, value=f"{value:.3f}")
                        else:
                            ws_kpi.cell(row=idx, column=2, value=value)
                            ws_kpi.cell(row=idx, column=2).number_format = '#,##0.00'
                    else:
                        ws_kpi.cell(row=idx, column=2, value=value)
                    
                    ws_kpi.cell(row=idx, column=2).alignment = arabic_alignment
            
            # Auto-adjust column widths
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
        except Exception as e:
            print(f"Error exporting detailed project report: {e}")
            return None
    
    def generate_2000_column_template(self, start_date: date, flow_type: str = "Daily") -> Optional[bytes]:
        """Generate Excel template with optimized columns for extensive time tracking"""
        try:
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Project Financial Data"
            
            # Header styling
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Generate date columns based on flow type
            # Smart calculation: use actual project date range if available, otherwise default
            all_projects = self.data_manager.get_all_projects()
            if all_projects:
                # Calculate optimal column count based on project dates
                project_dates = []
                for proj in all_projects:
                    if proj.get('start_date'):
                        project_dates.append(proj['start_date'])
                    if proj.get('end_date'):
                        project_dates.append(proj['end_date'])
                
                if project_dates:
                    min_date = min(project_dates)
                    max_date = max(project_dates)
                    # Calculate smart column count based on actual project duration
                    if flow_type == "Daily":
                        days_needed = (max_date - min_date).days + 90  # Add 90 days buffer
                        smart_columns = min(max(days_needed, 30), self.max_columns - 10)  # Minimum 30 days
                    elif flow_type == "Monthly":
                        months_needed = ((max_date.year - min_date.year) * 12 + max_date.month - min_date.month) + 6  # 6 month buffer
                        smart_columns = min(max(months_needed, 12), self.max_columns - 10)  # Minimum 1 year
                    else:
                        smart_columns = min(24, self.max_columns - 10)  # 2 years for yearly
                else:
                    smart_columns = min(365, self.max_columns - 10)  # Default to 1 year
            else:
                smart_columns = min(365, self.max_columns - 10)  # Default to 1 year
            
            date_columns = self._generate_date_columns_for_excel(start_date, flow_type, smart_columns)
            
            # Column headers matching Primavera-style format
            basic_headers = [
                "Project ID", "Project Name", "Parent Category", "Start Date", "Finish Date", 
                "BL Project Total Cost", "Executing Company", "Location", "Type", "Description"
            ]
            
            all_headers = basic_headers + date_columns
            
            # Add headers to first row - optimized bulk operation
            headers_to_use = all_headers[:self.max_columns]
            for col, header in enumerate(headers_to_use, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Apply formatting to header range in bulk
            header_range = f"A1:{get_column_letter(len(headers_to_use))}1"
            for cell in ws[header_range][0]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add cumulative row headers
            ws.cell(row=2, column=1, value="CUMULATIVE")
            ws.cell(row=2, column=1).font = Font(bold=True, color="FF0000")
            
            ws.cell(row=3, column=1, value="INTERVAL")
            ws.cell(row=3, column=1).font = Font(bold=True, color="0000FF")
            
            # Get projects and populate data - limit to first 20 for performance
            all_projects = self.data_manager.get_all_projects()
            projects_to_process = all_projects[:20]  # Limit to 20 projects for better performance
            current_row = 4
            
            for project in projects_to_process:
                # Basic project information using Primavera-style format
                project_data = [
                    project.get('project_id', f"P{project.get('id', '001'):03d}"),
                    project['project_name'],
                    project.get('parent_category_name', 'Uncategorized'),
                    project.get('start_date', ''),
                    project.get('end_date', ''),  # Maps to "Finish Date" in header
                    project.get('total_budget', 0),  # Maps to "BL Project Total Cost" in header
                    project.get('executing_company', ''),
                    project.get('project_location', ''),
                    project.get('project_type', ''),
                    project.get('project_description', '')
                ]
                
                # Add basic project data
                for col, value in enumerate(project_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                
                # Get progress data for this project
                progress_data = self.data_manager.get_progress_data(project['project_name'])
                
                # Populate financial data for each date column
                for col_idx, date_col in enumerate(date_columns, len(basic_headers) + 1):
                    if col_idx <= self.max_columns:
                        # Calculate cumulative and interval values
                        cumulative_value = self._get_cumulative_value(progress_data, date_col, flow_type)
                        interval_value = self._get_interval_value(progress_data, date_col, flow_type)
                        
                        # Add to cumulative row
                        ws.cell(row=2, column=col_idx, value=cumulative_value)
                        
                        # Add to interval row  
                        ws.cell(row=3, column=col_idx, value=interval_value)
                        
                        # Add project specific data
                        project_value = self._get_project_value_for_date(progress_data, date_col, flow_type)
                        ws.cell(row=current_row, column=col_idx, value=project_value)
                
                current_row += 1
            
            # Format columns
            for col in range(1, min(self.max_columns + 1, len(all_headers) + 1)):
                column_letter = ws.cell(row=1, column=col).column_letter
                if col > len(basic_headers):  # Financial data columns
                    ws.column_dimensions[column_letter].width = 12
                else:  # Basic info columns
                    ws.column_dimensions[column_letter].width = 15
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            print(f"Error generating 2000-column template: {e}")
            return None
    
    def _generate_date_columns_for_excel(self, start_date: date, flow_type: str, max_cols: int) -> List[str]:
        """Generate date columns for Excel template - optimized version"""
        # Use pandas for efficient date generation
        import pandas as pd
        
        if flow_type == "Daily":
            dates = pd.date_range(start=start_date, periods=max_cols, freq='D')
            return [d.strftime('%Y-%m-%d') for d in dates]
        elif flow_type == "Monthly":
            dates = pd.date_range(start=start_date, periods=max_cols, freq='MS')
            return [d.strftime('%Y-%m') for d in dates]
        elif flow_type == "Yearly":
            dates = pd.date_range(start=start_date, periods=max_cols, freq='YS')
            return [str(d.year) for d in dates]
        
        return []
    
    def _get_cumulative_value(self, progress_data: pd.DataFrame, date_col: str, flow_type: str) -> float:
        """Get cumulative financial value up to a specific date"""
        if progress_data.empty:
            return 0.0
        
        try:
            target_date = self._parse_date_column(date_col, flow_type)
            progress_data_copy = progress_data.copy()
            progress_data_copy['entry_date'] = pd.to_datetime(progress_data_copy['entry_date'])
            
            filtered_data = progress_data_copy[progress_data_copy['entry_date'] <= target_date]
            return filtered_data['actual_cost'].sum() if not filtered_data.empty else 0.0
        except:
            return 0.0
    
    def _get_interval_value(self, progress_data: pd.DataFrame, date_col: str, flow_type: str) -> float:
        """Get interval financial value for a specific period"""
        if progress_data.empty:
            return 0.0
        
        try:
            target_date = self._parse_date_column(date_col, flow_type)
            progress_data_copy = progress_data.copy()
            progress_data_copy['entry_date'] = pd.to_datetime(progress_data_copy['entry_date'])
            
            if flow_type == "Daily":
                filtered_data = progress_data_copy[progress_data_copy['entry_date'].dt.date == target_date.date()]
            elif flow_type == "Monthly":
                filtered_data = progress_data_copy[
                    (progress_data_copy['entry_date'].dt.year == target_date.year) &
                    (progress_data_copy['entry_date'].dt.month == target_date.month)
                ]
            elif flow_type == "Yearly":
                filtered_data = progress_data_copy[progress_data_copy['entry_date'].dt.year == target_date.year]
            
            return filtered_data['actual_cost'].sum() if not filtered_data.empty else 0.0
        except:
            return 0.0
    
    def _get_project_value_for_date(self, progress_data: pd.DataFrame, date_col: str, flow_type: str) -> float:
        """Get project-specific value for a date (using interval by default)"""
        return self._get_interval_value(progress_data, date_col, flow_type)
    
    def _parse_date_column(self, date_col: str, flow_type: str) -> pd.Timestamp:
        """Parse date column string to datetime object"""
        if flow_type == "Daily":
            return pd.to_datetime(date_col)
        elif flow_type == "Monthly":
            return pd.to_datetime(date_col + '-01')
        elif flow_type == "Yearly":
            return pd.to_datetime(f"{date_col}-01-01")
        else:
            return pd.to_datetime(date_col)
    
    def import_from_2000_column_excel(self, excel_file) -> bool:
        """Import data from 2000-column Excel format"""
        try:
            # Read Excel file
            df = pd.read_excel(excel_file, sheet_name=0)
            
            # DEBUG: Print dataframe info after reading
            print(f"DEBUG - Excel file imported successfully")
            print(f"DEBUG - DataFrame shape: {df.shape}")
            print(f"DEBUG - Column names: {list(df.columns)}")
            print(f"DEBUG - First 10 rows of imported data:")
            print(df.head(10))
            print(f"DEBUG - Data types:")
            print(df.dtypes)
            
            # Print specific rows that contain our data
            print(f"DEBUG - Row 1 (Project Names): {df.iloc[0].tolist() if len(df) > 0 else 'N/A'}")
            print(f"DEBUG - Row 6 (Dates): {df.iloc[5].tolist()[:10] if len(df) > 5 else 'N/A'}")
            print(f"DEBUG - Row 7 (Planned Total): {df.iloc[6].tolist()[:10] if len(df) > 6 else 'N/A'}")
            print(f"DEBUG - Row 8 (Cumulative): {df.iloc[7].tolist()[:10] if len(df) > 7 else 'N/A'}")
            print(f"DEBUG - Row 13 (Actual Cost): {df.iloc[12].tolist()[:10] if len(df) > 12 else 'N/A'}")
            
            # Extract basic project information with exact Excel column mapping as specified
            basic_columns = [
                "المشروع", "امر الشراء", "المقاول", "Start Date", "Finish Date", 
                "BL Project Total Cost", "Location", "Type", "Description"
            ]
            
            # Column mapping for exact Excel sheet columns as specified
            column_mapping = {
                'Project Name': ['المشروع', 'Project Name', 'project_name', 'Name'],
                'Project ID': ['امر الشراء', 'Project ID', 'project_id', 'ID'],  
                'Executing Company': ['المقاول', 'Executing Company', 'executing_company', 'Company'],
                'Start Date': ['Start Date', 'start_date', 'Start'],
                'Finish Date': ['Finish Date', 'End Date', 'end_date', 'Finish', 'End'],
                'BL Project Total Cost': ['BL Project Total Cost', 'Total Budget', 'total_budget', 'Budget', 'Total Cost'],
                'Location': ['Location', 'project_location', 'Site'],
                'Type': ['Type', 'project_type', 'Project Type'],
                'Description': ['Description', 'project_description', 'Notes']
            }
            
            # Get parent categories for mapping
            categories = self.data_manager.get_parent_categories()
            category_map = {cat['category_name']: cat['id'] for cat in categories}
            
            # Create reverse mapping for finding columns in the Excel file
            def find_column_value(row, target_column):
                """Find value from row using column mapping"""
                if target_column in column_mapping:
                    for possible_name in column_mapping[target_column]:
                        if possible_name in row and pd.notna(row[possible_name]):
                            return row[possible_name]
                return None
            
            for _, row in df.iterrows():
                project_name = find_column_value(row, 'Project Name')
                if project_name:
                    # Prepare project data using mapped columns
                    project_data = {
                        'project_name': project_name,
                        'project_id': find_column_value(row, 'Project ID') or '',
                        'parent_category_id': category_map.get(find_column_value(row, 'Parent Category'), None),
                        'executing_company': find_column_value(row, 'Executing Company') or '',
                        'consulting_company': '',  # Not in template
                        'start_date': find_column_value(row, 'Start Date') or '',
                        'end_date': find_column_value(row, 'Finish Date') or '',
                        'total_budget': find_column_value(row, 'BL Project Total Cost') or 0,
                        'project_location': find_column_value(row, 'Location') or '',
                        'project_type': find_column_value(row, 'Type') or '',
                        'project_description': find_column_value(row, 'Description') or '',
                        'display_order': 0,
                        'created_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Add or update project
                    self.data_manager.add_project(project_data)
                    
                    # Import financial data from date columns (exclude basic project info columns)
                    all_mapped_columns = []
                    for basic_col in basic_columns:
                        all_mapped_columns.extend(column_mapping.get(basic_col, [basic_col]))
                    
                    date_columns = [col for col in df.columns if col not in all_mapped_columns]
                    
                    for date_col in date_columns:
                        financial_value = row.get(date_col, 0)
                        
                        if pd.notna(financial_value) and financial_value > 0:
                            # Parse date and add progress data
                            try:
                                entry_date = self._parse_date_column(date_col, "Daily")
                                
                                # For simple Excel import, use the financial value as interval data
                                progress_data = {
                                    'project_name': project_data['project_name'],
                                    'entry_date': entry_date.strftime('%Y-%m-%d'),
                                    'planned_completion': 0,
                                    'planned_cost': 0,  # Set to 0 for cumulative flows (can be updated later)
                                    'actual_completion': 0,
                                    'actual_cost': float(financial_value),  # Use for interval flows
                                    'notes': ''
                                }
                                
                                self.data_manager.add_progress_data(progress_data)
                            except Exception as e:
                                print(f"Error importing date {date_col}: {e}")
            
            return True
            
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            return False
