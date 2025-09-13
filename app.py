import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import os
from data_manager import DataManager
from evm_calculator import EVMCalculator
from excel_exporter import ExcelExporter
from visualizations import create_s_curve, create_kpi_dashboard
from utils import format_currency, validate_date_range

# Configure page
st.set_page_config(
    page_title="Project Management of Abdullah Al-Saeed Engineering Consulting Company",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state
if 'data_manager' not in st.session_state:
    st.session_state.data_manager = DataManager()
if 'selected_project' not in st.session_state:
    st.session_state.selected_project = None

# CSS for Primavera-style interface
st.markdown("""
<style>
.rtl {
    direction: rtl;
    text-align: right;
}
.arabic-text {
    font-family: 'Arial', 'Tahoma', sans-serif;
    direction: rtl;
}

/* Primavera-style Table Styling */
.primavera-table {
    font-size: 12px !important;
    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
}

.primavera-table .stDataFrame {
    border: 1px solid #d0d0d0;
}

.primavera-table th {
    background-color: #e6f3ff !important;
    color: #333 !important;
    font-weight: bold !important;
    font-size: 11px !important;
    padding: 4px 8px !important;
    border: 1px solid #ccc !important;
}

.primavera-table td {
    font-size: 11px !important;
    padding: 3px 6px !important;
    border: 1px solid #e0e0e0 !important;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    max-width: 150px;
}

.project-row {
    background-color: #fafafa;
    border: 1px solid #e0e0e0;
    margin: 1px 0;
    padding: 4px;
    font-size: 11px;
}

.project-row-blue {
    background-color: #cce5ff;
    border: 1px solid #b8daff;
    margin: 1px 0;
    padding: 6px;
    font-size: 11px;
}

.project-row-green {
    background-color: #d4edda;
    border: 1px solid #c3e6cb;
    margin: 1px 0;
    padding: 6px;
    font-size: 11px;
}

.project-header-row {
    background-color: #ffffff;
    font-weight: bold;
    padding: 8px;
    border: 2px solid #333;
    margin: 2px 0;
    font-size: 12px;
}

.parent-category-header {
    background: linear-gradient(90deg, #4a90e2, #5ba0f2);
    color: white;
    font-weight: bold;
    padding: 8px;
    border-radius: 4px;
    margin: 5px 0;
    font-size: 12px;
}

.category-totals {
    background-color: #f0f8ff;
    border: 2px solid #4a90e2;
    padding: 6px;
    border-radius: 4px;
    font-weight: bold;
    font-size: 11px;
    margin: 2px 0;
}

.move-arrow {
    color: #666;
    cursor: pointer;
    font-size: 14px;
    margin: 0 2px;
}

.move-arrow:hover {
    color: #4a90e2;
}

.compact-button {
    padding: 2px 6px !important;
    font-size: 10px !important;
    height: 24px !important;
    min-height: 24px !important;
}

/* Professional Dashboard Theme */
.main-dashboard {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    padding: 20px;
    border-radius: 10px;
    margin: 10px 0;
}

.kpi-card {
    background: white;
    padding: 20px;
    border-radius: 10px;
    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    border-left: 4px solid #3498db;
    margin: 10px 0;
}

.status-good {
    color: #27ae60;
    font-weight: bold;
}

.status-warning {
    color: #f39c12;
    font-weight: bold;
}

.status-danger {
    color: #e74c3c;
    font-weight: bold;
}

.dashboard-header {
    background: linear-gradient(90deg, #2c3e50, #3498db);
    color: white;
    padding: 15px;
    border-radius: 8px;
    margin: 15px 0;
}

.chart-container {
    background: white;
    padding: 20px;
    border-radius: 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    margin: 10px 0;
}

/* Sidebar styling */
.css-1d391kg {
    background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%);
}

/* Metrics styling */
div[data-testid="metric-container"] {
    background: white;
    border: 1px solid #e0e0e0;
    padding: 1rem;
    border-radius: 8px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
}

/* Tab styling */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
}

.stTabs [data-baseweb="tab"] {
    background: #f8f9fa;
    border-radius: 8px;
    padding: 12px 24px;
    border: 1px solid #dee2e6;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(90deg, #3498db, #2980b9) !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)

def main():
    st.title("Project Management System")
    
    # Sidebar for navigation
    with st.sidebar:
        st.markdown('<div class="english"><h2>Main Menu</h2></div>', unsafe_allow_html=True)
        
        # Project selection with multi-select option
        projects = st.session_state.data_manager.get_all_projects()
        if projects:
            # Format project names with ID, dates, and cost
            project_display_names = []
            project_name_mapping = {}
            
            for p in projects:
                project_id = p.get('project_id', f"P{p.get('id', '001'):03d}")
                start_date = p.get('start_date', 'Not Set')
                end_date = p.get('end_date', 'Not Set')
                total_cost = p.get('total_budget', 0)
                
                display_name = f"ID: {project_id} | {p['project_name']} | {start_date} to {end_date} | ${total_cost:,.2f}"
                project_display_names.append(display_name)
                project_name_mapping[display_name] = p['project_name']
            
            project_names = [p['project_name'] for p in projects]
            
            # Multi-select projects interface
            st.markdown("### 🎯 اختيار المشاريع")
            
            # Select all checkbox
            select_all = st.checkbox("اختيار كل المشاريع", key="select_all_projects")
            
            if select_all:
                # Select all projects
                st.session_state.selected_projects = project_names
                st.session_state.selected_project = "All Projects"
                selected_projects_display = project_display_names
                st.info(f"تم اختيار جميع المشاريع ({len(project_names)} مشروع)")
            else:
                # Multi-select dropdown
                current_selection = st.session_state.get('multi_project_selector', [])
                selected_projects_display = st.multiselect(
                    "اختر المشاريع:",
                    project_display_names,
                    default=current_selection,
                    key="multi_project_selector"
                )
                
                if selected_projects_display:
                    selected_projects = [project_name_mapping[display] for display in selected_projects_display]
                    st.session_state.selected_projects = selected_projects
                    st.session_state.selected_project = selected_projects[0] if selected_projects else None
                    st.success(f"تم اختيار {len(selected_projects)} مشروع")
                else:
                    st.session_state.selected_projects = []
                    st.session_state.selected_project = None
                
        # Initialize selected_projects if not exists
        if 'selected_projects' not in st.session_state:
            st.session_state.selected_projects = []
        
        # Navigation tabs with professional dashboard design
        st.markdown("---")
        st.markdown('<div class="english"><h3>🏢 Control Panel</h3></div>', unsafe_allow_html=True)
        
        tab_choice = st.radio(
            "Sections",
            [
                "📝 Projects Management",
                "💰 Monthly Financial Data",
                "📈 Project Progress", 
                "👥 Resources",
                "📊 Charts",
                "📋 Reports",
                "🎥 Presentations",
                "⚙️ Settings"
            ],
            key="nav_radio",
            format_func=lambda x: x
        )
    
    # Main content area with new dashboard structure
    if tab_choice == "📝 Projects Management":
        data_entry_tab()
    elif tab_choice == "💰 Monthly Financial Data":
        financials_tab()
    elif tab_choice == "📈 Project Progress":
        progress_tab()
    elif tab_choice == "👥 Resources":
        resources_tab()
    elif tab_choice == "📊 Charts":
        charts_tab()
    elif tab_choice == "📋 Reports":
        reports_tab()
    elif tab_choice == "🎥 Presentations":
        powerpoint_tab()
    elif tab_choice == "⚙️ Settings":
        settings_tab()

def data_entry_tab():
    st.markdown('<div class="english"><h2>Projects Management</h2></div>', unsafe_allow_html=True)
    
    # Get all projects
    all_projects = st.session_state.data_manager.get_all_projects()
    
    if not all_projects:
        st.info("No projects loaded. Please import projects from Excel file in Settings or add a new project.")
        
        # Option to add new project
        if st.button("➕ Add New Project", use_container_width=True):
            st.session_state.show_new_project_form = True
        
        # Excel Template Management
        st.markdown("### Excel Template Management")
        excel_cols = st.columns(3)
        
        with excel_cols[0]:
            if st.button("📊 Generate 2000-Column Excel Template", use_container_width=True):
                excel_exporter = ExcelExporter(st.session_state.data_manager)
                template_data = excel_exporter.generate_2000_column_template(
                    start_date=date.today(),
                    flow_type="Daily"
                )
                
                if template_data:
                    st.download_button(
                        label="Download Excel Template",
                        data=template_data,
                        file_name=f"project_template_2000_columns_{date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success("2000-column Excel template generated successfully!")
                else:
                    st.error("Failed to generate Excel template")
        
        with excel_cols[1]:
            uploaded_file = st.file_uploader(
                "Import from 2000-Column Excel",
                type=['xlsx', 'xls'],
                help="Upload Excel file with 2000-column format"
            )
            
            if uploaded_file is not None:
                # Check MIME type and file extension
                if uploaded_file.type not in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                                              'application/vnd.ms-excel',
                                              'application/octet-stream']:  # Some browsers use this for xlsx
                    st.error(f"❌ ملف غير صالح! نوع الملف: {uploaded_file.type}")
                    st.info("يرجى رفع ملف Excel بصيغة .xlsx أو .xls فقط")
                elif not (uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls')):
                    st.error(f"❌ امتداد الملف غير صحيح: {uploaded_file.name}")
                    st.info("يرجى التأكد من أن الملف بصيغة .xlsx أو .xls")
                else:
                    if st.button("🔄 Import Excel Data", use_container_width=True):
                        try:
                            excel_exporter = ExcelExporter(st.session_state.data_manager)
                            result = excel_exporter.import_from_2000_column_excel(uploaded_file)
                            
                            if result and result.get('success'):
                                st.success(f"✅ تم استيراد البيانات بنجاح!")
                                st.success(f"📊 عدد المشاريع المستوردة: {result.get('imported_count', 0)}")
                                if result.get('imported_projects'):
                                    st.info(f"المشاريع: {', '.join(result.get('imported_projects', []))}")
                                st.rerun()
                            else:
                                error_msg = result.get('message', 'فشل استيراد البيانات') if result else 'فشل استيراد البيانات'
                                st.error(f"❌ {error_msg}")
                                st.info("تأكد من أن الملف يحتوي على البيانات في الصفوف الصحيحة:")
                                st.write("• الصف 6: التواريخ")
                                st.write("• الصف 7: Planned Total Cost")
                                st.write("• الصف 8: Cumulative Budgeted Cost")
                                st.write("• الصف 13: Actual Cost")
                        except Exception as e:
                            st.error(f"❌ خطأ في معالجة الملف: {str(e)}")
                            st.info("يرجى التحقق من:")
                            st.write("• صحة تنسيق ملف Excel")
                            st.write("• وجود البيانات في الصفوف المحددة")
                            st.write("• عدم وجود خلايا تالفة في الملف")
        
        with excel_cols[2]:
            if st.button("📈 Export Current Data to Excel", use_container_width=True):
                excel_exporter = ExcelExporter(st.session_state.data_manager)
                export_data = excel_exporter.generate_2000_column_template(
                    start_date=date.today().replace(day=1),
                    flow_type="Daily"
                )
                
                if export_data:
                    st.download_button(
                        label="Download Current Data",
                        data=export_data,
                        file_name=f"project_data_export_{date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success("Current project data exported!")
                else:
                    st.error("Failed to export project data")
        
        st.markdown("---")
        
        # Show new project form if needed
        if st.session_state.get('show_new_project_form', False):
            with st.expander("Add New Project", expanded=True):
                new_project_form()
        return
    
    st.success(f"Total loaded projects: {len(all_projects)}")
    
    # Projects list with edit/delete options
    # Table view with parent categories
    show_projects_table_view(all_projects)

def show_projects_table_view(all_projects):
    """Display projects in a Primavera-style table view with parent categories"""
    st.markdown('<div class="english primavera-table"><h3>Projects Management</h3></div>', unsafe_allow_html=True)
    
    # Get grouped projects by category
    grouped_projects = st.session_state.data_manager.get_projects_by_category()
    categories = st.session_state.data_manager.get_parent_categories()
    
    # Add controls for table view
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        show_uncategorized = st.checkbox("Show Uncategorized Projects", value=True)
    with col2:
        auto_expand = st.checkbox("Auto-expand All Categories", value=False)
    
    # Primavera-style table headers matching the image
    st.markdown('<div class="primavera-table">', unsafe_allow_html=True)
    header_cols = st.columns([0.6, 2.5, 0.8, 0.8, 0.8, 0.8, 1.2, 0.8])
    header_cols[0].markdown("**Project ID**")
    header_cols[1].markdown("**Project Name**")
    header_cols[2].markdown("**Start**") 
    header_cols[3].markdown("**Finish**")
    header_cols[4].markdown("**Schedule %<br>Complete**", unsafe_allow_html=True)
    header_cols[5].markdown("**Performance %<br>Complete**", unsafe_allow_html=True)
    header_cols[6].markdown("**BL Project Total<br>Cost**", unsafe_allow_html=True)
    header_cols[7].markdown("**Actual Data**")
    
    st.markdown("---")
    
    # Display projects grouped by parent categories
    for category_name, projects in grouped_projects.items():
        if category_name == 'Uncategorized' and not show_uncategorized:
            continue
        
        # Calculate category totals
        total_cost = sum(project.get('total_budget', 0) for project in projects)
        total_duration = calculate_category_duration(projects)
        
        # Category header like Primavera
        st.markdown(f'''
            <div class="project-header-row">
                ▼ {category_name}
            </div>
        ''', unsafe_allow_html=True)
        
        # Projects in this category
        with st.expander(f"View {category_name} Projects", expanded=auto_expand):
            for i, project in enumerate(projects):
                project_id = project.get('project_id', f"P{project.get('id', '001'):03d}")
                start_date = project.get('start_date', 'Not Set')
                end_date = project.get('end_date', 'Not Set')
                total_cost = project.get('total_budget', 0)
                
                # Get progress data for completion percentages
                progress_data = st.session_state.data_manager.get_progress_data(project['project_name'])
                schedule_complete = 0
                performance_complete = 0
                actual_data = "0.00"
                
                if not progress_data.empty:
                    latest_progress = progress_data.iloc[-1]
                    schedule_complete = latest_progress.get('planned_completion', 0)
                    performance_complete = latest_progress.get('actual_completion', 0)
                    actual_data = f"{performance_complete:.2f}"
                
                # Alternate row colors
                row_class = "project-row-blue" if i % 2 == 0 else "project-row-green"
                st.markdown(f'<div class="{row_class}">', unsafe_allow_html=True)
                row_cols = st.columns([0.6, 2.5, 0.8, 0.8, 0.8, 0.8, 1.2, 0.8])
                
                with row_cols[0]:
                    st.markdown(f"<small>{project_id}</small>", unsafe_allow_html=True)
                
                with row_cols[1]:
                    # Add indentation for sub-projects
                    indent = "&nbsp;&nbsp;&nbsp;&nbsp;" if category_name != "Main Projects" else ""
                    project_name_display = project['project_name']
                    if len(project_name_display) > 40:
                        project_name_display = project_name_display[:37] + "..."
                    st.markdown(f"<small>{indent}<strong>{project_name_display}</strong></small>", unsafe_allow_html=True)
                
                with row_cols[2]:
                    # Format date as DD-MMM-YY
                    if start_date != 'Not Set':
                        try:
                            date_obj = pd.to_datetime(start_date)
                            formatted_date = date_obj.strftime('%d-%b-%y')
                        except:
                            formatted_date = start_date
                    else:
                        formatted_date = ""
                    st.markdown(f"<small>{formatted_date}</small>", unsafe_allow_html=True)
                
                with row_cols[3]:
                    # Format date as DD-MMM-YY
                    if end_date != 'Not Set':
                        try:
                            date_obj = pd.to_datetime(end_date)
                            formatted_date = date_obj.strftime('%d-%b-%y')
                        except:
                            formatted_date = end_date
                    else:
                        formatted_date = ""
                    st.markdown(f"<small>{formatted_date}</small>", unsafe_allow_html=True)
                
                with row_cols[4]:
                    st.markdown(f"<small>{schedule_complete:.2f}%</small>", unsafe_allow_html=True)
                
                with row_cols[5]:
                    st.markdown(f"<small>{performance_complete:.2f}%</small>", unsafe_allow_html=True)
                
                with row_cols[6]:
                    st.markdown(f"<small>${total_cost:,.2f}</small>", unsafe_allow_html=True)
                
                with row_cols[7]:
                    st.markdown(f"<small>{actual_data}</small>", unsafe_allow_html=True)
                
                
                st.markdown('</div>', unsafe_allow_html=True)
                
                # Add action buttons below each row
                action_cols = st.columns([0.6, 0.6, 0.6, 0.6, 4])
                with action_cols[0]:
                    if st.button("✏️ Edit", key=f"edit_table_{category_name}_{i}", use_container_width=True):
                        st.session_state[f'editing_table_{category_name}_{i}'] = True
                
                with action_cols[1]:
                    if st.button("📊 View", key=f"view_table_{category_name}_{i}", use_container_width=True):
                        st.session_state.selected_projects = [project['project_name']]
                        st.success(f"Selected project: {project['project_name']}")
                
                with action_cols[2]:
                    if st.button("🗑️ Delete", key=f"delete_table_{category_name}_{i}", use_container_width=True):
                        st.session_state[f'confirm_delete_table_{category_name}_{i}'] = True
                
                with action_cols[3]:
                    if st.button("➕ Progress", key=f"progress_table_{category_name}_{i}", use_container_width=True):
                        st.session_state[f'show_add_progress_table_{category_name}_{i}'] = True
                
                # Remove the old row_cols[6] content
                
                # Handle inline editing
                if st.session_state.get(f'editing_table_{category_name}_{i}', False):
                    with st.form(f"edit_form_table_{category_name}_{i}"):
                        st.markdown("**Edit Project:**")
                        edit_cols = st.columns(2)
                        
                        with edit_cols[0]:
                            new_project_name = st.text_input("Project Name", value=project['project_name'])
                            new_project_id = st.text_input("Project ID", value=project_id)
                            new_start_date = st.date_input("Start Date", value=pd.to_datetime(start_date) if start_date != 'Not Set' else date.today())
                        
                        with edit_cols[1]:
                            new_end_date = st.date_input("End Date", value=pd.to_datetime(end_date) if end_date != 'Not Set' else date.today())
                            new_total_cost = st.number_input("Total Cost", value=float(total_cost))
                            new_location = st.text_input("Location", value=project.get('project_location', ''))
                        
                        submit_cols = st.columns(2)
                        with submit_cols[0]:
                            if st.form_submit_button("Update"):
                                # Update project logic here
                                st.session_state[f'editing_table_{category_name}_{i}'] = False
                                st.success("Project updated successfully")
                                st.rerun()
                        
                        with submit_cols[1]:
                            if st.form_submit_button("Cancel"):
                                st.session_state[f'editing_table_{category_name}_{i}'] = False
                                st.rerun()
                
                # Handle delete confirmation
                if st.session_state.get(f'confirm_delete_table_{category_name}_{i}', False):
                    st.warning(f"Are you sure you want to delete project: {project['project_name']}?")
                    del_cols = st.columns(2)
                    
                    with del_cols[0]:
                        if st.button("Yes, Delete", key=f"confirm_yes_table_{category_name}_{i}"):
                            success = st.session_state.data_manager.delete_project(project['project_name'])
                            if success:
                                st.success("Project deleted successfully")
                                st.session_state[f'confirm_delete_table_{category_name}_{i}'] = False
                                st.rerun()
                            else:
                                st.error("Error deleting project")
                    
                    with del_cols[1]:
                        if st.button("Cancel", key=f"confirm_no_table_{category_name}_{i}"):
                            st.session_state[f'confirm_delete_table_{category_name}_{i}'] = False
                            st.rerun()
                
                # Handle add progress
                if st.session_state.get(f'show_add_progress_table_{category_name}_{i}', False):
                    with st.form(f"progress_form_table_{category_name}_{i}"):
                        st.markdown("**Add Progress Data:**")
                        progress_cols = st.columns(2)
                        
                        with progress_cols[0]:
                            entry_date = st.date_input("Entry Date", value=date.today())
                            planned_completion = st.slider("Planned Completion %", 0.0, 100.0, 0.0)
                            planned_cost = st.number_input("Planned Cost", value=0.0)
                        
                        with progress_cols[1]:
                            actual_completion = st.slider("Actual Completion %", 0.0, 100.0, 0.0)
                            actual_cost = st.number_input("Actual Cost", value=0.0)
                            notes = st.text_area("Notes", height=60)
                        
                        submit_progress_cols = st.columns(2)
                        with submit_progress_cols[0]:
                            if st.form_submit_button("Add Progress"):
                                progress_data = {
                                    'project_name': project['project_name'],
                                    'entry_date': entry_date.strftime('%Y-%m-%d'),
                                    'planned_completion': planned_completion,
                                    'planned_cost': planned_cost,
                                    'actual_completion': actual_completion,
                                    'actual_cost': actual_cost,
                                    'notes': notes
                                }
                                
                                success = st.session_state.data_manager.add_progress_data(progress_data)
                                if success:
                                    st.success("Progress data added successfully")
                                    st.session_state[f'show_add_progress_table_{category_name}_{i}'] = False
                                    st.rerun()
                                else:
                                    st.error("Error adding progress data")
                        
                        with submit_progress_cols[1]:
                            if st.form_submit_button("Cancel"):
                                st.session_state[f'show_add_progress_table_{category_name}_{i}'] = False
                                st.rerun()
                
    st.markdown('</div>', unsafe_allow_html=True)

def calculate_category_duration(projects):
    """Calculate total duration for projects in a category"""
    total_days = 0
    for project in projects:
        start_date = project.get('start_date')
        end_date = project.get('end_date')
        
        if start_date and end_date and start_date != 'Not Set' and end_date != 'Not Set':
            try:
                start = pd.to_datetime(start_date)
                end = pd.to_datetime(end_date)
                duration = (end - start).days
                total_days += max(0, duration)
            except:
                continue
    
    return total_days

def move_project_in_category(project, category_name, direction):
    """Move project up or down within its category"""
    try:
        current_order = project.get('display_order', 0)
        if direction == "up":
            new_order = current_order - 1
        else:
            new_order = current_order + 1
        
        success = st.session_state.data_manager.update_project_parent_category(
            project['project_name'], 
            project.get('parent_category_id'),
            new_order
        )
        
        if success:
            st.rerun()
    except Exception as e:
        st.error(f"Error moving project: {e}")

def show_original_project_list(all_projects):
    """Original project list view - keeping for reference"""  
    for i, project in enumerate(all_projects):
        project_id = project.get('project_id', f"P{project.get('id', '001'):03d}")
        start_date = project.get('start_date', 'Not Set')
        end_date = project.get('end_date', 'Not Set')
        total_cost = project.get('total_budget', 0)
        
        project_title = f"📁 ID: {project_id} | {project['project_name']} | {start_date} to {end_date} | ${total_cost:,.2f}"
        
        with st.expander(project_title, expanded=False):
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                st.write(f"**تاريخ البدء:** {project.get('start_date', 'غير محدد')}")
                st.write(f"**تاريخ الانتهاء:** {project.get('end_date', 'غير محدد')}")
                st.write(f"**الميزانية:** {project.get('total_budget', 0):,.2f} ريال")
                
                # Get latest progress
                progress = st.session_state.data_manager.get_progress_data(project['project_name'])
                if not progress.empty:
                    latest = progress.iloc[-1]
                    st.write(f"**آخر إنجاز فعلي:** {latest.get('actual_completion', 0):.2f}%")
                    st.write(f"**آخر تكلفة فعلية:** {latest.get('actual_cost', 0):,.2f} ريال")
                else:
                    st.write("**الحالة:** لا توجد بيانات إنجاز")
            
            with col2:
                if st.button(f"✏️ تعديل", key=f"edit_{i}"):
                    st.session_state[f'editing_{i}'] = True
                
                if st.button(f"📊 عرض التفاصيل", key=f"details_{i}"):
                    # Set as selected project and switch to charts
                    st.session_state.selected_projects = [project['project_name']]
                    st.success(f"تم اختيار المشروع: {project['project_name']}")
                    st.info("يمكنك الآن الانتقال إلى المخططات أو لوحة المراقبة لعرض التفاصيل")
            
            with col3:
                if st.button(f"🗑️ حذف", key=f"delete_{i}"):
                    st.session_state[f'confirm_delete_{i}'] = True
                
                # Add progress data button
                add_progress_btn = st.button(f"➕ إضافة إنجاز", key=f"add_progress_{i}")
                if add_progress_btn:
                    st.session_state[f'show_add_progress_{i}'] = True
            
            # Handle editing
            if st.session_state.get(f'editing_{i}', False):
                st.markdown("**تعديل بيانات المشروع:**")
                edit_project_form(project, i)
            
            # Handle delete confirmation
            if st.session_state.get(f'confirm_delete_{i}', False):
                st.warning(f"هل أنت متأكد من حذف المشروع: {project['project_name']}؟")
                col_yes, col_no = st.columns(2)
                
                with col_yes:
                    if st.button("نعم، احذف", key=f"confirm_yes_{i}"):
                        success = st.session_state.data_manager.delete_project(project['project_name'])
                        if success:
                            st.success("تم حذف المشروع بنجاح")
                            st.session_state[f'confirm_delete_{i}'] = False
                            st.rerun()
                        else:
                            st.error("خطأ في حذف المشروع")
                
                with col_no:
                    if st.button("إلغاء", key=f"confirm_no_{i}"):
                        st.session_state[f'confirm_delete_{i}'] = False
                        st.rerun()
            
            # Handle add progress
            if st.session_state.get(f'show_add_progress_{i}', False):
                st.markdown("**إضافة بيانات إنجاز جديدة:**")
                add_progress_form(project['project_name'], i)
    
    # Add new project button
    st.divider()
    if st.button("➕ إضافة مشروع جديد", use_container_width=True):
        st.session_state.show_new_project_form = True
    
    # Show new project form if needed
    if st.session_state.get('show_new_project_form', False):
        with st.expander("إضافة مشروع جديد", expanded=True):
            new_project_form()

def project_basic_info():
    st.markdown('<div class="rtl"><h3>معلومات المشروع الأساسية</h3></div>', unsafe_allow_html=True)
    
    with st.form("project_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            project_name = st.text_input("اسم المشروع *", key="project_name_input")
            executing_company = st.text_input("الشركة المنفذة *", key="executing_company_input")
            start_date = st.date_input("تاريخ البدء *", key="start_date_input")
            total_budget = st.number_input("الميزانية الإجمالية (ريال) *", min_value=0.0, key="budget_input")
        
        with col2:
            consulting_company = st.text_input("الشركة الاستشارية", key="consulting_company_input")
            end_date = st.date_input("تاريخ الانتهاء المخطط *", key="end_date_input")
            project_location = st.text_input("موقع المشروع", key="location_input")
            project_type = st.selectbox("نوع المشروع", 
                                      ["", "مباني سكنية", "مباني تجارية", "مشاريع بنية تحتية", 
                                       "مشاريع صناعية", "أخرى"], key="project_type_input")
        
        project_description = st.text_area("وصف المشروع", key="description_input")
        
        submitted = st.form_submit_button("حفظ معلومات المشروع")
        
        if submitted:
            if project_name and executing_company and start_date and end_date and total_budget > 0:
                if start_date <= end_date:
                    project_data = {
                        'project_name': project_name,
                        'executing_company': executing_company,
                        'consulting_company': consulting_company,
                        'start_date': start_date,
                        'end_date': end_date,
                        'total_budget': total_budget,
                        'project_location': project_location,
                        'project_type': project_type,
                        'project_description': project_description,
                        'created_date': datetime.now()
                    }
                    
                    success = st.session_state.data_manager.add_project(project_data)
                    if success:
                        st.success("تم حفظ معلومات المشروع بنجاح!")
                        st.session_state.selected_project = project_name
                        st.rerun()
                    else:
                        st.error("خطأ في حفظ البيانات. يرجى المحاولة مرة أخرى.")
                else:
                    st.error("تاريخ الانتهاء يجب أن يكون بعد تاريخ البدء")
            else:
                st.error("يرجى ملء جميع الحقول المطلوبة (*)")

def daily_monthly_data():
    st.markdown('<div class="rtl"><h3>إدخال الإنجاز الفعلي فقط</h3></div>', unsafe_allow_html=True)
    st.info("البيانات المخططة (Planned) يتم إدخالها من ملف Excel فقط")
    
    if not st.session_state.selected_project:
        st.warning("يرجى اختيار مشروع أولاً من القائمة الجانبية")
        return
    
    # Actual progress data entry form only
    with st.form("actual_progress_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            entry_date = st.date_input("تاريخ الإدخال", value=date.today())
            actual_completion = st.number_input("نسبة الإنجاز الفعلي (%)", min_value=0.0, max_value=100.0, step=0.1)
            actual_cost = st.number_input("التكلفة الفعلية حتى التاريخ (ريال)", min_value=0.0)
        
        with col2:
            notes = st.text_area("ملاحظات")
            st.markdown("**ملاحظة:** سيتم إضافة هذه البيانات إلى ملف Excel تلقائياً عند التصدير")
        
        submitted = st.form_submit_button("حفظ الإنجاز الفعلي")
        
        if submitted:
            # Get latest planned data for this project
            existing_progress = st.session_state.data_manager.get_progress_data(st.session_state.selected_project)
            planned_completion = 0
            planned_cost = 0
            
            if not existing_progress.empty:
                latest_planned = existing_progress.iloc[-1]
                planned_completion = latest_planned.get('planned_completion', 0)
                planned_cost = latest_planned.get('planned_cost', 0)
            
            progress_data = {
                'project_name': st.session_state.selected_project,
                'entry_date': entry_date,
                'planned_completion': planned_completion,  # Keep existing planned data
                'planned_cost': planned_cost,  # Keep existing planned cost
                'actual_completion': actual_completion,
                'actual_cost': actual_cost,
                'notes': f"{notes} (إدخال فعلي من الواجهة - {datetime.now().strftime('%Y-%m-%d %H:%M')})"
            }
            
            success = st.session_state.data_manager.add_progress_data(progress_data)
            if success:
                st.success("تم حفظ الإنجاز الفعلي بنجاح! سيظهر في ملف Excel عند التصدير.")
                
                # Auto-export updated Excel file
                from excel_exporter import ExcelExporter
                exporter = ExcelExporter(st.session_state.data_manager)
                
                # Get current project data for export
                project_data = st.session_state.data_manager.get_project_by_name(st.session_state.selected_project)
                if project_data:
                    project_data['project_id'] = project_data.get('project_description', '').replace('Project ID: ', '') or 'AUTO'
                    updated_template = exporter.export_project_template([project_data])
                    
                    if updated_template:
                        st.download_button(
                            label="⬇️ تحميل ملف Excel المحدث",
                            data=updated_template,
                            file_name=f"{st.session_state.selected_project}_updated.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                st.rerun()
            else:
                st.error("خطأ في حفظ البيانات")
    
    # Display existing progress data
    st.markdown('<div class="rtl"><h4>بيانات التقدم المحفوظة</h4></div>', unsafe_allow_html=True)
    progress_data = st.session_state.data_manager.get_progress_data(st.session_state.selected_project)
    
    if not progress_data.empty:
        # Format the dataframe for display
        display_df = progress_data.copy()
        display_df.columns = ['تاريخ الإدخال', 'نسبة الإنجاز المخطط (%)', 'التكلفة المخططة', 
                             'نسبة الإنجاز الفعلي (%)', 'التكلفة الفعلية', 'ملاحظات']
        display_df['التكلفة المخططة'] = display_df['التكلفة المخططة'].apply(format_currency)
        display_df['التكلفة الفعلية'] = display_df['التكلفة الفعلية'].apply(format_currency)
        
        st.dataframe(display_df, use_container_width=True)
    else:
        st.info("لا توجد بيانات تقدم محفوظة لهذا المشروع")

def resources_equipment():
    st.markdown('<div class="rtl"><h3>الموارد والمعدات</h3></div>', unsafe_allow_html=True)
    
    if not st.session_state.selected_project:
        st.warning("يرجى اختيار مشروع أولاً من القائمة الجانبية")
        return
    
    resource_tab1, resource_tab2 = st.tabs(["العمالة", "المعدات"])
    
    with resource_tab1:
        labor_resources()
    
    with resource_tab2:
        equipment_resources()

def labor_resources():
    with st.form("labor_form"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            job_title = st.text_input("المسمى الوظيفي")
            quantity = st.number_input("العدد", min_value=1, step=1)
        
        with col2:
            daily_rate = st.number_input("الأجر اليومي (ريال)", min_value=0.0)
            start_date = st.date_input("تاريخ البدء", key="labor_start")
        
        with col3:
            end_date = st.date_input("تاريخ الانتهاء", key="labor_end")
            notes = st.text_input("ملاحظات")
        
        submitted = st.form_submit_button("إضافة عامل/مجموعة عمال")
        
        if submitted and job_title and quantity > 0:
            labor_data = {
                'project_name': st.session_state.selected_project,
                'resource_type': 'labor',
                'name': job_title,
                'quantity': quantity,
                'daily_rate': daily_rate,
                'start_date': start_date,
                'end_date': end_date,
                'notes': notes
            }
            
            success = st.session_state.data_manager.add_resource(labor_data)
            if success:
                st.success("تم إضافة العمالة بنجاح!")
            else:
                st.error("خطأ في حفظ البيانات")

def equipment_resources():
    with st.form("equipment_form"):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            equipment_name = st.text_input("اسم المعدة/الآلة")
            quantity = st.number_input("العدد", min_value=1, step=1)
        
        with col2:
            daily_rate = st.number_input("تكلفة التشغيل اليومية (ريال)", min_value=0.0)
            start_date = st.date_input("تاريخ البدء", key="equipment_start")
        
        with col3:
            end_date = st.date_input("تاريخ الانتهاء", key="equipment_end")
            notes = st.text_input("ملاحظات")
        
        submitted = st.form_submit_button("إضافة معدة")
        
        if submitted and equipment_name and quantity > 0:
            equipment_data = {
                'project_name': st.session_state.selected_project,
                'resource_type': 'equipment',
                'name': equipment_name,
                'quantity': quantity,
                'daily_rate': daily_rate,
                'start_date': start_date,
                'end_date': end_date,
                'notes': notes
            }
            
            success = st.session_state.data_manager.add_resource(equipment_data)
            if success:
                st.success("تم إضافة المعدة بنجاح!")
            else:
                st.error("خطأ في حفظ البيانات")

def reports_tab():
    st.markdown('<div class="rtl"><h2>التقارير</h2></div>', unsafe_allow_html=True)
    
    report_type = st.selectbox(
        "نوع التقرير",
        ["تقرير التدفق النقدي", "تقرير KPI للمحفظة", "تقرير تفصيلي للمشروع"]
    )
    
    if report_type == "تقرير التدفق النقدي":
        cash_flow_report()
    elif report_type == "تقرير KPI للمحفظة":
        portfolio_kpi_report()
    elif report_type == "تقرير تفصيلي للمشروع":
        detailed_project_report()

def cash_flow_report():
    st.markdown('<div class="rtl"><h3>تقرير التدفق النقدي</h3></div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input("من تاريخ", key="cash_flow_start")
        # Use selected projects from sidebar
        selected_projects = st.session_state.get('selected_projects', [])
        if selected_projects:
            if len(selected_projects) == 1:
                project_selection = selected_projects[0]
            else:
                project_selection = "المشاريع المحددة"
            st.info(f"المشاريع المحددة: {', '.join(selected_projects)}")
        else:
            project_selection = "جميع المشاريع"
            st.warning("لم يتم تحديد مشاريع، سيتم عرض جميع المشاريع")
    
    with col2:
        end_date = st.date_input("إلى تاريخ", key="cash_flow_end")
    
    if st.button("إنشاء تقرير التدفق النقدي"):
        if validate_date_range(start_date, end_date):
            exporter = ExcelExporter(st.session_state.data_manager)
            
            if not selected_projects or len(selected_projects) > 1:
                report_data = exporter.generate_portfolio_cash_flow_report(start_date, end_date)
            else:
                report_data = exporter.generate_project_cash_flow_report(selected_projects[0], start_date, end_date)
            
            if report_data is not None and not report_data.empty:
                st.dataframe(report_data, use_container_width=True)
                
                # Export to Excel
                excel_buffer = exporter.export_cash_flow_to_excel(report_data, project_selection, start_date, end_date)
                if excel_buffer:
                    st.download_button(
                        label="تحميل كملف Excel",
                        data=excel_buffer,
                        file_name=f"cash_flow_report_{start_date}_{end_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("لا توجد بيانات للفترة المحددة")
        else:
            st.error("تاريخ النهاية يجب أن يكون بعد تاريخ البداية")

def portfolio_kpi_report():
    st.markdown('<div class="rtl"><h3>تقرير KPI للمحفظة</h3></div>', unsafe_allow_html=True)
    
    calculator = EVMCalculator(st.session_state.data_manager)
    portfolio_kpi = calculator.calculate_portfolio_kpi()
    
    if portfolio_kpi:
        # Display KPI metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("القيمة المخططة (PV)", format_currency(portfolio_kpi.get('total_pv', 0)))
            st.metric("مؤشر أداء التكلفة (CPI)", f"{portfolio_kpi.get('avg_cpi', 0):.2f}")
        
        with col2:
            st.metric("القيمة المكتسبة (EV)", format_currency(portfolio_kpi.get('total_ev', 0)))
            st.metric("مؤشر أداء الجدولة (SPI)", f"{portfolio_kpi.get('avg_spi', 0):.2f}")
        
        with col3:
            st.metric("التكلفة الفعلية (AC)", format_currency(portfolio_kpi.get('total_ac', 0)))
            st.metric("انحراف التكلفة (CV)", format_currency(portfolio_kpi.get('total_cv', 0)))
        
        with col4:
            st.metric("عدد المشاريع", portfolio_kpi.get('total_projects', 0))
            st.metric("انحراف الجدولة (SV)", format_currency(portfolio_kpi.get('total_sv', 0)))
        
        # Export functionality
        if st.button("تصدير تقرير KPI"):
            exporter = ExcelExporter(st.session_state.data_manager)
            excel_buffer = exporter.export_portfolio_kpi_to_excel(portfolio_kpi)
            if excel_buffer:
                st.download_button(
                    label="تحميل تقرير KPI كملف Excel",
                    data=excel_buffer,
                    file_name=f"portfolio_kpi_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.warning("لا توجد بيانات كافية لحساب مؤشرات الأداء")

def detailed_project_report():
    st.markdown('<div class="rtl"><h3>تقرير تفصيلي للمشروع</h3></div>', unsafe_allow_html=True)
    
    if not st.session_state.selected_project:
        st.warning("يرجى اختيار مشروع من القائمة الجانبية")
        return
    
    calculator = EVMCalculator(st.session_state.data_manager)
    project_kpi = calculator.calculate_project_kpi(st.session_state.selected_project)
    
    if project_kpi:
        # Project basic info
        project_info = st.session_state.data_manager.get_project_info(st.session_state.selected_project)
        if project_info:
            st.markdown(f"**اسم المشروع:** {project_info['project_name']}")
            st.markdown(f"**الشركة المنفذة:** {project_info['executing_company']}")
            st.markdown(f"**الميزانية الإجمالية:** {format_currency(project_info['total_budget'])}")
        
        # KPI metrics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("القيمة المخططة (PV)", format_currency(project_kpi.get('pv', 0)))
            st.metric("مؤشر أداء التكلفة (CPI)", f"{project_kpi.get('cpi', 0):.2f}")
        
        with col2:
            st.metric("القيمة المكتسبة (EV)", format_currency(project_kpi.get('ev', 0)))
            st.metric("مؤشر أداء الجدولة (SPI)", f"{project_kpi.get('spi', 0):.2f}")
        
        with col3:
            st.metric("التكلفة الفعلية (AC)", format_currency(project_kpi.get('ac', 0)))
            st.metric("حالة المشروع", project_kpi.get('status', 'غير محدد'))

def charts_tab():
    st.markdown('<div class="rtl"><h2>المخططات والرسوم البيانية</h2></div>', unsafe_allow_html=True)
    
    chart_type = st.selectbox(
        "نوع المخطط",
        ["منحنى S للتكلفة", "مقارنة أداء المشاريع", "تقدم المشروع عبر الزمن"]
    )
    
    if chart_type == "منحنى S للتكلفة":
        s_curve_chart()
    elif chart_type == "مقارنة أداء المشاريع":
        performance_comparison_chart()
    elif chart_type == "تقدم المشروع عبر الزمن":
        project_progress_chart()

def s_curve_chart():
    st.markdown('<div class="rtl"><h3>منحنى S للتكلفة</h3></div>', unsafe_allow_html=True)
    
    # Use selected projects from sidebar
    selected_projects = st.session_state.get('selected_projects', [])
    
    if not selected_projects:
        st.warning("يرجى اختيار مشروع أو أكثر من القائمة الجانبية")
        return
    
    st.info(f"عرض منحنى S للمشاريع: {', '.join(selected_projects)}")
    
    fig = create_s_curve(st.session_state.data_manager, selected_projects)
    if fig:
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("لا توجد بيانات كافية لإنشاء منحنى S للمشاريع المحددة")

def performance_comparison_chart():
    st.markdown('<div class="rtl"><h3>مقارنة أداء المشاريع</h3></div>', unsafe_allow_html=True)
    
    calculator = EVMCalculator(st.session_state.data_manager)
    portfolio_data = calculator.get_all_projects_performance()
    
    if portfolio_data:
        df = pd.DataFrame(portfolio_data)
        
        # Create performance comparison chart
        col1, col2 = st.columns(2)
        
        with col1:
            # CPI comparison
            st.bar_chart(df.set_index('project_name')['cpi'])
            st.caption("مؤشر أداء التكلفة (CPI)")
        
        with col2:
            # SPI comparison  
            st.bar_chart(df.set_index('project_name')['spi'])
            st.caption("مؤشر أداء الجدولة (SPI)")
    else:
        st.warning("لا توجد بيانات كافية للمقارنة")

def project_progress_chart():
    st.markdown('<div class="rtl"><h3>تقدم المشروع عبر الزمن</h3></div>', unsafe_allow_html=True)
    
    selected_projects = st.session_state.get('selected_projects', [])
    
    if not selected_projects:
        st.warning("يرجى اختيار مشروع أو أكثر من القائمة الجانبية")
        return
    
    if len(selected_projects) == 1:
        # Single project detailed view
        project_name = selected_projects[0]
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        if not progress_data.empty:
            progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
            progress_data = progress_data.sort_values('entry_date')
            
            st.line_chart(
                progress_data.set_index('entry_date')[['planned_completion', 'actual_completion']]
            )
            st.caption(f"مقارنة نسبة الإنجاز المخطط مقابل الفعلي - {project_name}")
        else:
            st.warning(f"لا توجد بيانات تقدم للمشروع: {project_name}")
    else:
        # Multiple projects comparison
        st.info(f"عرض مقارنة التقدم لـ {len(selected_projects)} مشروع")
        
        all_data = {}
        for project_name in selected_projects:
            progress_data = st.session_state.data_manager.get_progress_data(project_name)
            if not progress_data.empty:
                progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
                progress_data = progress_data.sort_values('entry_date')
                # Get latest progress
                latest_actual = progress_data['actual_completion'].iloc[-1]
                latest_planned = progress_data['planned_completion'].iloc[-1]
                all_data[project_name] = {'Actual': latest_actual, 'Planned': latest_planned}
        
        if all_data:
            df = pd.DataFrame(all_data).T
            st.bar_chart(df)
            st.caption("مقارنة آخر نسب الإنجاز للمشاريع المحددة")
        else:
            st.warning("لا توجد بيانات تقدم للمشاريع المحددة")

def dashboard_tab():
    st.markdown('<div class="rtl"><h2>لوحة المراقبة</h2></div>', unsafe_allow_html=True)
    
    # Filters
    col1, col2, col3 = st.columns(3)
    
    with col1:
        status_filter = st.selectbox(
            "فلتر حسب الحالة",
            ["جميع المشاريع", "متقدم", "متأخر", "على المسار"]
        )
    
    with col2:
        spi_threshold = st.slider("حد مؤشر SPI", 0.5, 1.5, 1.0, 0.1)
    
    with col3:
        cpi_threshold = st.slider("حد مؤشر CPI", 0.5, 1.5, 1.0, 0.1)
    
    # Dashboard content
    calculator = EVMCalculator(st.session_state.data_manager)
    dashboard_data = calculator.get_dashboard_data(status_filter, spi_threshold, cpi_threshold)
    
    if dashboard_data:
        create_kpi_dashboard(dashboard_data)
    else:
        st.warning("لا توجد مشاريع تطابق المعايير المحددة")


def powerpoint_tab():
    st.markdown('<div class="rtl"><h2>إنشاء عروض PowerPoint تلقائية</h2></div>', unsafe_allow_html=True)
    
    selected_projects = st.session_state.get('selected_projects', [])
    
    if not selected_projects:
        st.warning("يرجى اختيار المشاريع المراد عرضها من القائمة الجانبية")
        return
    
    st.success(f"سيتم إنشاء العرض للمشاريع التالية: {', '.join(selected_projects)}")
    
    # Date range selection
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input("تاريخ البداية للبيانات المعروضة", key="ppt_start")
    
    with col2:
        end_date = st.date_input("تاريخ النهاية للبيانات المعروضة", key="ppt_end")
    
    st.divider()
    st.markdown('<div class="rtl"><h3>اختر نوع العرض المطلوب</h3></div>', unsafe_allow_html=True)
    
    # Define 10 presentation templates with descriptions and previews
    presentation_templates = [
        {
            "title": "العرض التنفيذي الشامل",
            "description": "عرض مخصص للإدارة العليا يحتوي على الملخص التنفيذي + مؤشرات الأداء الرئيسية + التوصيات",
            "contents": ["شريحة العنوان", "الملخص التنفيذي", "مؤشرات الأداء الرئيسية", "نظرة عامة على المشاريع", "التوصيات والقرارات"],
            "charts": ["KPI Dashboard", "Progress Overview", "Financial Summary"],
            "icon": "👔",
            "slides_count": 8
        },
        {
            "title": "تقرير التقدم التفصيلي",
            "description": "تقرير شامل للتقدم يتضمن Gantt Chart + S-Curve + تحليل الإنجاز لكل مشروع",
            "contents": ["عنوان التقرير", "جدولة المشاريع", "منحنيات الأداء", "تفاصيل كل مشروع", "تحليل الانحرافات"],
            "charts": ["Gantt Chart", "S-Curve", "Progress Tracking", "Variance Analysis"],
            "icon": "📊",
            "slides_count": 12
        },
        {
            "title": "العرض المالي والميزانيات",
            "description": "تركز على الجانب المالي: التكاليف + الميزانيات + التدفقات النقدية + تحليل العائد",
            "contents": ["نظرة مالية عامة", "تحليل الميزانيات", "التدفقات النقدية", "مؤشرات الربحية", "توقعات مالية"],
            "charts": ["Cost Analysis", "Budget vs Actual", "Cash Flow", "Financial KPIs"],
            "icon": "💰",
            "slides_count": 10
        },
        {
            "title": "مقارنة أداء المشاريع",
            "description": "مقارنة شاملة بين المشاريع المختارة مع Bar Charts + Pie Charts + جداول مقارنة",
            "contents": ["مقدمة المقارنة", "مقارنة الأداء", "مقارنة التكاليف", "مقارنة الجداول الزمنية", "أفضل الممارسات"],
            "charts": ["Comparative Bar Charts", "Performance Pie Charts", "Timeline Comparison"],
            "icon": "⚖️", 
            "slides_count": 9
        },
        {
            "title": "لوحة المراقبة التفاعلية",
            "description": "عرض يحاكي لوحة المراقبة مع مؤشرات حية + إشارات إنذار + حالة المشاريع",
            "contents": ["لوحة المراقبة الرئيسية", "مؤشرات التحذير", "حالة المشاريع", "الإجراءات المطلوبة", "المتابعة"],
            "charts": ["Real-time Dashboard", "Alert Indicators", "Status Overview", "Action Items"],
            "icon": "🎛️",
            "slides_count": 7
        },
        {
            "title": "تحليل المخاطر والجودة",
            "description": "تقييم المخاطر + مؤشرات الجودة + خطط التخفيف + إجراءات الوقاية",
            "contents": ["تحليل المخاطر", "مؤشرات الجودة", "خطط التخفيف", "إجراءات الوقاية", "المراجعة المستمرة"],
            "charts": ["Risk Matrix", "Quality Indicators", "Mitigation Plans", "Prevention Measures"],
            "icon": "🛡️",
            "slides_count": 11
        },
        {
            "title": "العرض الزمني والجدولة",
            "description": "يركز على الجدولة الزمنية: Timeline + المراحل + التسليمات + الأحداث المهمة",
            "contents": ["الجدولة العامة", "المراحل الرئيسية", "التسليمات المهمة", "التواريخ الحرجة", "خطط التعديل"],
            "charts": ["Project Timeline", "Milestones Chart", "Critical Path", "Schedule Adjustments"],
            "icon": "📅",
            "slides_count": 8
        },
        {
            "title": "تقرير الموارد والفرق",
            "description": "إدارة الموارد البشرية + التخصيصات + الكفاءات + توزيع الأعمال على الفرق",
            "contents": ["الموارد المخصصة", "كفاءة الفرق", "توزيع الأعمال", "التطوير المطلوب", "خطط التحسين"],
            "charts": ["Resource Allocation", "Team Efficiency", "Workload Distribution", "Skill Development"],
            "icon": "👥",
            "slides_count": 9
        },
        {
            "title": "عرض العملاء والأصحاب",
            "description": "مصمم خصيصاً لتقديمه للعملاء وأصحاب المشاريع مع التركيز على القيمة المضافة",
            "contents": ["رسالة ترحيب", "إنجازات المشروع", "القيمة المحققة", "الجودة المقدمة", "الخطوات القادمة"],
            "charts": ["Achievement Highlights", "Value Creation", "Quality Metrics", "Future Roadmap"],
            "icon": "🤝",
            "slides_count": 10
        },
        {
            "title": "التقرير الشامل المتكامل",
            "description": "تقرير شامل يجمع كل العناصر: مالي + تقني + إداري + مخاطر + توصيات شاملة",
            "contents": ["ملخص شامل", "التحليل المالي", "التحليل التقني", "إدارة المخاطر", "التوصيات الشاملة", "الملاحق"],
            "charts": ["Comprehensive Dashboard", "All Chart Types", "Detailed Analysis", "Complete KPIs"],
            "icon": "📋",
            "slides_count": 15
        }
    ]
    
    # Display templates in a grid
    cols = st.columns(2)
    selected_template = None
    
    for i, template in enumerate(presentation_templates):
        col = cols[i % 2]
        
        with col:
            with st.container():
                st.markdown(f"### {template['icon']} {template['title']}")
                st.write(template['description'])
                
                # Show contents preview
                with st.expander(f"محتويات العرض ({template['slides_count']} شريحة)"):
                    for content in template['contents']:
                        st.write(f"• {content}")
                    
                    st.markdown("**الرسوم البيانية المدرجة:**")
                    for chart in template['charts']:
                        st.write(f"📊 {chart}")
                
                if st.button(f"اختر هذا العرض", key=f"select_{i}", use_container_width=True):
                    selected_template = template
                    
                st.divider()
    
    # Generate selected presentation
    if selected_template:
        st.success(f"تم اختيار: {selected_template['title']}")
        
        with st.expander("خيارات إضافية للعرض المختار", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                include_company_logo = st.checkbox("تضمين شعار الشركة", value=True)
                include_recommendations = st.checkbox("تضمين التوصيات التلقائية", value=True)
                include_appendix = st.checkbox("تضمين الملاحق والبيانات التفصيلية", value=False)
            
            with col2:
                language_preference = st.selectbox("لغة العرض", ["العربية", "English", "Arabic + English"])
                color_scheme = st.selectbox("نظام الألوان", ["أزرق مهني", "رمادي أعمال", "أخضر طبيعي", "ذهبي فاخر"])
                custom_title = st.text_input("عنوان مخصص (اختياري)", placeholder=selected_template['title'])
        
        if st.button("🎥 إنشاء العرض المختار", use_container_width=True, type="primary"):
            with st.spinner(f"جاري إنشاء {selected_template['title']}..."):
                try:
                    from powerpoint_generator import PowerPointGenerator
                    
                    ppt_generator = PowerPointGenerator(st.session_state.data_manager)
                    presentation_data = ppt_generator.create_specialized_presentation(
                        template_type=selected_template['title'],
                        projects=selected_projects, 
                        start_date=start_date, 
                        end_date=end_date,
                        custom_options={
                            'include_logo': include_company_logo,
                            'include_recommendations': include_recommendations,
                            'include_appendix': include_appendix,
                            'language': language_preference,
                            'color_scheme': color_scheme,
                            'custom_title': custom_title or selected_template['title']
                        }
                    )
                    
                    if presentation_data:
                        filename = f"{selected_template['title'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pptx"
                        
                        st.success("تم إنشاء العرض التقديمي بنجاح!")
                        
                        st.download_button(
                            label="📥 تحميل العرض التقديمي",
                            data=presentation_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                            use_container_width=True
                        )
                        
                        # Display final presentation info
                        st.info(f"""
                        **تفاصيل العرض النهائي:**
                        - النوع: {selected_template['title']}
                        - عدد الشرائح: {selected_template['slides_count']}
                        - عدد المشاريع: {len(selected_projects)}
                        - فترة البيانات: {start_date} إلى {end_date}
                        - اللغة: {language_preference}
                        - نظام الألوان: {color_scheme}
                        """)
                    else:
                        st.error("فشل في إنشاء العرض التقديمي. يرجى المحاولة مرة أخرى.")
                        
                except Exception as e:
                    st.error(f"خطأ في إنشاء العرض: {str(e)}")
                    st.info("سيتم إنشاء عرض أساسي بدلاً من ذلك")
    
    else:
        st.info("👆 اختر نوع العرض من القائمة أعلاه لبدء الإنشاء")

def edit_project_form(project, index):
    """Form for editing existing project"""
    with st.form(f"edit_project_{index}"):
        col1, col2 = st.columns(2)
        
        with col1:
            project_name = st.text_input("اسم المشروع*", value=project['project_name'])
            executing_company = st.text_input("الشركة المنفذة*", value=project.get('executing_company', ''))
            consulting_company = st.text_input("الشركة الاستشارية*", value=project.get('consulting_company', ''))
            project_location = st.text_input("موقع المشروع", value=project.get('project_location', ''))
        
        with col2:
            start_date = st.date_input("تاريخ البدء*", value=pd.to_datetime(project['start_date']).date() if project.get('start_date') else date.today())
            end_date = st.date_input("تاريخ الانتهاء*", value=pd.to_datetime(project['end_date']).date() if project.get('end_date') else date.today())
            total_budget = st.number_input("إجمالي الميزانية (ريال)*", value=float(project.get('total_budget', 0)), min_value=0.0)
            project_type = st.selectbox("نوع المشروع", 
                                      ["مشروع إنشائي", "مشروع صيانة", "مشروع تطوير", "مشروع استشاري"],
                                      index=0 if not project.get('project_type') else ["مشروع إنشائي", "مشروع صيانة", "مشروع تطوير", "مشروع استشاري"].index(project.get('project_type', "مشروع إنشائي")))
        
        project_description = st.text_area("وصف المشروع", value=project.get('project_description', ''))
        
        col_save, col_cancel = st.columns(2)
        
        with col_save:
            save_changes = st.form_submit_button("💾 حفظ التعديلات", use_container_width=True)
        
        with col_cancel:
            cancel_edit = st.form_submit_button("❌ إلغاء", use_container_width=True)
        
        if save_changes:
            if project_name and executing_company and consulting_company:
                updated_data = {
                    'project_name': project_name,
                    'executing_company': executing_company,
                    'consulting_company': consulting_company,
                    'start_date': start_date,
                    'end_date': end_date,
                    'total_budget': total_budget,
                    'project_location': project_location,
                    'project_type': project_type,
                    'project_description': project_description,
                    'created_date': project.get('created_date', datetime.now())
                }
                
                if start_date <= end_date:
                    # For now, use add_project since update_project method needs to be added to DataManager
                    success = st.session_state.data_manager.add_project(updated_data)
                    if success:
                        st.success("تم تحديث المشروع بنجاح!")
                        st.session_state[f'editing_{index}'] = False
                        st.rerun()
                    else:
                        st.error("خطأ في تحديث المشروع")
                else:
                    st.error("تاريخ الانتهاء يجب أن يكون بعد تاريخ البدء")
            else:
                st.error("يرجى ملء جميع الحقول المطلوبة (*)")
        
        if cancel_edit:
            st.session_state[f'editing_{index}'] = False
            st.rerun()


def add_progress_form(project_name, index):
    """Form for adding progress data to specific project"""
    with st.form(f"add_progress_{index}"):
        col1, col2 = st.columns(2)
        
        with col1:
            entry_date = st.date_input("تاريخ الإدخال", value=date.today())
            actual_completion = st.number_input("نسبة الإنجاز الفعلي (%)", min_value=0.0, max_value=100.0, step=0.1)
            actual_cost = st.number_input("التكلفة الفعلية حتى التاريخ (ريال)", min_value=0.0)
        
        with col2:
            notes = st.text_area("ملاحظات")
            st.info("سيتم إضافة هذه البيانات إلى قاعدة البيانات وملف Excel عند التصدير")
        
        submitted = st.form_submit_button("حفظ الإنجاز الفعلي")
        cancel = st.form_submit_button("إلغاء")
        
        if submitted:
            # Get latest planned data for this project
            existing_progress = st.session_state.data_manager.get_progress_data(project_name)
            planned_completion = 0
            planned_cost = 0
            
            if not existing_progress.empty:
                latest_planned = existing_progress.iloc[-1]
                planned_completion = latest_planned.get('planned_completion', 0)
                planned_cost = latest_planned.get('planned_cost', 0)
            
            progress_data = {
                'project_name': project_name,
                'entry_date': entry_date,
                'planned_completion': planned_completion,
                'planned_cost': planned_cost,
                'actual_completion': actual_completion,
                'actual_cost': actual_cost,
                'notes': f"{notes} (إدخال من قائمة المشاريع - {datetime.now().strftime('%Y-%m-%d %H:%M')})"
            }
            
            success = st.session_state.data_manager.add_progress_data(progress_data)
            if success:
                st.success("تم حفظ الإنجاز الفعلي بنجاح!")
                st.session_state[f'show_add_progress_{index}'] = False
                st.rerun()
            else:
                st.error("خطأ في حفظ البيانات")
        
        if cancel:
            st.session_state[f'show_add_progress_{index}'] = False
            st.rerun()


def new_project_form():
    """Form for adding completely new project"""
    with st.form("new_project_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            project_name = st.text_input("اسم المشروع*")
            executing_company = st.text_input("الشركة المنفذة*")
            consulting_company = st.text_input("الشركة الاستشارية*", value="شركة عبد الله السعيد للاستشارات الهندسية")
            project_location = st.text_input("موقع المشروع")
        
        with col2:
            start_date = st.date_input("تاريخ البدء*", value=date.today())
            end_date = st.date_input("تاريخ الانتهاء*", value=date.today())
            total_budget = st.number_input("إجمالي الميزانية (ريال)*", min_value=0.0)
            project_type = st.selectbox("نوع المشروع", 
                                      ["مشروع إنشائي", "مشروع صيانة", "مشروع تطوير", "مشروع استشاري"])
        
        project_description = st.text_area("وصف المشروع")
        
        col_submit, col_cancel = st.columns(2)
        
        with col_submit:
            submitted = st.form_submit_button("➕ إضافة المشروع", use_container_width=True)
        
        with col_cancel:
            cancelled = st.form_submit_button("❌ إلغاء", use_container_width=True)
        
        if submitted:
            if project_name and executing_company and consulting_company and total_budget > 0:
                project_data = {
                    'project_name': project_name,
                    'executing_company': executing_company,
                    'consulting_company': consulting_company,
                    'start_date': start_date,
                    'end_date': end_date,
                    'total_budget': total_budget,
                    'project_location': project_location,
                    'project_type': project_type,
                    'project_description': project_description,
                    'created_date': datetime.now()
                }
                
                if start_date <= end_date:
                    success = st.session_state.data_manager.add_project(project_data)
                    if success:
                        st.success(f"تم إضافة المشروع '{project_name}' بنجاح!")
                        st.session_state.show_new_project_form = False
                        st.rerun()
                    else:
                        st.error("خطأ في إضافة المشروع. تأكد من عدم وجود مشروع بنفس الاسم")
                else:
                    st.error("تاريخ الانتهاء يجب أن يكون بعد تاريخ البدء")
            else:
                st.error("يرجى ملء جميع الحقول المطلوبة (*)")
        
        if cancelled:
            st.session_state.show_new_project_form = False
            st.rerun()



def financials_tab():
    """Financial data screen with monthly and cumulative cash flows"""
    
    # Add CSS for the financial table
    st.markdown("""
    <style>
    .financial-table {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Arial', 'Tahoma', sans-serif;
        font-size: 11px;
        direction: rtl;
        text-align: center;
        background-color: white;
    }
    .financial-table th, .financial-table td {
        border: 1px solid #ddd;
        padding: 4px 6px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        max-width: 120px;
        min-width: 90px;
    }
    .financial-table th {
        background-color: #4a90e2;
        color: white;
        font-weight: bold;
        position: sticky;
        top: 0;
        z-index: 10;
    }
    .financial-table .project-header {
        background-color: #f8f9fa;
        font-weight: bold;
        color: #333;
        text-align: right;
        padding: 8px 10px;
        min-width: 250px;
        max-width: 300px;
        position: sticky;
        right: 80px;
        z-index: 5;
        border-left: 2px solid #4a90e2;
        word-wrap: break-word;
        white-space: normal;
    }
    .financial-table .purchase-order {
        background-color: #e3f2fd !important;
        font-weight: bold;
        color: #1976d2 !important;
        text-align: center !important;
        padding: 8px 4px;
        min-width: 60px;
        max-width: 80px;
        position: sticky;
        right: 0;
        z-index: 6;
        border-left: 2px solid #4a90e2;
    }
    .financial-table tr:nth-child(even) {
        background-color: #f9f9f9;
    }
    .financial-table tr:hover {
        background-color: #e3f2fd;
    }
    .financial-table .amount {
        text-align: center;
        font-family: 'Courier New', monospace;
        font-size: 10px;
        font-weight: 500;
        min-width: 100px;
        max-width: 140px;
        white-space: nowrap;
        padding: 4px 8px;
    }
    .financial-table .positive-amount {
        color: #2e7d32;
        background-color: #e8f5e8;
    }
    .financial-table .zero-amount {
        color: #666;
        background-color: #f5f5f5;
    }
    .financial-table .project-desc {
        font-size: 10px;
        color: #666;
        font-style: italic;
    }
    .table-container {
        max-height: 600px;
        overflow: auto;
        border: 1px solid #ddd;
        border-radius: 4px;
    }
    .cash-flow-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 10px;
        border-radius: 8px;
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Date range selector at the top
    st.markdown("""
    <div class="cash-flow-header">
        <h2>💰 التدفقات النقدية - Cash Flows</h2>
    </div>
    """, unsafe_allow_html=True)
    
    # Use projects selected from sidebar
    all_projects = st.session_state.data_manager.get_all_projects()
    
    if not all_projects:
        st.warning("لا توجد مشاريع متاحة. يرجى إضافة مشاريع أولاً.")
        return
        
    # Get selected projects from sidebar
    selected_project_names = st.session_state.get('selected_projects', [])
    
    if not selected_project_names:
        st.warning("يرجى اختيار مشروع واحد على الأقل من القائمة الجانبية.")
        return
        
    selected_projects = [proj for proj in all_projects if proj['project_name'] in selected_project_names]
    
    if not selected_projects:
        st.warning("المشاريع المختارة غير متاحة.")
        return
    
    st.markdown(f"**المشاريع المختارة:** {len(selected_projects)} مشروع")
    
    # Date range selection
    st.markdown("### 📅 اختيار المدى الزمني")
    date_cols = st.columns(2)
    
    with date_cols[0]:
        # Get default start date from earliest project start among selected projects
        default_start = date.today().replace(day=1)
        if selected_projects:
            try:
                earliest_start_dates = []
                for project in selected_projects:
                    if project.get('start_date'):
                        project_start = pd.to_datetime(project['start_date']).date()
                        earliest_start_dates.append(project_start)
                
                if earliest_start_dates:
                    default_start = min(earliest_start_dates)
            except:
                pass
                
        date_range_start = st.date_input(
            "📅 من تاريخ",
            value=default_start,
            key="cash_flow_date_start"
        )
    
    with date_cols[1]:
        # Get default end date from latest project end among selected projects
        default_end = date.today()
        if selected_projects:
            try:
                latest_end_dates = []
                for project in selected_projects:
                    if project.get('end_date'):
                        project_end = pd.to_datetime(project['end_date']).date()
                        latest_end_dates.append(project_end)
                
                if latest_end_dates:
                    default_end = max(latest_end_dates)
            except:
                pass
                
        date_range_end = st.date_input(
            "📅 إلى تاريخ", 
            value=default_end,
            key="cash_flow_date_end"
        )
    
    # Use selected projects instead of all projects
    
    # Divider
    st.divider()
    
    # SECTION 1: Monthly Cash Flows (Non-Cumulative) - Row 7 Data
    st.markdown("""
    <div class="cash-flow-header">
        <h3>📊 التدفقات النقدية الشهرية (غير التراكمية) - Monthly Cash Flows</h3>
        <p style="font-size: 12px; margin: 5px 0;">البيانات من الصف 7: Planned Total Cost</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Generate monthly columns based on selected date range
    monthly_columns = generate_date_columns(date_range_start, date_range_end, "Monthly")
    display_monthly = monthly_columns  # Show all columns in the selected date range
    
    if len(monthly_columns) > 24:
        st.info(f"عرض {len(monthly_columns)} شهر في النطاق المحدد من {date_range_start} إلى {date_range_end}. قد يتطلب التمرير الأفقي.")
    
    # Build Monthly Cash Flow table
    monthly_table_html = '<div class="table-container">'
    monthly_table_html += '<table class="financial-table">'
    
    # Table header
    monthly_table_html += '<thead><tr>'
    monthly_table_html += '<th class="project-header" style="min-width: 80px; background-color: #e3f2fd; color: #1976d2;">أمر الشراء</th>'
    monthly_table_html += '<th class="project-header">اسم المشروع / الوصف</th>'
    
    for date_col in display_monthly:
        formatted_date = pd.to_datetime(date_col + '-01').strftime('%m/%Y')
        monthly_table_html += f'<th style="min-width: 100px;">{formatted_date}</th>'
    
    monthly_table_html += '</tr></thead>'
    monthly_table_html += '<tbody>'
    
    # Monthly data rows
    for project in selected_projects:
        project_name = project['project_name']
        project_desc = project.get('project_description', '')
        total_budget = project.get('total_budget', 0)
        
        # Get progress data (Excel imported data)
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        # Filter by date range
        if not progress_data.empty:
            progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
            progress_data = progress_data[
                (progress_data['entry_date'].dt.date >= date_range_start) &
                (progress_data['entry_date'].dt.date <= date_range_end)
            ]
        
        monthly_table_html += '<tr>'
        
        # Purchase Order column
        project_po = project.get('purchase_order', project.get('project_id', ''))
        po_cell = f'<td class="purchase-order">{project_po}</td>'
        monthly_table_html += po_cell
        
        # Project name column
        budget_formatted = f"{total_budget:,.2f}" if total_budget > 0 else "-"
        project_cell = f"""
        <td class="project-header" style="word-wrap: break-word; white-space: normal;">
            <div><strong>{project_name}</strong></div>
            <div class="project-desc">{project_desc[:50] + '...' if len(project_desc) > 50 else project_desc}</div>
            <div class="project-desc">الميزانية: {budget_formatted}</div>
        </td>
        """
        monthly_table_html += project_cell
        
        # Monthly financial data (using Row 7 - Planned Total Cost for intervals)
        for date_col in display_monthly:
            # Check if the month is before project start or after project end
            period_date = pd.to_datetime(date_col + '-01')
            project_end = get_project_end_date(project_name)
            project_start = get_project_start_date(project_name)
            
            # Check if period is before project start
            if project_start and period_date.replace(day=1) < project_start.replace(day=1):
                formatted_value = "قبل بداية المشروع"
                cell_class = "amount"
                style = "color: #888888; font-style: italic; font-size: 10px;"
                monthly_table_html += f'<td class="{cell_class}" style="{style}">{formatted_value}</td>'
            # Check if period is after project end
            elif project_end and period_date.replace(day=1) > project_end.replace(day=1):
                formatted_value = "خارج مدة المشروع"
                cell_class = "amount"
                style = "color: #ff6b6b; font-style: italic; font-size: 10px;"
                monthly_table_html += f'<td class="{cell_class}" style="{style}">{formatted_value}</td>'
            else:
                financial_value = get_financial_data_for_date(
                    progress_data, date_col, "Interval flows", "Monthly"
                )
                
                if financial_value and financial_value > 0:
                    formatted_value = f"{financial_value:,.2f}"
                    cell_class = "amount positive-amount"
                else:
                    formatted_value = "–"
                    cell_class = "amount zero-amount"
                
                monthly_table_html += f'<td class="{cell_class}">{formatted_value}</td>'
        
        monthly_table_html += '</tr>'
    
    monthly_table_html += '</tbody></table></div>'
    
    # Display monthly table
    st.markdown(monthly_table_html, unsafe_allow_html=True)
    
    # Export section for monthly data
    st.markdown("#### 📤 تصدير البيانات الشهرية")
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        excel_data_monthly = create_financial_excel_export(
            selected_projects, monthly_columns, "Interval flows", "Monthly"
        )
        if excel_data_monthly:
            st.download_button(
                label="📊 تحميل البيانات الشهرية",
                data=excel_data_monthly,
                file_name=f"monthly_cash_flows_{date_range_start}_{date_range_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_monthly",
                use_container_width=True
            )
    
    with col2:
        if excel_data_monthly:
            st.success(f"✅ جاهز للتحميل ({len(selected_projects)} مشروع)")
        else:
            st.error("❌ خطأ في إعداد ملف التصدير")
    
    # Divider between sections
    st.divider()
    
    # SECTION 2: Cumulative Cash Flows - Row 8 Data  
    st.markdown("""
    <div class="cash-flow-header">
        <h3>📈 التدفقات النقدية التراكمية - Cumulative Cash Flows</h3>
        <p style="font-size: 12px; margin: 5px 0;">البيانات من الصف 8: Cum Budgeted Total Cost</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Build Cumulative Cash Flow table (same structure, different data)
    cumulative_table_html = '<div class="table-container">'
    cumulative_table_html += '<table class="financial-table">'
    
    # Table header
    cumulative_table_html += '<thead><tr>'
    cumulative_table_html += '<th class="project-header" style="min-width: 80px; background-color: #e3f2fd; color: #1976d2;">أمر الشراء</th>'
    cumulative_table_html += '<th class="project-header">اسم المشروع / الوصف</th>'
    
    for date_col in display_monthly:
        formatted_date = pd.to_datetime(date_col + '-01').strftime('%m/%Y')
        cumulative_table_html += f'<th style="min-width: 100px;">{formatted_date}</th>'
    
    cumulative_table_html += '</tr></thead>'
    cumulative_table_html += '<tbody>'
    
    # Cumulative data rows
    for project in selected_projects:
        project_name = project['project_name']
        project_desc = project.get('project_description', '')
        total_budget = project.get('total_budget', 0)
        
        # Get progress data
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        # Filter by date range
        if not progress_data.empty:
            progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
            progress_data = progress_data[
                (progress_data['entry_date'].dt.date >= date_range_start) &
                (progress_data['entry_date'].dt.date <= date_range_end)
            ]
        
        cumulative_table_html += '<tr>'
        
        # Purchase Order column
        project_po = project.get('purchase_order', project.get('project_id', ''))
        po_cell = f'<td class="purchase-order">{project_po}</td>'
        cumulative_table_html += po_cell
        
        # Project name column
        budget_formatted = f"{total_budget:,.2f}" if total_budget > 0 else "-"
        project_cell = f"""
        <td class="project-header" style="word-wrap: break-word; white-space: normal;">
            <div><strong>{project_name}</strong></div>
            <div class="project-desc">{project_desc[:50] + '...' if len(project_desc) > 50 else project_desc}</div>
            <div class="project-desc">الميزانية: {budget_formatted}</div>
        </td>
        """
        cumulative_table_html += project_cell
        
        # Cumulative financial data (using Row 8 - Cumulative Budgeted Cost)
        for date_col in display_monthly:
            # Check if the month is before project start or after project end
            period_date = pd.to_datetime(date_col + '-01')
            project_end = get_project_end_date(project_name)
            project_start = get_project_start_date(project_name)
            
            # Check if period is before project start
            if project_start and period_date.replace(day=1) < project_start.replace(day=1):
                formatted_value = "قبل بداية المشروع"
                cell_class = "amount"
                style = "color: #888888; font-style: italic; font-size: 10px;"
                cumulative_table_html += f'<td class="{cell_class}" style="{style}">{formatted_value}</td>'
            # Check if period is after project end
            elif project_end and period_date.replace(day=1) > project_end.replace(day=1):
                formatted_value = "خارج مدة المشروع"
                cell_class = "amount"
                style = "color: #ff6b6b; font-style: italic; font-size: 10px;"
                cumulative_table_html += f'<td class="{cell_class}" style="{style}">{formatted_value}</td>'
            else:
                financial_value = get_financial_data_for_date(
                    progress_data, date_col, "Cumulative flows", "Monthly"
                )
                
                if financial_value and financial_value > 0:
                    formatted_value = f"{financial_value:,.2f}"
                    cell_class = "amount positive-amount"
                else:
                    formatted_value = "–"
                    cell_class = "amount zero-amount"
                
                cumulative_table_html += f'<td class="{cell_class}">{formatted_value}</td>'
        
        cumulative_table_html += '</tr>'
    
    cumulative_table_html += '</tbody></table></div>'
    
    # Display cumulative table
    st.markdown(cumulative_table_html, unsafe_allow_html=True)
    
    # Export section for cumulative data
    st.markdown("#### 📤 تصدير البيانات التراكمية")
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        excel_data_cumulative = create_financial_excel_export(
            selected_projects, monthly_columns, "Cumulative flows", "Monthly"
        )
        if excel_data_cumulative:
            st.download_button(
                label="📈 تحميل البيانات التراكمية",
                data=excel_data_cumulative,
                file_name=f"cumulative_cash_flows_{date_range_start}_{date_range_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_cumulative",
                use_container_width=True
            )
    
    with col2:
        if excel_data_cumulative:
            st.success(f"✅ جاهز للتحميل ({len(selected_projects)} مشروع)")
        else:
            st.error("❌ خطأ في إعداد ملف التصدير")
    
    # Combined export section
    st.markdown("#### 📋 تصدير شامل")
    col_combined1, col_combined2 = st.columns([1, 1])
    
    with col_combined1:
        # Create combined Excel with both monthly and cumulative data
        if st.button("📊 تصدير جميع البيانات", key="export_all", use_container_width=True):
            combined_excel = create_combined_financial_export(
                selected_projects, monthly_columns, date_range_start, date_range_end
            )
            if combined_excel:
                st.download_button(
                    label="📥 تحميل التقرير الشامل",
                    data=combined_excel,
                    file_name=f"financial_report_complete_{date_range_start}_{date_range_end}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_combined"
                )
    
    # Show debugging info in expander
    with st.expander("🔍 معلومات التصحيح (لتشخيص المشاكل)"):
        st.write(f"عدد المشاريع: {len(selected_projects)}")
        st.write(f"عدد أعمدة التاريخ: {len(monthly_columns)}")
        st.write(f"نطاق التاريخ: {date_range_start} إلى {date_range_end}")
        
        # Show sample progress data for first project
        if selected_projects:
            sample_project = selected_projects[0]['project_name']
            sample_data = st.session_state.data_manager.get_progress_data(sample_project)
            st.write(f"بيانات عينة من مشروع '{sample_project}': {len(sample_data)} صف")
            if not sample_data.empty:
                st.dataframe(sample_data.head())

def show_detailed_financial_analysis(all_projects, display_columns, data_type, flow_type, date_range_start, date_range_end):
    """Show detailed financial analysis with additional metrics"""
    st.markdown("### 📊 التحليل المالي التفصيلي")
    
    # Calculate summary statistics
    total_projects = len(all_projects)
    total_budget = sum(project.get('total_budget', 0) for project in all_projects)
    
    # Create summary cards
    summary_cols = st.columns(4)
    
    with summary_cols[0]:
        st.metric("عدد المشاريع", total_projects)
    
    with summary_cols[1]:
        st.metric("إجمالي الميزانية", f"{total_budget:,.2f}")
    
    with summary_cols[2]:
        # Calculate total actual spending using improved cash flow data
        total_actual = 0
        for project in all_projects:
            cash_flow_data = st.session_state.data_manager.get_cash_flow_data(
                project['project_name'], date_range_start, date_range_end
            )
            if not cash_flow_data.empty:
                total_actual += cash_flow_data['actual_cost'].sum()
        st.metric("إجمالي الإنفاق الفعلي", f"{total_actual:,.2f}")
    
    with summary_cols[3]:
        if total_budget > 0:
            utilization = (total_actual / total_budget) * 100
            st.metric("نسبة الاستخدام", f"{utilization:.1f}%")
        else:
            st.metric("نسبة الاستخدام", "0%")
    
    # Project performance analysis
    st.markdown("#### 📈 تحليل أداء المشاريع")
    
    performance_data = []
    for project in all_projects:
        project_name = project['project_name']
        total_budget = project.get('total_budget', 0)
        cash_flow_data = st.session_state.data_manager.get_cash_flow_data(
            project_name, date_range_start, date_range_end
        )
        
        if not cash_flow_data.empty:
            actual_spent = cash_flow_data['actual_cost'].sum()
            planned_spent = cash_flow_data['planned_cost'].sum()
            
            performance_data.append({
                'المشروع': project_name,
                'الميزانية المخططة': total_budget,
                'الإنفاق الفعلي': actual_spent,
                'الإنفاق المخطط': planned_spent,
                'نسبة التنفيذ': f"{(actual_spent/total_budget*100):.1f}%" if total_budget > 0 else "0%"
            })
    
    if performance_data:
        df_performance = pd.DataFrame(performance_data)
        st.dataframe(df_performance, use_container_width=True)
    
    # Timeline analysis
    st.markdown("#### 📅 تحليل الجدول الزمني")
    
    if display_columns:
        timeline_data = {}
        for date_col in display_columns[:10]:  # Show first 10 for performance
            total_for_date = 0
            for project in all_projects:
                cash_flow_data = st.session_state.data_manager.get_cash_flow_data(
                    project['project_name'], date_range_start, date_range_end
                )
                value = get_financial_data_for_date(cash_flow_data, date_col, data_type, flow_type)
                if value:
                    total_for_date += value
            
            if flow_type == "Monthly":
                formatted_date = pd.to_datetime(date_col + '-01').strftime('%Y-%m')
            elif flow_type == "Daily":
                formatted_date = pd.to_datetime(date_col).strftime('%Y-%m-%d')
            else:
                formatted_date = date_col
                
            timeline_data[formatted_date] = total_for_date
        
        if timeline_data:
            timeline_df = pd.DataFrame(list(timeline_data.items()), columns=['التاريخ', 'إجمالي القيمة'])
            st.line_chart(timeline_df.set_index('التاريخ'))

def generate_date_columns(start_date, end_date, flow_type):
    """Generate date columns based on flow type and date range"""
    date_columns = []
    current_date = start_date
    
    if flow_type == "Daily":
        while current_date <= end_date:
            date_columns.append(current_date.strftime('%Y-%m-%d'))
            current_date = current_date.replace(day=current_date.day + 1) if current_date.day < 28 else current_date.replace(month=current_date.month + 1, day=1) if current_date.month < 12 else current_date.replace(year=current_date.year + 1, month=1, day=1)
    elif flow_type == "Monthly":
        while current_date <= end_date:
            date_columns.append(current_date.strftime('%Y-%m'))
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
    elif flow_type == "Yearly":
        while current_date.year <= end_date.year:
            date_columns.append(str(current_date.year))
            current_date = current_date.replace(year=current_date.year + 1)
    
    return date_columns

@st.cache_data(ttl=600)  # Cache for 10 minutes as date parsing is expensive
def parse_excel_maybe_date(value):
    """Parse a value that might be a date in various formats
    Returns a date object or None
    """
    from datetime import datetime, date, timedelta
    import pandas as pd
    
    if value is None:
        return None
    
    try:
        # Handle string dates like "2023-12-31"
        if isinstance(value, str):
            if value.strip():
                return datetime.strptime(value, '%Y-%m-%d').date()
            return None
        
        # Handle Excel numeric dates (days since 1899-12-30)
        elif isinstance(value, (int, float)):
            if value > 0:
                # Excel uses 1900-01-01 as day 1, but with a leap year bug
                # So we use 1899-12-30 as base and add the days
                excel_base = date(1899, 12, 30)
                return excel_base + timedelta(days=int(value))
            return None
        
        # Handle pandas Timestamp
        elif hasattr(value, 'date'):
            return value.date()
        
        # Handle datetime.date directly
        elif isinstance(value, date):
            return value
        
        # Handle datetime.datetime
        elif isinstance(value, datetime):
            return value.date()
            
    except Exception as e:
        print(f"DEBUG - parse_excel_maybe_date: Failed to parse {value}: {e}")
    
    return None

@st.cache_data(ttl=600)  # Cache for 10 minutes as Excel parsing is expensive
def extract_excel_row_data(notes_str, row_number):
    """Extract data from notes field based on Excel row number (R7, R8, etc.) with improved error handling
    Special handling for date rows (17 and 20)"""
    if not notes_str or pd.isna(notes_str):
        return None
    
    try:
        # Parse notes format: R7:0|R8:0|R9:0.0|R10:0.0|R11:0.0013|R12:1.0|R13:0.0
        notes_str = str(notes_str).strip()
        row_key = f"R{row_number}:"
        
        if row_key in notes_str:
            # Find the value after row_key
            start_idx = notes_str.find(row_key) + len(row_key)
            end_idx = notes_str.find("|", start_idx)
            if end_idx == -1:
                end_idx = len(notes_str)
            
            value_str = notes_str[start_idx:end_idx].strip()
            
            # Handle empty or invalid values and convert common symbols to zero
            if not value_str or value_str.lower() in ['', 'null', 'none', 'nan']:
                return None
            
            # Special handling for date rows (17 and 20)
            if row_number in [17, 20]:
                # Check if it's a date string format (YYYY-MM-DD)
                if '-' in value_str and len(value_str) >= 10:
                    print(f"DEBUG - extract_excel_row_data: Found date string for R{row_number}: {value_str}")
                    return value_str  # Return date string as-is
                # Check if it's 0 (no date)
                elif value_str == '0' or value_str == '0.0':
                    return None
                # Try parsing as float (Excel date serial)
                elif value_str.replace('.', '').replace('-', '').isdigit():
                    try:
                        date_serial = float(value_str)
                        if date_serial > 40000:  # Valid Excel date range
                            print(f"DEBUG - extract_excel_row_data: Found Excel date serial for R{row_number}: {date_serial}")
                            return date_serial
                    except:
                        pass
                # If not a valid date, return None
                print(f"DEBUG - extract_excel_row_data: No valid date found for R{row_number}: {value_str}")
                return None
            
            # Clean and normalize the value string for non-date fields
            if isinstance(value_str, str):
                value_str = value_str.strip()
                
                # Convert common "empty" or "no data" symbols to zero
                empty_symbols = ['-', '—', '–', '_', 'n/a', 'na', 'غير متوفر', 'لا يوجد', 'فارغ']
                if value_str.lower() in empty_symbols or value_str in empty_symbols:
                    return 0.0  # Convert empty symbols to zero
                
                # Remove common text indicators and thousands separators
                value_str = value_str.replace(',', '').replace('٬', '').replace(' ', '')
                
                # Check for percentage
                if '%' in value_str:
                    try:
                        return float(value_str.replace('%', '').strip())
                    except:
                        return None
                
                # Handle Arabic numerals (convert to English)
                arabic_to_english = {
                    '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
                    '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9'
                }
                for arabic, english in arabic_to_english.items():
                    value_str = value_str.replace(arabic, english)
            
            # Try to convert to float
            try:
                result = float(value_str)
                # Validate result is a reasonable number
                if abs(result) > 1e15:  # Extremely large numbers are likely errors
                    print(f"DEBUG - extract_excel_row_data: Suspiciously large value {result} for R{row_number}")
                    return None
                return result
            except (ValueError, TypeError):
                print(f"DEBUG - extract_excel_row_data: Cannot convert '{value_str}' to number for R{row_number}")
                return None
        
        return None
    except Exception as e:
        print(f"DEBUG - extract_excel_row_data: Error parsing R{row_number}: {e}")
        return None

def get_progress_percentage_for_period(progress_data, period_start, period_end, row_number, is_cumulative=True):
    """Get progress percentage from Excel row data for a specific period
    
    Args:
        progress_data: DataFrame with progress data
        period_start: Start date of period
        period_end: End date of period  
        row_number: Excel row number (9=Planned Weekly %, 10=Planned Monthly %, 13=Actual %)
        is_cumulative: If True, get cumulative value at end of period. If False, get last value in period.
    """
    if progress_data.empty:
        return None
    
    try:
        # Ensure entry_date is datetime
        progress_data_copy = progress_data.copy()
        progress_data_copy['entry_date'] = pd.to_datetime(progress_data_copy['entry_date'])
        
        # Filter data within the period
        filtered_data = progress_data_copy[
            (progress_data_copy['entry_date'] >= pd.to_datetime(period_start)) &
            (progress_data_copy['entry_date'] <= pd.to_datetime(period_end))
        ]
        
        if filtered_data.empty:
            # If no data in period, get the last available data before period_end
            before_period = progress_data_copy[progress_data_copy['entry_date'] <= pd.to_datetime(period_end)]
            if not before_period.empty:
                last_row = before_period.iloc[-1]
                return extract_excel_row_data(last_row.get('notes', ''), row_number)
            return None
        
        # Get the last row in the period
        last_row = filtered_data.iloc[-1]
        return extract_excel_row_data(last_row.get('notes', ''), row_number)
        
    except Exception as e:
        return None

def get_max_progress_percentage_for_period(progress_data, period_start, period_end, row_number):
    """Get maximum progress percentage from Excel row data for a specific period
    Used specifically for actual progress (row 13) to get the highest achieved percentage in the period
    
    Args:
        progress_data: DataFrame with progress data
        period_start: Start date of period
        period_end: End date of period  
        row_number: Excel row number (typically 13 for actual progress)
    """
    if progress_data.empty:
        return None
    
    try:
        # Ensure entry_date is datetime
        progress_data_copy = progress_data.copy()
        progress_data_copy['entry_date'] = pd.to_datetime(progress_data_copy['entry_date'])
        
        # Filter data within the period
        filtered_data = progress_data_copy[
            (progress_data_copy['entry_date'] >= pd.to_datetime(period_start)) &
            (progress_data_copy['entry_date'] <= pd.to_datetime(period_end))
        ]
        
        if filtered_data.empty:
            # If no data in period, get the last available data before period_end
            before_period = progress_data_copy[progress_data_copy['entry_date'] <= pd.to_datetime(period_end)]
            if not before_period.empty:
                last_row = before_period.iloc[-1]
                return extract_excel_row_data(last_row.get('notes', ''), row_number)
            return None
        
        # Extract all values for the specified row in the period
        values = []
        for _, row in filtered_data.iterrows():
            value = extract_excel_row_data(row.get('notes', ''), row_number)
            if value is not None and value > 0:
                values.append(value)
        
        if values:
            return max(values)  # Return the maximum value found in the period
        else:
            # If no valid values found, return the last available value
            last_row = filtered_data.iloc[-1]
            return extract_excel_row_data(last_row.get('notes', ''), row_number)
        
    except Exception as e:
        return None

def get_project_end_date(project_name):
    """Get project end date"""
    try:
        projects_list = st.session_state.data_manager.get_all_projects()
        projects_df = pd.DataFrame(projects_list)
        
        if projects_df.empty:
            return None
            
        project = projects_df[projects_df['project_name'] == project_name]
        
        if project.empty:
            return None
            
        project_end = pd.to_datetime(project.iloc[0]['end_date'])
        return project_end
        
    except Exception as e:
        return None

def get_project_start_date(project_name):
    """Get project start date"""
    try:
        projects_list = st.session_state.data_manager.get_all_projects()
        projects_df = pd.DataFrame(projects_list)
        
        if projects_df.empty:
            return None
            
        project = projects_df[projects_df['project_name'] == project_name]
        
        if project.empty:
            return None
            
        project_start = pd.to_datetime(project.iloc[0]['start_date'])
        return project_start
        
    except Exception as e:
        return None

def is_date_beyond_project_end(project_name, check_date):
    """Check if a date is beyond the project end date"""
    try:
        projects_list = st.session_state.data_manager.get_all_projects()
        projects_df = pd.DataFrame(projects_list)
        
        if projects_df.empty:
            return False
            
        project = projects_df[projects_df['project_name'] == project_name]
        
        if project.empty:
            return False
            
        project_end = pd.to_datetime(project.iloc[0]['end_date'])
        check_dt = pd.to_datetime(check_date)
        
        
        return check_dt > project_end
        
    except Exception as e:
        return False

def calculate_elapsed_percentage_beyond_end_monthly(project_name, target_date):
    """Calculate elapsed percentage for monthly view when target date is beyond project end
    Formula: (Last day of month - Project start date) / (Project end date - Project start date) × 100
    """
    try:
        projects_list = st.session_state.data_manager.get_all_projects()
        projects_df = pd.DataFrame(projects_list)
        
        if projects_df.empty:
            return None
            
        project = projects_df[projects_df['project_name'] == project_name]
        
        if project.empty:
            return None
            
        project_start = pd.to_datetime(project.iloc[0]['start_date'])
        project_end = pd.to_datetime(project.iloc[0]['end_date'])
        target_dt = pd.to_datetime(target_date)
        
        # Calculate total project duration in days
        total_duration = (project_end - project_start).days
        
        # Calculate days from project start to target date (last day of month)
        elapsed_days = (target_dt - project_start).days
        
        # Calculate percentage using the specified formula
        if total_duration > 0:
            percentage = elapsed_days / total_duration
            return percentage
        else:
            return None
            
    except Exception as e:
        return None

def calculate_elapsed_percentage_beyond_end_weekly(project_name, thursday_date):
    """Calculate elapsed percentage for weekly view when Thursday date is beyond project end
    Formula: (Thursday date - Project start date) / (Project end date - Project start date) × 100
    """
    try:
        projects_list = st.session_state.data_manager.get_all_projects()
        projects_df = pd.DataFrame(projects_list)
        
        if projects_df.empty:
            return None
            
        project = projects_df[projects_df['project_name'] == project_name]
        
        if project.empty:
            return None
            
        project_start = pd.to_datetime(project.iloc[0]['start_date'])
        project_end = pd.to_datetime(project.iloc[0]['end_date'])
        thursday_dt = pd.to_datetime(thursday_date)
        
        # Calculate total project duration in days
        total_duration = (project_end - project_start).days
        
        # Calculate days from project start to Thursday date
        elapsed_days = (thursday_dt - project_start).days
        
        # Calculate percentage using the specified formula
        if total_duration > 0:
            percentage = elapsed_days / total_duration
            return percentage
        else:
            return None
            
    except Exception as e:
        return None

def generate_monthly_columns(start_date, end_date):
    """Generate monthly columns for progress tracking"""
    columns = []
    current_date = pd.to_datetime(start_date).replace(day=1)
    end_date = pd.to_datetime(end_date)
    
    while current_date <= end_date:
        # Format as "Month Year" in Arabic
        month_name = current_date.strftime('%B %Y')
        # Convert to Arabic month names
        month_mapping = {
            'January': 'يناير', 'February': 'فبراير', 'March': 'مارس',
            'April': 'أبريل', 'May': 'مايو', 'June': 'يونيو',
            'July': 'يوليو', 'August': 'أغسطس', 'September': 'سبتمبر',
            'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
        }
        
        for eng, ar in month_mapping.items():
            month_name = month_name.replace(eng, ar)
        
        columns.append({
            'date_key': current_date.strftime('%Y-%m'),
            'display_name': month_name,
            'start_date': current_date,
            'end_date': current_date + pd.offsets.MonthEnd(0)
        })
        
        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    
    return columns

def generate_weekly_columns(start_date, end_date):
    """Generate weekly columns based on Thursdays"""
    columns = []
    current_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)
    
    # Find the first Thursday
    days_ahead = 3 - current_date.weekday()  # Thursday is weekday 3
    if days_ahead < 0:  # Target day already happened this week
        days_ahead += 7
    
    first_thursday = current_date + pd.Timedelta(days=days_ahead)
    
    current_thursday = first_thursday
    while current_thursday <= end_date:
        week_start = current_thursday - pd.Timedelta(days=6)  # Monday
        week_end = current_thursday
        
        columns.append({
            'date_key': current_thursday.strftime('%Y-%m-%d'),
            'display_name': current_thursday.strftime('%d-%m-%Y'),
            'thursday_date': current_thursday,
            'week_start': week_start,
            'week_end': week_end,
            'month_year': current_thursday.strftime('%Y-%m')
        })
        
        current_thursday += pd.Timedelta(days=7)
    
    return columns

def get_financial_data_for_date(progress_data, date_col, data_type, flow_type):
    """Get financial data for a specific date column from imported Excel data
    
    Exact mapping as specified:
    - For Interval flows: Row 7 (Planned Total Cost) calculated as monthly intervals
    - For Cumulative flows: Row 8 (Cumulative Budgeted Cost) - last value up to date
    """
    
    if progress_data.empty:
        return 0
    
    try:
        # Convert date_col to datetime for comparison
        if flow_type == "Monthly":
            target_date = pd.to_datetime(date_col + '-01')
        elif flow_type == "Daily":
            target_date = pd.to_datetime(date_col)
        elif flow_type == "Yearly":
            target_date = pd.to_datetime(f"{date_col}-01-01")
        else:
            return 0
        
        # Ensure entry_date is datetime
        progress_data_copy = progress_data.copy()
        progress_data_copy['entry_date'] = pd.to_datetime(progress_data_copy['entry_date'])
        
        if data_type == "Cumulative flows":
            # Row 8: Cumulative Budgeted Cost - get last value up to end of target month
            if flow_type == "Monthly":
                # For cumulative flows, use end of month instead of beginning
                month_end_date = target_date + pd.offsets.MonthEnd(0)
                filtered_data = progress_data_copy[progress_data_copy['entry_date'] <= month_end_date]
            else:
                filtered_data = progress_data_copy[progress_data_copy['entry_date'] <= target_date]
            
            if not filtered_data.empty:
                last_row = filtered_data.iloc[-1]
                return extract_excel_row_data(last_row.get('notes', ''), 8)
            return 0
        else:  # Interval flows - Row 7: Planned Total Cost
            # Row 7: Sum all values directly from Row 7 within the month
            if flow_type == "Monthly":
                # Define month boundaries
                current_year = target_date.year
                current_month = target_date.month
                month_start = pd.Timestamp(current_year, current_month, 1)
                month_end = month_start + pd.offsets.MonthEnd(0)
                
                # Filter data for entries within the target month
                month_data = progress_data_copy[
                    (progress_data_copy['entry_date'] >= month_start) &
                    (progress_data_copy['entry_date'] <= month_end)
                ]
                
                # Sum all Row 7 values within the month
                total_monthly_value = 0
                if not month_data.empty:
                    for _, row in month_data.iterrows():
                        row_7_value = extract_excel_row_data(row.get('notes', ''), 7)
                        if row_7_value and row_7_value > 0:
                            total_monthly_value += row_7_value
                
                return max(0, total_monthly_value)  # Ensure non-negative values
                
            elif flow_type == "Daily":
                # For monthly intervals: get current month cumulative minus previous month cumulative
                current_year = target_date.year
                current_month = target_date.month
                
                # Get cumulative value up to end of current month
                current_month_end = pd.Timestamp(current_year, current_month, 1) + pd.offsets.MonthEnd(0)
                current_cumulative_data = progress_data_copy[progress_data_copy['entry_date'] <= current_month_end]
                current_cumulative = current_cumulative_data['planned_cost'].sum() if not current_cumulative_data.empty else 0
                
                # Get cumulative value up to end of previous month
                if current_month == 1:
                    previous_year = current_year - 1
                    previous_month = 12
                else:
                    previous_year = current_year
                    previous_month = current_month - 1
                    
                previous_month_end = pd.Timestamp(previous_year, previous_month, 1) + pd.offsets.MonthEnd(0)
                previous_cumulative_data = progress_data_copy[progress_data_copy['entry_date'] <= previous_month_end]
                previous_cumulative = previous_cumulative_data['planned_cost'].sum() if not previous_cumulative_data.empty else 0
                
                result = current_cumulative - previous_cumulative
                print(f"DEBUG - Monthly interval SUMIFS: Current cumulative ({current_cumulative}) - Previous cumulative ({previous_cumulative}) = {result}")
                
            elif flow_type == "Yearly":
                # For yearly intervals: get current year cumulative minus previous year cumulative
                current_year = target_date.year
                
                # Get cumulative value up to end of current year
                current_year_end = pd.Timestamp(current_year, 12, 31)
                current_cumulative_data = progress_data_copy[progress_data_copy['entry_date'] <= current_year_end]
                current_cumulative = current_cumulative_data['planned_cost'].sum() if not current_cumulative_data.empty else 0
                
                # Get cumulative value up to end of previous year
                previous_year_end = pd.Timestamp(current_year - 1, 12, 31)
                previous_cumulative_data = progress_data_copy[progress_data_copy['entry_date'] <= previous_year_end]
                previous_cumulative = previous_cumulative_data['planned_cost'].sum() if not previous_cumulative_data.empty else 0
                
                result = current_cumulative - previous_cumulative
                print(f"DEBUG - Yearly interval SUMIFS: Current cumulative ({current_cumulative}) - Previous cumulative ({previous_cumulative}) = {result}")
            
            return max(0, result) if 'result' in locals() else 0
    
    except Exception as e:
        return 0
    
    return 0

def create_financial_excel_export(all_projects, date_columns, data_type, flow_type):
    """Create Excel export with full financial data table and RTL support"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Data"
        
        # Set worksheet RTL direction
        ws.sheet_view.rightToLeft = True
        
        # Define styles matching the interface
        po_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        project_header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") 
        date_header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        
        po_font = Font(bold=True, color="1976D2", name="Arial", size=11)
        project_font = Font(bold=True, color="333333", name="Arial", size=11)
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        data_font = Font(name="Courier New", size=10)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # Headers with proper formatting
        col_idx = 1
        
        # PO Header
        po_cell = ws.cell(row=1, column=col_idx, value="أمر الشراء")
        po_cell.font = po_font
        po_cell.fill = po_header_fill
        po_cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1
        
        # Project Name Header  
        proj_cell = ws.cell(row=1, column=col_idx, value="اسم المشروع / الوصف")
        proj_cell.font = project_font
        proj_cell.fill = project_header_fill
        proj_cell.alignment = right_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 35
        col_idx += 1
        
        # Date Headers
        for date_col in date_columns:
            formatted_date = pd.to_datetime(date_col + '-01').strftime('%m/%Y')
            date_cell = ws.cell(row=1, column=col_idx, value=formatted_date)
            date_cell.font = header_font
            date_cell.fill = date_header_fill
            date_cell.alignment = center_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1
        
        # Data rows
        for row, project in enumerate(all_projects, 2):
            project_name = project['project_name']
            project_po = project.get('purchase_order', project.get('project_id', ''))
            project_desc = project.get('project_description', '')
            total_budget = project.get('total_budget', 0)
            progress_data = st.session_state.data_manager.get_progress_data(project_name)
            
            # Purchase Order column
            po_cell = ws.cell(row=row, column=1, value=project_po)
            po_cell.font = Font(bold=True, color="1976D2", name="Arial", size=10)
            po_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            po_cell.alignment = center_alignment
            
            # Project info column with description and budget
            budget_formatted = f"{total_budget:,.2f}" if total_budget > 0 else "-"
            project_info = f"{project_name}\n{project_desc[:50] + '...' if len(project_desc) > 50 else project_desc}\nالميزانية: {budget_formatted}"
            
            proj_cell = ws.cell(row=row, column=2, value=project_info)
            proj_cell.font = Font(bold=True, color="333333", name="Arial", size=10)
            proj_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            proj_cell.alignment = right_alignment
            
            # Financial data for each date
            for col, date_col in enumerate(date_columns, 3):
                # Check if the month is after the project end month
                period_date = pd.to_datetime(date_col + '-01')
                project_end = get_project_end_date(project_name)
                
                if project_end and period_date.replace(day=1) > project_end.replace(day=1):
                    display_value = "خارج مدة المشروع"
                    cell_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                    font_color = "FF6B6B"
                else:
                    financial_value = get_financial_data_for_date(
                        progress_data, date_col, data_type, flow_type
                    )
                    
                    if financial_value and financial_value > 0:
                        display_value = f"{financial_value:,.2f}"
                        cell_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        font_color = "2E7D32"
                    else:
                        display_value = "–"
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        font_color = "666666"
                
                data_cell = ws.cell(row=row, column=col, value=display_value)
                data_cell.font = Font(color=font_color, name="Courier New", size=10, italic=(display_value == "خارج مدة المشروع"))
                data_cell.fill = cell_fill
                data_cell.alignment = center_alignment
        
        # Freeze panes to keep PO and project name visible
        ws.freeze_panes = ws['C2']  # Freeze first two columns
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust row heights for wrapped text
        for row_num in range(2, len(all_projects) + 2):
            ws.row_dimensions[row_num].height = 60
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    except Exception as e:
        st.error(f"Error creating Excel export: {e}")
        return None

def create_combined_financial_export(all_projects, date_columns, date_start, date_end):
    """Create combined Excel export with both monthly and cumulative data"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Monthly Data sheet
        ws_monthly = wb.create_sheet("Monthly Data")
        ws_monthly.title = "Monthly Data"
        
        # Set RTL direction for monthly sheet
        ws_monthly.sheet_view.rightToLeft = True
        
        # Define styles
        po_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        project_header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        date_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        po_font = Font(bold=True, color="1976D2", name="Arial", size=11)
        project_font = Font(bold=True, color="333333", name="Arial", size=11)
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # Headers with proper formatting
        col_idx = 1
        
        # PO Header
        po_cell = ws_monthly.cell(row=1, column=col_idx, value="أمر الشراء")
        po_cell.font = po_font
        po_cell.fill = po_header_fill
        po_cell.alignment = center_alignment
        ws_monthly.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1
        
        # Project Name Header
        proj_cell = ws_monthly.cell(row=1, column=col_idx, value="اسم المشروع / الوصف")
        proj_cell.font = project_font
        proj_cell.fill = project_header_fill
        proj_cell.alignment = right_alignment
        ws_monthly.column_dimensions[get_column_letter(col_idx)].width = 35
        col_idx += 1
        
        # Date Headers
        for date_col in date_columns:
            formatted_date = pd.to_datetime(date_col + '-01').strftime('%m/%Y')
            date_cell = ws_monthly.cell(row=1, column=col_idx, value=formatted_date)
            date_cell.font = header_font
            date_cell.fill = date_header_fill
            date_cell.alignment = center_alignment
            ws_monthly.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1
        
        # Monthly data rows
        for row, project in enumerate(all_projects, 2):
            project_name = project['project_name']
            project_po = project.get('purchase_order', project.get('project_id', ''))
            project_desc = project.get('project_description', '')
            total_budget = project.get('total_budget', 0)
            progress_data = st.session_state.data_manager.get_progress_data(project_name)
            
            # PO Column
            po_cell = ws_monthly.cell(row=row, column=1, value=project_po)
            po_cell.font = Font(bold=True, color="1976D2", name="Arial", size=10)
            po_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            po_cell.alignment = center_alignment
            
            # Project info with description and budget
            budget_formatted = f"{total_budget:,.2f}" if total_budget > 0 else "-"
            project_info = f"{project_name}\n{project_desc[:50] + '...' if len(project_desc) > 50 else project_desc}\nالميزانية: {budget_formatted}"
            
            proj_cell = ws_monthly.cell(row=row, column=2, value=project_info)
            proj_cell.font = Font(bold=True, color="333333", name="Arial", size=10)
            proj_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            proj_cell.alignment = right_alignment
            
            for col, date_col in enumerate(date_columns, 3):
                # Check if the month is after the project end month
                period_date = pd.to_datetime(date_col + '-01')
                project_end = get_project_end_date(project_name)
                
                if project_end and period_date.replace(day=1) > project_end.replace(day=1):
                    display_value = "خارج مدة المشروع"
                    cell_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                    font_color = "FF6B6B"
                else:
                    financial_value = get_financial_data_for_date(
                        progress_data, date_col, "Interval flows", "Monthly"
                    )
                    
                    if financial_value and financial_value > 0:
                        display_value = f"{financial_value:,.2f}"
                        cell_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        font_color = "2E7D32"
                    else:
                        display_value = "–"
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        font_color = "666666"
                
                data_cell = ws_monthly.cell(row=row, column=col, value=display_value)
                data_cell.font = Font(color=font_color, name="Courier New", size=10, italic=(display_value == "خارج مدة المشروع"))
                data_cell.fill = cell_fill
                data_cell.alignment = center_alignment
        
        # Auto-adjust row heights
        for row_num in range(2, len(all_projects) + 2):
            ws_monthly.row_dimensions[row_num].height = 60
        
        # Freeze panes
        ws_monthly.freeze_panes = ws_monthly['C2']
        
        # Create Cumulative Data sheet
        ws_cumulative = wb.create_sheet("Cumulative Data")
        
        # Set RTL direction for cumulative sheet
        ws_cumulative.sheet_view.rightToLeft = True
        
        # Cumulative data headers with green theme
        cumulative_header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        
        col_idx = 1
        
        # PO Header
        po_cell = ws_cumulative.cell(row=1, column=col_idx, value="أمر الشراء")
        po_cell.font = po_font
        po_cell.fill = po_header_fill
        po_cell.alignment = center_alignment
        ws_cumulative.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1
        
        # Project Name Header
        proj_cell = ws_cumulative.cell(row=1, column=col_idx, value="اسم المشروع / الوصف")
        proj_cell.font = project_font
        proj_cell.fill = project_header_fill
        proj_cell.alignment = right_alignment
        ws_cumulative.column_dimensions[get_column_letter(col_idx)].width = 35
        col_idx += 1
        
        # Date Headers with cumulative theme
        for date_col in date_columns:
            formatted_date = pd.to_datetime(date_col + '-01').strftime('%m/%Y')
            date_cell = ws_cumulative.cell(row=1, column=col_idx, value=formatted_date)
            date_cell.font = header_font
            date_cell.fill = cumulative_header_fill
            date_cell.alignment = center_alignment
            ws_cumulative.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1
        
        # Cumulative data rows
        for row, project in enumerate(all_projects, 2):
            project_name = project['project_name']
            project_po = project.get('purchase_order', project.get('project_id', ''))
            project_desc = project.get('project_description', '')
            total_budget = project.get('total_budget', 0)
            progress_data = st.session_state.data_manager.get_progress_data(project_name)
            
            # PO Column
            po_cell = ws_cumulative.cell(row=row, column=1, value=project_po)
            po_cell.font = Font(bold=True, color="1976D2", name="Arial", size=10)
            po_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            po_cell.alignment = center_alignment
            
            # Project info with description and budget
            budget_formatted = f"{total_budget:,.2f}" if total_budget > 0 else "-"
            project_info = f"{project_name}\n{project_desc[:50] + '...' if len(project_desc) > 50 else project_desc}\nالميزانية: {budget_formatted}"
            
            proj_cell = ws_cumulative.cell(row=row, column=2, value=project_info)
            proj_cell.font = Font(bold=True, color="333333", name="Arial", size=10)
            proj_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            proj_cell.alignment = right_alignment
            
            for col, date_col in enumerate(date_columns, 3):
                # Check if the month is after the project end month
                period_date = pd.to_datetime(date_col + '-01')
                project_end = get_project_end_date(project_name)
                
                if project_end and period_date.replace(day=1) > project_end.replace(day=1):
                    display_value = "خارج مدة المشروع"
                    cell_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                    font_color = "FF6B6B"
                else:
                    financial_value = get_financial_data_for_date(
                        progress_data, date_col, "Cumulative flows", "Monthly"
                    )
                    
                    if financial_value and financial_value > 0:
                        display_value = f"{financial_value:,.2f}"
                        cell_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        font_color = "2E7D32"
                    else:
                        display_value = "–"
                        cell_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        font_color = "666666"
                
                data_cell = ws_cumulative.cell(row=row, column=col, value=display_value)
                data_cell.font = Font(color=font_color, name="Courier New", size=10, italic=(display_value == "خارج مدة المشروع"))
                data_cell.fill = cell_fill
                data_cell.alignment = center_alignment
        
        # Auto-adjust row heights
        for row_num in range(2, len(all_projects) + 2):
            ws_cumulative.row_dimensions[row_num].height = 60
        
        # Freeze panes
        ws_cumulative.freeze_panes = ws_cumulative['C2']
        
        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
        ws_summary.sheet_view.rightToLeft = True
        
        # Summary information in Arabic
        summary_title = ws_summary.cell(row=1, column=1, value="ملخص التقرير المالي")
        summary_title.font = Font(bold=True, size=16, name="Arial")
        summary_title.alignment = Alignment(horizontal="right", vertical="center")
        
        # Total budget calculation
        total_budget = sum(project.get('total_budget', 0) for project in all_projects)
        
        ws_summary.cell(row=3, column=1, value="المدى الزمني:").font = Font(bold=True, name="Arial", size=11)
        ws_summary.cell(row=3, column=1).alignment = Alignment(horizontal="right")
        ws_summary.cell(row=3, column=2, value=f"{date_start} إلى {date_end}").font = Font(name="Arial", size=11)
        ws_summary.cell(row=3, column=2).alignment = Alignment(horizontal="right")
        
        ws_summary.cell(row=4, column=1, value="عدد المشاريع:").font = Font(bold=True, name="Arial", size=11)
        ws_summary.cell(row=4, column=1).alignment = Alignment(horizontal="right")
        ws_summary.cell(row=4, column=2, value=len(all_projects)).font = Font(name="Arial", size=11)
        ws_summary.cell(row=4, column=2).alignment = Alignment(horizontal="center")
        
        ws_summary.cell(row=5, column=1, value="إجمالي الميزانية:").font = Font(bold=True, name="Arial", size=11)
        ws_summary.cell(row=5, column=1).alignment = Alignment(horizontal="right")
        ws_summary.cell(row=5, column=2, value=f"{total_budget:,.2f}").font = Font(name="Arial", size=11)
        ws_summary.cell(row=5, column=2).alignment = Alignment(horizontal="center")
        
        ws_summary.cell(row=6, column=1, value="تاريخ الإنشاء:").font = Font(bold=True, name="Arial", size=11)
        ws_summary.cell(row=6, column=1).alignment = Alignment(horizontal="right")
        ws_summary.cell(row=6, column=2, value=pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')).font = Font(name="Arial", size=11)
        ws_summary.cell(row=6, column=2).alignment = Alignment(horizontal="center")
        
        # Project list
        projects_title = ws_summary.cell(row=8, column=1, value="المشاريع المشمولة:")
        projects_title.font = Font(bold=True, name="Arial", size=11)
        projects_title.alignment = Alignment(horizontal="right")
        
        for idx, project in enumerate(all_projects, 9):
            project_po = project.get('purchase_order', project.get('project_id', ''))
            ws_summary.cell(row=idx, column=1, value=f"{idx-8}.").font = Font(name="Arial", size=10)
            ws_summary.cell(row=idx, column=1).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=idx, column=2, value=project_po).font = Font(bold=True, color="1976D2", name="Arial", size=10)
            ws_summary.cell(row=idx, column=2).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=idx, column=3, value=project['project_name']).font = Font(name="Arial", size=10)
            ws_summary.cell(row=idx, column=3).alignment = Alignment(horizontal="right", wrap_text=True)
        
        # Add borders to all sheets
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for sheet in [ws_monthly, ws_cumulative]:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    except Exception as e:
        st.error(f"Error creating combined Excel export: {e}")
        return None


def get_workforce_count_for_period(progress_data, start_date, end_date):
    """Calculate workforce count from progress data for a given period"""
    try:
        if progress_data.empty:
            return None
            
        # Filter data for the period
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        period_data = progress_data[
            (progress_data['entry_date'] >= pd.to_datetime(start_date)) &
            (progress_data['entry_date'] <= pd.to_datetime(end_date))
        ]
        
        if period_data.empty:
            return None
        
        # Extract workforce count from notes field (R12 represents workforce count)
        workforce_values = []
        for _, row in period_data.iterrows():
            workforce_count = extract_excel_row_data(row.get('notes', ''), 12)
            if workforce_count > 0:
                workforce_values.append(workforce_count)
        
        if workforce_values:
            return sum(workforce_values) / len(workforce_values)  # Average workforce
        return None
    except Exception as e:
        print(f"Error calculating workforce count: {e}")
        return None


def get_equipment_count_for_period(progress_data, start_date, end_date):
    """Calculate equipment count from progress data for a given period"""
    try:
        if progress_data.empty:
            return None
            
        # Filter data for the period
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        period_data = progress_data[
            (progress_data['entry_date'] >= pd.to_datetime(start_date)) &
            (progress_data['entry_date'] <= pd.to_datetime(end_date))
        ]
        
        if period_data.empty:
            return None
        
        # Estimate equipment count based on workforce (typical ratio 1:5 equipment to workforce)
        workforce_values = []
        for _, row in period_data.iterrows():
            workforce_count = extract_excel_row_data(row.get('notes', ''), 12)
            if workforce_count > 0:
                workforce_values.append(workforce_count)
        
        if workforce_values:
            avg_workforce = sum(workforce_values) / len(workforce_values)
            return avg_workforce / 5  # Estimated equipment count (1 equipment per 5 workers)
        return None
    except Exception as e:
        print(f"Error calculating equipment count: {e}")
        return None


def get_elapsed_time_for_period(progress_data, start_date, end_date):
    """Calculate elapsed time from progress data for a given period"""
    try:
        if progress_data.empty:
            return None
            
        # Filter data for the period
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        period_data = progress_data[
            (progress_data['entry_date'] >= pd.to_datetime(start_date)) &
            (progress_data['entry_date'] <= pd.to_datetime(end_date))
        ]
        
        if period_data.empty:
            return None
        
        # Extract elapsed time from notes field (R11 represents elapsed percentage)
        elapsed_values = []
        for _, row in period_data.iterrows():
            elapsed_percentage = extract_excel_row_data(row.get('notes', ''), 11)
            if elapsed_percentage > 0:
                # Convert percentage to days (assume project duration of 1000 days for calculation)
                elapsed_days = elapsed_percentage * 1000
                elapsed_values.append(elapsed_days)
        
        if elapsed_values:
            return sum(elapsed_values) / len(elapsed_values)  # Average elapsed time
        return None
    except Exception as e:
        print(f"Error calculating elapsed time: {e}")
        return None


# New functions for monthly resource tracking (using rows 20, 21, 22)
def get_monthly_manpower_count_for_period(progress_data, start_date, end_date):
    """Calculate monthly manpower count from progress data for a given period (row 21)
    Uses R20 as date reference, fallback to entry_date since R20/R21 are often zero"""
    try:
        if progress_data.empty:
            print(f"DEBUG - Monthly manpower: No progress data available for period {start_date} to {end_date}")
            return None
            
        # Convert period dates for comparison
        target_start = pd.to_datetime(start_date)
        target_end = pd.to_datetime(end_date)
        target_date = target_start  # Use start of period as target
        
        print(f"DEBUG - Monthly manpower: Looking for R21 values (monthly manpower per requirements)")
        
        # Find all data with R21 (monthly manpower) values
        all_data = progress_data[progress_data['notes'].str.contains('R21:', na=False)]
        
        if all_data.empty:
            print(f"DEBUG - Monthly manpower: No R21 data found in dataset")
            return None
        
        # Since R20/R21 are often zero, use entry_date with R21 values as per requirements
        print(f"DEBUG - Monthly manpower: R20 dates often zero, using entry_date fallback with R21 values from correct row")
        
        # Calculate distance from target date using entry_date
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        all_data['distance'] = abs((all_data['entry_date'] - target_date).dt.days)
        
        # Extract R21 values (monthly manpower as per requirements)
        valid_entries = []
        for _, row in all_data.iterrows():
            notes = row.get('notes', '')
            r21_value = extract_excel_row_data(notes, 21)  # Row 21 = monthly manpower per requirements
            
            if r21_value is not None and isinstance(r21_value, (int, float)) and r21_value >= 0:
                valid_entries.append({
                    'entry_date': row['entry_date'],
                    'r21_value': float(r21_value),
                    'distance': row['distance']
                })
        
        if not valid_entries:
            print(f"DEBUG - Monthly manpower: No valid R21 values found")
            return None
        
        # Sort by distance only - return closest value regardless of whether it's zero
        # As per specifications: if the cell contains 0, show 0
        valid_entries.sort(key=lambda x: x['distance'])
        closest_entry = valid_entries[0]
        return closest_entry['r21_value']
        
    except Exception as e:
        print(f"Error calculating monthly manpower count: {e}")
        return None


def get_monthly_equipment_count_for_period(progress_data, start_date, end_date):
    """Calculate monthly equipment count from progress data for a given period (row 22)
    Uses R20 as date reference with closest date matching logic"""
    try:
        if progress_data.empty:
            print(f"DEBUG - Monthly equipment: No progress data available for period {start_date} to {end_date}")
            return None
            
        # Convert period dates for comparison
        target_start = pd.to_datetime(start_date)
        target_end = pd.to_datetime(end_date)
        target_date = target_start  # Use start of period as target
        
        print(f"DEBUG - Monthly equipment: Looking for R20 date closest to {target_date.date()}")
        
        # Find all data with R22 (equipment) and R20 (date) values
        all_data = progress_data[progress_data['notes'].str.contains('R22:', na=False) & 
                                progress_data['notes'].str.contains('R20:', na=False)]
        
        if all_data.empty:
            print(f"DEBUG - Monthly equipment: No R22+R20 data found in dataset")
            return None
        
        # Since R20/R22 are often zero, use entry_date with R22 values as per requirements
        print(f"DEBUG - Monthly equipment: R20 dates often zero, using entry_date fallback with R22 values from correct row")
        
        # Calculate distance from target date using entry_date
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        all_data['distance'] = abs((all_data['entry_date'] - target_date).dt.days)
        
        # Extract R22 values (monthly equipment as per requirements)
        valid_entries = []
        for _, row in all_data.iterrows():
            notes = row.get('notes', '')
            r22_value = extract_excel_row_data(notes, 22)  # Row 22 = monthly equipment per requirements
            
            if r22_value is not None and isinstance(r22_value, (int, float)) and r22_value >= 0:
                valid_entries.append({
                    'entry_date': row['entry_date'],
                    'r22_value': float(r22_value),
                    'distance': row['distance']
                })
        
        if not valid_entries:
            print(f"DEBUG - Monthly equipment: No valid R22 values found")
            return None
        
        # Sort by distance only - return closest value regardless of whether it's zero
        # As per specifications: if the cell contains 0, show 0
        valid_entries.sort(key=lambda x: x['distance'])
        closest_entry = valid_entries[0]
        return closest_entry['r22_value']
        
    except Exception as e:
        print(f"Error calculating monthly equipment count: {e}")
        return None


def get_monthly_date_for_period(progress_data, start_date, end_date):
    """Extract monthly date data from progress data for a given period (row 20)"""
    try:
        if progress_data.empty:
            return None
            
        # Filter data for the period
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        period_data = progress_data[
            (progress_data['entry_date'] >= pd.to_datetime(start_date)) &
            (progress_data['entry_date'] <= pd.to_datetime(end_date))
        ]
        
        if period_data.empty:
            return None
        
        # Extract date from notes field (R20 represents monthly date data)
        date_values = []
        for _, row in period_data.iterrows():
            date_value = extract_excel_row_data(row.get('notes', ''), 20)
            if date_value:
                date_values.append(date_value)
        
        if date_values:
            return date_values[0]  # Return first date found
        return None
    except Exception as e:
        print(f"Error extracting monthly date: {e}")
        return None


# New functions for weekly resource tracking (using rows 17, 18, 19)
def get_weekly_manpower_count_for_period(progress_data, start_date, end_date):
    """Calculate weekly manpower count from progress data for a given period (row 18)
    Uses R17 as date reference with entry_date fallback since R17 values are zero"""
    try:
        if progress_data.empty:
            # No progress data available
            return None
            
        # Convert dates for processing
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        target_start = pd.to_datetime(start_date)
        target_end = pd.to_datetime(end_date)
        
        # Find the Thursday in the week (target date for weekly data per requirements)
        target_thursday = target_start
        while target_thursday.weekday() != 3:  # 3 = Thursday
            target_thursday = target_thursday + pd.Timedelta(days=1)
            if target_thursday > target_end:
                break
        
        # Looking for R17 date closest to Thursday
        
        # Find data with R18 (manpower) values
        all_data = progress_data[progress_data['notes'].str.contains('R18:', na=False)]
        if all_data.empty:
            # No R18 data found
            return None
        
        # Check R17 dates first, fallback to entry_date if R17 is zero
        print(f"DEBUG - Weekly manpower: R17 dates are zero, using entry_date fallback with R18 values from correct row")
        
        # Calculate distance from target Thursday using entry_date
        all_data['distance'] = abs((all_data['entry_date'] - target_thursday).dt.days)
        
        # Extract R18 values (weekly manpower as per requirements)
        valid_entries = []
        for _, row in all_data.iterrows():
            notes = row.get('notes', '')
            r18_value = extract_excel_row_data(notes, 18)  # Row 18 = weekly manpower per requirements
            
            if r18_value is not None and isinstance(r18_value, (int, float)) and r18_value >= 0:
                valid_entries.append({
                    'entry_date': row['entry_date'],
                    'r18_value': float(r18_value),
                    'distance': row['distance']
                })
        
        if not valid_entries:
            print(f"DEBUG - Weekly manpower: No valid R18 values found")
            return None
        
        # Sort by distance only - return closest value regardless of whether it's zero
        # As per specifications: if the cell contains 0, show 0
        valid_entries.sort(key=lambda x: x['distance'])
        closest_entry = valid_entries[0]
        return closest_entry['r18_value']
            
    except Exception as e:
        print(f"Error calculating weekly manpower count: {e}")
        return None


def get_weekly_equipment_count_for_period(progress_data, start_date, end_date):
    """Calculate weekly equipment count from progress data for a given period (row 19)
    Uses R17 as date reference with entry_date fallback since R17 values are zero"""
    try:
        if progress_data.empty:
            # No progress data available
            return None
            
        # Convert dates for processing
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        target_start = pd.to_datetime(start_date)
        target_end = pd.to_datetime(end_date)
        
        # Find the Thursday in the week (target date for weekly data per requirements)
        target_thursday = target_start
        while target_thursday.weekday() != 3:  # 3 = Thursday
            target_thursday = target_thursday + pd.Timedelta(days=1)
            if target_thursday > target_end:
                break
        
        # Looking for R17 date closest to Thursday
        
        # Find data with R19 (equipment) values
        all_data = progress_data[progress_data['notes'].str.contains('R19:', na=False)]
        if all_data.empty:
            # No R19 data found
            return None
        
        # Check R17 dates first, fallback to entry_date if R17 is zero
        # R17 dates are zero, using entry_date fallback
        
        # Calculate distance from target Thursday using entry_date
        all_data['distance'] = abs((all_data['entry_date'] - target_thursday).dt.days)
        
        # Extract R19 values (weekly equipment as per requirements)
        valid_entries = []
        for _, row in all_data.iterrows():
            notes = row.get('notes', '')
            r19_value = extract_excel_row_data(notes, 19)  # Row 19 = weekly equipment per requirements
            
            if r19_value is not None and isinstance(r19_value, (int, float)) and r19_value >= 0:
                valid_entries.append({
                    'entry_date': row['entry_date'],
                    'r19_value': float(r19_value),
                    'distance': row['distance']
                })
        
        if not valid_entries:
            # No valid R19 values found
            return None
        
        # Sort by distance only - return closest value regardless of whether it's zero
        # As per specifications: if the cell contains 0, show 0
        valid_entries.sort(key=lambda x: x['distance'])
        closest_entry = valid_entries[0]
        return closest_entry['r19_value']
            
    except Exception as e:
        print(f"Error calculating weekly equipment count: {e}")
        return None


def get_weekly_date_for_period(progress_data, start_date, end_date):
    """Extract weekly date data from progress data for a given period (row 17) with Thursday preference"""
    try:
        # Function called for period
        if progress_data.empty:
            # No progress data available
            return None
            
        # Processing progress data
        # Filter data for the period
        progress_data['entry_date'] = pd.to_datetime(progress_data['entry_date'])
        period_data = progress_data[
            (progress_data['entry_date'] >= pd.to_datetime(start_date)) &
            (progress_data['entry_date'] <= pd.to_datetime(end_date))
        ]
        
        # Filtered data for period
        if period_data.empty:
            # No data found in period
            return None
        
        # Extract date from notes field (R17 represents weekly date data)
        found_dates = []
        valid_entries = 0
        for _, row in period_data.iterrows():
            notes = row.get('notes', '')
            if 'R17:' not in notes:
                # R17 not found in notes
                continue
                
            date_value = extract_excel_row_data(notes, 17)
            valid_entries += 1
            
            if date_value:
                # Got date value from R17
                try:
                    # Handle different date formats and types
                    if isinstance(date_value, (int, float)):
                        # Excel date format (days since 1900-01-01)
                        if date_value > 0:
                            base_date = pd.to_datetime('1900-01-01')
                            actual_date = base_date + pd.Timedelta(days=int(date_value) - 2)
                            found_dates.append(actual_date)
                            # Converted Excel date
                        else:
                            # Invalid Excel date value
                            pass
                    elif isinstance(date_value, str) and date_value.strip():
                        # String date format
                        print(f"DEBUG - Weekly date: Attempting to parse string date: '{date_value}'")
                        parsed_date = pd.to_datetime(date_value)
                        found_dates.append(parsed_date)
                        print(f"DEBUG - Weekly date: Parsed string date {date_value} to {parsed_date}")
                    else:
                        print(f"DEBUG - Weekly date: Unable to parse date value {date_value} of type {type(date_value)}")
                except Exception as date_error:
                    print(f"DEBUG - Weekly date: Error parsing date {date_value}: {date_error}")
                    continue
            else:
                print(f"DEBUG - Weekly date: date_value is None or empty")
        
        print(f"DEBUG - Weekly date: Processed {valid_entries} entries, found {len(found_dates)} valid dates")
        
        if found_dates:
            # Find the closest Thursday to the found dates
            target_date = found_dates[0]
            print(f"DEBUG - Weekly date: Target date {target_date} (weekday: {target_date.weekday()})")
            
            # Check if target date is already Thursday (weekday 3)
            if target_date.weekday() == 3:
                print(f"DEBUG - Weekly date: Target date is already Thursday")
                return target_date
            
            # Find the closest Thursday
            days_until_thursday = (3 - target_date.weekday()) % 7  # Thursday is weekday 3
            days_since_thursday = (target_date.weekday() - 3) % 7
            
            print(f"DEBUG - Weekly date: Days until Thursday: {days_until_thursday}, Days since Thursday: {days_since_thursday}")
            
            if days_until_thursday <= days_since_thursday:
                # Next Thursday is closer
                closest_thursday = target_date + pd.Timedelta(days=days_until_thursday)
                print(f"DEBUG - Weekly date: Using next Thursday {closest_thursday}")
            else:
                # Previous Thursday is closer
                closest_thursday = target_date - pd.Timedelta(days=days_since_thursday)
                print(f"DEBUG - Weekly date: Using previous Thursday {closest_thursday}")
            
            return closest_thursday
        else:
            print(f"DEBUG - Weekly date: No valid dates found in period")
            return None
    except Exception as e:
        print(f"Error extracting weekly date: {e}")
        return None


def show_monthly_progress_table(selected_projects, monthly_periods):
    """Display monthly progress tracking table with 3 horizontal rows per project"""
    st.markdown("""
    <div class="progress-section">
        <h3>📊 جدول نسب الإنجاز الشهرية - Monthly Progress Tracking</h3>
        <p style="font-size: 12px; margin: 5px 0;">
            المخطط (صف 10) | المنفذ (صف 13 - أكبر نسبة) | المدة المنقضية (صف 11)
        </p>
    </div>
    <style>
    .monthly-table-container {
        overflow-x: auto;
        position: relative;
        direction: rtl;
    }
    .monthly-table {
        border-collapse: collapse;
        font-size: 11px;
        min-width: 100%;
        direction: rtl;
    }
    .monthly-table .fixed-columns {
        position: sticky;
        right: 0;
        z-index: 10;
        background-color: white;
        border-left: 2px solid #ddd;
    }
    .monthly-table th, .monthly-table td {
        border: 1px solid #ddd;
        padding: 4px;
        text-align: center;
        white-space: normal;
    }
    .monthly-table .project-name {
        writing-mode: horizontal-tb;
        text-align: right;
        min-width: 150px;
        max-width: 200px;
        font-weight: bold;
        background-color: #f8f9fa;
        padding: 8px;
        word-wrap: break-word;
        white-space: normal;
    }
    .monthly-table .purchase-order {
        writing-mode: horizontal-tb;
        text-align: center;
        min-width: 80px;
        max-width: 100px;
        font-weight: bold;
        background-color: #e3f2fd;
        padding: 8px;
        color: #1976d2;
    }
    .monthly-table .row-label {
        font-size: 9px;
        font-weight: bold;
        width: 80px;
        background-color: #f0f2f6;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if not monthly_periods:
        st.warning("لا توجد فترات شهرية في النطاق المحدد")
        return
    
    # Build HTML table with new structure
    table_html = '<div class="monthly-table-container">'
    table_html += '<table class="monthly-table">'
    
    # Table header
    table_html += '<thead>'
    table_html += '<tr style="background-color: #2c3e50; color: white;">'
    table_html += '<th rowspan="2" class="fixed-columns purchase-order" style="min-width: 80px;">أمر الشراء</th>'
    table_html += '<th rowspan="2" class="fixed-columns project-name" style="min-width: 150px;">اسم المشروع</th>'
    table_html += '<th rowspan="2" class="fixed-columns row-label" style="min-width: 80px;">النوع</th>'
    
    for period in monthly_periods:
        table_html += f'<th style="background-color: #34495e; min-width: 80px;">{period["display_name"]}</th>'
    
    table_html += '</tr>'
    table_html += '</thead>'
    table_html += '<tbody>'
    
    # Data rows - 2 rows per project (manpower + equipment only)
    for project in selected_projects:
        project_name = project['project_name']
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        # Row 1: Planned (صف 10)
        table_html += '<tr>'
        project_po = project.get('purchase_order', project.get('project_id', ''))
        table_html += f'<td rowspan="3" class="fixed-columns purchase-order">{project_po}</td>'
        table_html += f'<td rowspan="3" class="fixed-columns project-name">{project_name}</td>'
        table_html += '<td class="fixed-columns row-label" style="color: #1f77b4;">مخطط</td>'
        
        for period in monthly_periods:
            # Check if the month is after the project end month
            period_end = pd.to_datetime(period['end_date'])
            project_end = get_project_end_date(project_name)
            
            if project_end and period_end.replace(day=1) > project_end.replace(day=1):
                display_value = "خارج مدة المشروع"
                style = "color: #ff6b6b; font-style: italic; font-size: 8px;"
            else:
                planned_value = get_progress_percentage_for_period(
                    progress_data, period['start_date'], period['end_date'], 10
                )
                if planned_value is not None:
                    display_value = f"{planned_value * 100:.2f}%"
                    style = "color: #1f77b4; font-weight: bold;"
                else:
                    display_value = "–"
                    style = "color: #999;"
            table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 2: Actual (صف 13)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #2e8b57;">منفذ</td>'
        
        for period in monthly_periods:
            # Check if the month is after the project end month
            period_end = pd.to_datetime(period['end_date'])
            project_end = get_project_end_date(project_name)
            
            if project_end and period_end.replace(day=1) > project_end.replace(day=1):
                display_value = "خارج مدة المشروع"
                style = "color: #ff6b6b; font-style: italic; font-size: 8px;"
            else:
                actual_value = get_max_progress_percentage_for_period(
                    progress_data, period['start_date'], period['end_date'], 13
                )
                if actual_value is not None:
                    display_value = f"{actual_value * 100:.2f}%"
                    style = "color: #2e8b57; font-weight: bold;"
                else:
                    display_value = "–"
                    style = "color: #999;"
            table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 3: Elapsed (صف 11)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #e67e22;">منقضية</td>'
        
        for period in monthly_periods:
            # Check if the month is after the project end month
            period_end = pd.to_datetime(period['end_date'])
            project_end = get_project_end_date(project_name)
            
            if project_end and period_end.replace(day=1) > project_end.replace(day=1):
                elapsed_value = calculate_elapsed_percentage_beyond_end_monthly(project_name, period['end_date'])
                if elapsed_value is not None:
                    display_value = f"{elapsed_value * 100:.2f}%"
                    style = "color: #e67e22; font-weight: bold;"
                else:
                    display_value = "–"
                    style = "color: #999;"
            else:
                elapsed_value = get_progress_percentage_for_period(
                    progress_data, period['start_date'], period['end_date'], 11
                )
                if elapsed_value is not None:
                    display_value = f"{elapsed_value * 100:.2f}%"
                    style = "color: #e67e22; font-weight: bold;"
                else:
                    display_value = "–"
                    style = "color: #999;"
            table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
    
    table_html += '</tbody></table></div>'
    
    # Display table
    st.markdown(table_html, unsafe_allow_html=True)
    
    # Export option
    col1, col2 = st.columns([1, 3])
    with col1:
        if st.button("📊 تصدير الجدول الشهري", key="export_monthly_progress"):
            excel_data = create_monthly_progress_excel(selected_projects, monthly_periods)
            if excel_data:
                st.download_button(
                    label="📥 تحميل الإنجاز الشهري",
                    data=excel_data,
                    file_name=f"monthly_progress_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_monthly_progress"
                )

def show_weekly_progress_table(selected_projects, weekly_periods):
    """Display weekly progress tracking table with 3 horizontal rows per project"""
    st.markdown("""
    <div class="progress-section">
        <h3>📅 جدول نسب الإنجاز الأسبوعية - Weekly Progress Tracking</h3>
        <p style="font-size: 12px; margin: 5px 0;">
            المخطط (صف 10) | المنفذ (إدخال يدوي) | المدة المنقضية (صف 11) - يوم الخميس
        </p>
    </div>
    <style>
    .weekly-table-container {
        overflow-x: auto;
        position: relative;
        direction: rtl;
    }
    .weekly-table {
        border-collapse: collapse;
        font-size: 10px;
        min-width: 100%;
        direction: rtl;
    }
    .weekly-table .fixed-columns {
        position: sticky;
        right: 0;
        z-index: 10;
        background-color: white;
        border-left: 2px solid #ddd;
    }
    .weekly-table th, .weekly-table td {
        border: 1px solid #ddd;
        padding: 3px;
        text-align: center;
        white-space: nowrap;
    }
    .weekly-table .project-name {
        writing-mode: horizontal-tb;
        text-align: right;
        min-width: 150px;
        max-width: 200px;
        font-weight: bold;
        background-color: #f8f9fa;
        padding: 8px;
        word-wrap: break-word;
        white-space: normal;
    }
    .weekly-table .purchase-order {
        writing-mode: horizontal-tb;
        text-align: center;
        min-width: 60px;
        max-width: 80px;
        font-weight: bold;
        background-color: #e3f2fd;
        padding: 4px;
        color: #1976d2;
    }
    .weekly-table .row-label {
        font-size: 8px;
        font-weight: bold;
        width: 60px;
        background-color: #f0f2f6;
    }
    .weekly-table .month-header {
        background-color: #2c3e50;
        color: white;
        font-size: 11px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if not weekly_periods:
        st.warning("لا توجد فترات أسبوعية في النطاق المحدد")
        return
    
    # Group weeks by month for better organization
    weeks_by_month = {}
    for week in weekly_periods:
        month_key = week['month_year']
        if month_key not in weeks_by_month:
            weeks_by_month[month_key] = []
        weeks_by_month[month_key].append(week)
    
    # Build HTML table with new structure
    table_html = '<div class="weekly-table-container">'
    table_html += '<table class="weekly-table">'
    
    # Table header with month groupings
    table_html += '<thead>'
    
    # Main header row
    table_html += '<tr class="month-header">'
    table_html += '<th rowspan="2" class="fixed-columns purchase-order" style="min-width: 60px;">كود المشروع (E3)</th>'
    table_html += '<th rowspan="2" class="fixed-columns project-name" style="min-width: 150px;">اسم المشروع</th>'
    table_html += '<th rowspan="2" class="fixed-columns row-label" style="min-width: 60px;">النوع</th>'
    
    for month_key, weeks in weeks_by_month.items():
        month_name = pd.to_datetime(month_key + '-01').strftime('%B %Y')
        # Convert to Arabic
        month_mapping = {
            'January': 'يناير', 'February': 'فبراير', 'March': 'مارس',
            'April': 'أبريل', 'May': 'مايو', 'June': 'يونيو',
            'July': 'يوليو', 'August': 'أغسطس', 'September': 'سبتمبر',
            'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
        }
        for eng, ar in month_mapping.items():
            month_name = month_name.replace(eng, ar)
        
        table_html += f'<th colspan="{len(weeks)}" style="min-width: {len(weeks)*60}px;">{month_name}</th>'
    
    table_html += '</tr>'
    
    # Week header row (Thursday dates)
    table_html += '<tr style="background-color: #34495e; color: white;">'
    for month_key, weeks in weeks_by_month.items():
        for week in weeks:
            table_html += f'<th style="min-width: 60px; font-size: 8px;">{week["display_name"]}</th>'
    table_html += '</tr>'
    table_html += '</thead>'
    table_html += '<tbody>'
    
    # Data rows - 2 rows per project (manpower + equipment only)
    for project in selected_projects:
        project_name = project['project_name']
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        # Row 1: Planned (صف 10)
        table_html += '<tr>'
        project_po = project.get('purchase_order', project.get('project_id', ''))
        table_html += f'<td rowspan="3" class="fixed-columns purchase-order">{project_po}</td>'
        table_html += f'<td rowspan="3" class="fixed-columns project-name">{project_name}</td>'
        table_html += '<td class="fixed-columns row-label" style="color: #1f77b4;">مخطط</td>'
        
        for month_key, weeks in weeks_by_month.items():
            for week in weeks:
                # Check if the week is after the project end week
                thursday_date = pd.to_datetime(week['thursday_date'])
                project_end = get_project_end_date(project_name)
                
                # For weekly view, check if Thursday is in a week after project end week
                if project_end:
                    # Get the Thursday of the week containing project end date
                    project_end_weekday = project_end.weekday()
                    if project_end_weekday == 3:  # If project ends on Thursday
                        project_end_thursday = project_end
                    elif project_end_weekday < 3:  # If project ends before Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(3 - project_end_weekday))
                    else:  # If project ends after Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(7 + 3 - project_end_weekday))
                    
                    # Show message if current Thursday is after the project end week Thursday
                    if thursday_date > project_end_thursday:
                        display_value = "خارج مدة المشروع"
                        style = "color: #ff6b6b; font-style: italic; font-size: 7px;"
                    else:
                        planned_value = get_progress_percentage_for_period(
                            progress_data, week['thursday_date'], week['thursday_date'], 10
                        )
                        if planned_value is not None:
                            display_value = f"{planned_value * 100:.2f}%"
                            style = "color: #1f77b4; font-weight: bold; font-size: 9px;"
                        else:
                            display_value = "–"
                            style = "color: #999; font-size: 9px;"
                else:
                    planned_value = get_progress_percentage_for_period(
                        progress_data, week['thursday_date'], week['thursday_date'], 10
                    )
                    if planned_value is not None:
                        display_value = f"{planned_value * 100:.2f}%"
                        style = "color: #1f77b4; font-weight: bold; font-size: 9px;"
                    else:
                        display_value = "–"
                        style = "color: #999; font-size: 9px;"
                table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 2: Actual (إدخال يدوي)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #2e8b57;">منفذ</td>'
        
        for month_key, weeks in weeks_by_month.items():
            for week in weeks:
                # Check if the week is after the project end week
                thursday_date = pd.to_datetime(week['thursday_date'])
                project_end = get_project_end_date(project_name)
                
                # For weekly view, check if Thursday is in a week after project end week
                if project_end:
                    # Get the Thursday of the week containing project end date
                    project_end_weekday = project_end.weekday()
                    if project_end_weekday == 3:  # If project ends on Thursday
                        project_end_thursday = project_end
                    elif project_end_weekday < 3:  # If project ends before Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(3 - project_end_weekday))
                    else:  # If project ends after Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(7 + 3 - project_end_weekday))
                    
                    # Show message if current Thursday is after the project end week Thursday
                    if thursday_date > project_end_thursday:
                        display_value = "خارج مدة المشروع"
                        style = "color: #ff6b6b; font-style: italic; font-size: 7px;"
                    else:
                        # Placeholder for manual input - will be enhanced later
                        display_value = "يدوي"
                        style = "color: #ff9800; font-style: italic; font-size: 8px; cursor: pointer;"
                else:
                    # Placeholder for manual input - will be enhanced later
                    display_value = "يدوي"
                    style = "color: #ff9800; font-style: italic; font-size: 8px; cursor: pointer;"
                table_html += f'<td style="{style}" title="اضغط للإدخال اليدوي">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 3: Elapsed (صف 11)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #e67e22;">منقضية</td>'
        
        for month_key, weeks in weeks_by_month.items():
            for week in weeks:
                # Check if the week is after the project end week
                thursday_date = pd.to_datetime(week['thursday_date'])
                project_end = get_project_end_date(project_name)
                
                # For weekly view, check if Thursday is in a week after project end week
                if project_end:
                    # Get the Thursday of the week containing project end date
                    project_end_weekday = project_end.weekday()
                    if project_end_weekday == 3:  # If project ends on Thursday
                        project_end_thursday = project_end
                    elif project_end_weekday < 3:  # If project ends before Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(3 - project_end_weekday))
                    else:  # If project ends after Thursday
                        project_end_thursday = project_end + pd.Timedelta(days=(7 + 3 - project_end_weekday))
                    
                    # Show calculation if current Thursday is after the project end week Thursday
                    if thursday_date > project_end_thursday:
                        elapsed_value = calculate_elapsed_percentage_beyond_end_weekly(project_name, week['thursday_date'])
                        if elapsed_value is not None:
                            display_value = f"{elapsed_value * 100:.2f}%"
                            style = "color: #e67e22; font-weight: bold; font-size: 9px;"
                        else:
                            display_value = "–"
                            style = "color: #999; font-size: 9px;"
                    else:
                        elapsed_value = get_progress_percentage_for_period(
                            progress_data, week['thursday_date'], week['thursday_date'], 11
                        )
                        if elapsed_value is not None:
                            display_value = f"{elapsed_value * 100:.2f}%"
                            style = "color: #e67e22; font-weight: bold; font-size: 9px;"
                        else:
                            display_value = "–"
                            style = "color: #999; font-size: 9px;"
                else:
                    elapsed_value = get_progress_percentage_for_period(
                        progress_data, week['thursday_date'], week['thursday_date'], 11
                    )
                    if elapsed_value is not None:
                        display_value = f"{elapsed_value * 100:.1f}%"
                        style = "color: #e67e22; font-weight: bold; font-size: 9px;"
                    else:
                        display_value = "–"
                        style = "color: #999; font-size: 9px;"
                table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
    
    table_html += '</tbody></table></div>'
    
    # Display table
    st.markdown(table_html, unsafe_allow_html=True)
    
    # Manual input section
    st.markdown("#### ✏️ إدخال الإنجاز الأسبوعي اليدوي")
    
    if selected_projects and weekly_periods:
        # Simplified input form to match new table design
        with st.expander("📝 إدخال نسب الإنجاز الأسبوعية", expanded=False):
            st.markdown("""
            <div style="background-color: #e8f5e8; padding: 10px; border-radius: 5px; margin: 10px 0; border-left: 4px solid #2e8b57;">
                <strong>💡 ملاحظة:</strong> في التصميم الجديد، يمكنك رؤية 3 صفوف لكل مشروع:<br>
                • <span style="color: #1f77b4;">مخطط</span> - من صف 10 (الخميس)<br>
                • <span style="color: #2e8b57;">منفذ</span> - إدخال يدوي<br>
                • <span style="color: #e67e22;">منقضية</span> - من صف 11 (الخميس)
            </div>
            """, unsafe_allow_html=True)
            
            # Create input form
            input_col1, input_col2, input_col3 = st.columns([3, 3, 2])
            
            with input_col1:
                # Use first selected project or allow selection from selected projects
                if len(selected_projects) == 1:
                    selected_project = selected_projects[0]['project_name']
                    st.info(f"🏗️ المشروع المحدد: {selected_project}")
                else:
                    selected_project = st.selectbox(
                        "🏗️ اختر المشروع من المختارة:",
                        [proj['project_name'] for proj in selected_projects],
                        key="weekly_input_project"
                    )
            
            with input_col2:
                # Show weeks in a more user-friendly format
                week_options = []
                for week in weekly_periods:
                    week_options.append({
                        'display': f"{week['display_name']} ({week['month_year']})",
                        'value': week
                    })
                
                selected_week_idx = st.selectbox(
                    "📅 اختر الأسبوع (الخميس):",
                    range(len(week_options)),
                    format_func=lambda x: week_options[x]['display'],
                    key="weekly_input_week"
                )
                selected_week = week_options[selected_week_idx]['value']
            
            with input_col3:
                actual_percentage = st.number_input(
                    "📊 الإنجاز المنفذ %:",
                    min_value=0.0,
                    max_value=100.0,
                    step=0.1,
                    format="%.1f",
                    key="weekly_actual_input"
                )
            
            # Action buttons
            save_col1, save_col2 = st.columns([1, 3])
            with save_col1:
                if st.button("💾 حفظ الإنجاز", key="save_weekly_progress"):
                    st.success(f"✅ تم حفظ {actual_percentage:.1f}% للمشروع '{selected_project}' في أسبوع {selected_week['display_name']}")
                    st.info("ℹ️ سيتم ربط البيانات بقاعدة البيانات في التحديث القادم")
            
            # Show current project progress overview
            st.markdown("---")
            st.markdown(f"**📋 نظرة عامة على المشروع المحدد: {selected_project}**")
            
            # Get progress data for selected project only
            project_progress = st.session_state.data_manager.get_progress_data(selected_project)
            
            if not project_progress.empty:
                # Display last 3 weeks data for this project
                overview_data = []
                for week in weekly_periods[-3:]:  # Last 3 weeks
                    planned = get_progress_percentage_for_period(project_progress, week['thursday_date'], week['thursday_date'], 10)
                    elapsed = get_progress_percentage_for_period(project_progress, week['thursday_date'], week['thursday_date'], 11)
                    
                    overview_data.append({
                        'الأسبوع': week['display_name'],
                        'مخطط %': f"{planned * 100:.1f}%" if planned else "–",
                        'منفذ %': "إدخال يدوي مطلوب",
                        'منقضية %': f"{elapsed * 100:.1f}%" if elapsed else "–"
                    })
                
                if overview_data:
                    overview_df = pd.DataFrame(overview_data)
                    st.dataframe(overview_df, width=800, hide_index=True)
            else:
                st.info("لا توجد بيانات متاحة لهذا المشروع")
    
    else:
        st.info("يرجى اختيار المشاريع والفترة الزمنية أولاً من الشريط الجانبي")

def create_monthly_progress_excel(selected_projects, monthly_periods):
    """Create Excel export for monthly progress data with RTL support and enhanced formatting"""
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Monthly Progress"
        
        # Set worksheet RTL direction
        ws.sheet_view.rightToLeft = True
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        po_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        project_header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") 
        period_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        po_font = Font(bold=True, color="1976D2", name="Arial", size=10)
        project_font = Font(bold=True, color="333333", name="Arial", size=10)
        data_font = Font(name="Arial", size=9)
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # Build headers with proper structure matching the display
        col_idx = 1
        
        # PO Header
        po_cell = ws.cell(row=1, column=col_idx, value="أمر الشراء")
        po_cell.font = Font(bold=True, color="1976D2", name="Arial", size=11)
        po_cell.fill = po_header_fill
        po_cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        col_idx += 1
        
        # Project Name Header  
        proj_cell = ws.cell(row=1, column=col_idx, value="اسم المشروع")
        proj_cell.font = Font(bold=True, color="333333", name="Arial", size=11) 
        proj_cell.fill = project_header_fill
        proj_cell.alignment = right_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
        col_idx += 1
        
        # Row Type Header
        type_cell = ws.cell(row=1, column=col_idx, value="النوع")
        type_cell.font = header_font
        type_cell.fill = period_header_fill  
        type_cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = 10
        col_idx += 1
        
        # Period Headers
        for period in monthly_periods:
            period_cell = ws.cell(row=1, column=col_idx, value=period['display_name'])
            period_cell.font = header_font
            period_cell.fill = period_header_fill
            period_cell.alignment = center_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1
        
        # Data rows - 2 rows per project (manpower + equipment only) matching display structure
        current_row = 2
        for project in selected_projects:
            project_name = project['project_name']
            project_po = project.get('purchase_order', project.get('project_id', ''))
            progress_data = st.session_state.data_manager.get_progress_data(project_name)
            
            # Row 1: Planned (مخطط)
            col_idx = 1
            ws.cell(row=current_row, column=col_idx, value=project_po).font = po_font
            ws.cell(row=current_row, column=col_idx).alignment = center_alignment
            ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value=project_name).font = project_font
            ws.cell(row=current_row, column=col_idx).alignment = right_alignment
            ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            col_idx += 1
            ws.cell(row=current_row, column=col_idx, value="مخطط").font = Font(bold=True, color="1F77B4")
            ws.cell(row=current_row, column=col_idx).alignment = center_alignment
            
            col_idx += 1
            for period in monthly_periods:
                planned_value = get_progress_percentage_for_period(
                    progress_data, period['start_date'], period['end_date'], 10
                )
                display_val = f"{planned_value * 100:.2f}%" if planned_value is not None else "–"
                cell = ws.cell(row=current_row, column=col_idx, value=display_val)
                cell.font = Font(color="1F77B4", bold=True)
                cell.alignment = center_alignment
                col_idx += 1
            
            current_row += 1
            
            # Row 2: Actual (منفذ)  
            col_idx = 3  # Skip merged PO and project name
            ws.cell(row=current_row, column=col_idx, value="منفذ").font = Font(bold=True, color="2E8B57")
            ws.cell(row=current_row, column=col_idx).alignment = center_alignment
            
            col_idx += 1
            for period in monthly_periods:
                actual_value = get_max_progress_percentage_for_period(
                    progress_data, period['start_date'], period['end_date'], 13
                )
                display_val = f"{actual_value * 100:.2f}%" if actual_value is not None else "–"
                cell = ws.cell(row=current_row, column=col_idx, value=display_val)
                cell.font = Font(color="2E8B57", bold=True)
                cell.alignment = center_alignment
                col_idx += 1
                
            current_row += 1
            
            # Row 3: Elapsed (منقضية)
            col_idx = 3  # Skip merged PO and project name
            ws.cell(row=current_row, column=col_idx, value="منقضية").font = Font(bold=True, color="E67E22") 
            ws.cell(row=current_row, column=col_idx).alignment = center_alignment
            
            col_idx += 1
            for period in monthly_periods:
                if is_date_beyond_project_end(project_name, period['end_date']):
                    elapsed_value = calculate_elapsed_percentage_beyond_end_monthly(project_name, period['end_date'])
                else:
                    elapsed_value = get_progress_percentage_for_period(
                        progress_data, period['start_date'], period['end_date'], 11
                    )
                display_val = f"{elapsed_value * 100:.2f}%" if elapsed_value is not None else "–"
                cell = ws.cell(row=current_row, column=col_idx, value=display_val)
                cell.font = Font(color="E67E22", bold=True)
                cell.alignment = center_alignment
                col_idx += 1
                
            current_row += 1
            
            # Merge PO and project name cells for the 3 rows
            ws.merge_cells(start_row=current_row-3, start_column=1, end_row=current_row-1, end_column=1)
            ws.merge_cells(start_row=current_row-3, start_column=2, end_row=current_row-1, end_column=2)
        
        # Freeze panes to keep PO and project name visible
        ws.freeze_panes = ws['C2']  # Freeze first two columns
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    except Exception as e:
        st.error(f"Error creating Excel export: {e}")
        return None

def show_financial_summary_charts(all_projects, data_type):
    """Show summary charts for financial data"""
    st.markdown("### Financial Summary Charts")
    
    # Create summary data
    chart_data = []
    for project in all_projects:
        project_name = project['project_name']
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        if not progress_data.empty:
            total_actual_cost = progress_data['actual_cost'].sum()
            total_planned_cost = progress_data['planned_cost'].sum()
            budget = project.get('total_budget', 0)
            
            chart_data.append({
                'Project': project_name,
                'Actual Cost': total_actual_cost,
                'Planned Cost': total_planned_cost,
                'Budget': budget
            })
    
    if chart_data:
        import plotly.express as px
        chart_df = pd.DataFrame(chart_data)
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Cost comparison chart
            fig1 = px.bar(
                chart_df, 
                x='Project', 
                y=['Actual Cost', 'Planned Cost', 'Budget'],
                title="Cost Comparison by Project",
                barmode='group'
            )
            st.plotly_chart(fig1, use_container_width=True)
        
        with col2:
            # Cost variance chart
            chart_df['Cost Variance'] = chart_df['Actual Cost'] - chart_df['Planned Cost']
            fig2 = px.bar(
                chart_df, 
                x='Project', 
                y='Cost Variance',
                title="Cost Variance by Project",
                color='Cost Variance',
                color_continuous_scale=['red', 'yellow', 'green']
            )
            st.plotly_chart(fig2, use_container_width=True)


def progress_tab():
    """Enhanced Project progress tracking tab with monthly and weekly views"""
    st.markdown("""
    <div class="rtl">
        <h2>📈 تقدم المشاريع - Progress Tracking</h2>
        <p>عرض تفصيلي لنسب الإنجاز الشهرية والأسبوعية</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Use projects selected from sidebar
    all_projects = st.session_state.data_manager.get_all_projects()
    
    if not all_projects:
        st.warning("لا توجد مشاريع متاحة. يرجى إضافة مشاريع أولاً.")
        return
        
    # Get selected projects from sidebar
    selected_project_names = st.session_state.get('selected_projects', [])
    
    if not selected_project_names:
        st.warning("يرجى اختيار مشروع واحد على الأقل من القائمة الجانبية.")
        return
        
    selected_projects = [proj for proj in all_projects if proj['project_name'] in selected_project_names]
    
    if not selected_projects:
        st.warning("المشاريع المختارة غير متاحة.")
        return
    
    st.markdown(f"**المشاريع المختارة:** {len(selected_projects)} مشروع")
    
    # Date range selection for progress tracking
    st.markdown("### 📅 اختيار المدى الزمني")
    date_cols = st.columns(2)
    
    with date_cols[0]:
        start_date = st.date_input(
            "📅 من تاريخ", 
            value=date.today().replace(month=1, day=1),  # Start of current year
            key="progress_date_start"
        )
    with date_cols[1]:
        end_date = st.date_input(
            "📅 إلى تاريخ", 
            value=date.today() + timedelta(days=365),  # One year from now
            key="progress_date_end"
        )
    
    # Generate time period columns
    monthly_periods = generate_monthly_columns(start_date, end_date)
    weekly_periods = generate_weekly_columns(start_date, end_date)
    
    # Display both sections
    show_monthly_progress_table(selected_projects, monthly_periods)
    st.divider()
    show_weekly_progress_table(selected_projects, weekly_periods)


def get_project_code_from_e3(project):
    """Extract project code from E3 cell (stored in project data)"""
    try:
        # First try purchase_order field
        if 'purchase_order' in project and project['purchase_order']:
            return project['purchase_order']
        
        # Then try project_id field
        if 'project_id' in project and project['project_id']:
            return project['project_id']
        
        # Fallback to project name abbreviation
        project_name = project.get('project_name', '')
        if project_name:
            # Create abbreviation from first letters of words
            words = project_name.split()
            if len(words) >= 2:
                return ''.join([word[0].upper() for word in words[:3]])
            else:
                return project_name[:6].upper()
        
        return 'N/A'
    except Exception as e:
        print(f"DEBUG - get_project_code_from_e3 error: {e}")
        return 'N/A'


@st.cache_data(ttl=300)  # Cache for 5 minutes to improve performance
def get_closest_monthly_value(progress_data, period_start, period_end, value_row, date_row):
    """Get closest monthly value using date matching logic with R20/R21/R22 rows
    
    Args:
        progress_data: DataFrame with progress data
        period_start: Start date of the period
        period_end: End date of the period
        value_row: Row number for the value (18 for manpower, 19 for equipment)
        date_row: Row number for the date (17 for monthly)
    """
    try:
        if progress_data.empty:
            return None
            
        # Convert dates to datetime for comparison
        from datetime import datetime
        import pandas as pd
        
        if isinstance(period_start, str):
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
        if isinstance(period_end, str):
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        
        # Target date for matching (middle of period)
        target_date = period_start + (period_end - period_start) / 2
        
        print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Looking for R{value_row} values (monthly {'manpower' if value_row == 21 else 'equipment'} per requirements)")
        print(f"DEBUG - Monthly: Progress data has {len(progress_data)} rows")
        
        # Extract values and dates from notes field
        closest_value = None
        closest_distance = float('inf')
        closest_date = None
        
        for _, row in progress_data.iterrows():
            if pd.isna(row.get('notes')):
                continue
                
            # Extract value from the specified row (R21 or R22)
            value = extract_excel_row_data(row['notes'], value_row)
            if value is None:
                continue
            # Allow zero values as per specifications - zeros should be displayed
            value = float(value)
                
            # Extract date from R20 (monthly date row)
            date_value = extract_excel_row_data(row['notes'], date_row)
            print(f"DEBUG - Monthly: Extracted R{date_row} date: {date_value}")
            
            # Process the date value whether it's a string, number, or None
            actual_date = None
            if date_value is None or date_value == 0:
                print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: R{date_row} dates are zero, using entry_date fallback with R{value_row} values from correct row")
                
                # Use entry_date from the row
                if 'entry_date' in row and pd.notna(row['entry_date']):
                    try:
                        if isinstance(row['entry_date'], str):
                            entry_date = datetime.strptime(row['entry_date'], '%Y-%m-%d').date()
                        elif hasattr(row['entry_date'], 'date'):
                            # Handle pandas Timestamp
                            entry_date = row['entry_date'].date()
                        else:
                            # Already a date object
                            entry_date = row['entry_date']
                        
                        # Ensure target_date is also a date object for comparison
                        if hasattr(target_date, 'date'):
                            target_date = target_date.date()
                        
                        # Calculate distance from target date with monthly matching (ignore day)
                        target_month_year = (target_date.year, target_date.month)
                        entry_month_year = (entry_date.year, entry_date.month)
                        
                        if target_month_year == entry_month_year:
                            # Perfect month match - use this value
                            closest_distance = 0
                            closest_value = value
                            print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Perfect month match found for {entry_month_year}")
                            break  # Perfect match found
                        else:
                            # Calculate month distance
                            month_distance = abs((target_date.year - entry_date.year) * 12 + (target_date.month - entry_date.month))
                            
                            # Apply tie-breaker logic: prefer closer months, and for equal distances prefer later dates
                            should_update = False
                            if month_distance < closest_distance:
                                should_update = True
                            elif month_distance == closest_distance and closest_date is not None:
                                # Tie-breaker: prefer later date when distances are equal
                                if entry_date > closest_date:
                                    should_update = True
                            
                            if should_update:
                                closest_distance = month_distance
                                closest_value = value
                                closest_date = entry_date
                            
                    except Exception as e:
                        print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Error parsing entry_date: {e}")
                        continue
            elif isinstance(date_value, str):
                # Handle string dates like "2023-12-31"
                try:
                    actual_date = datetime.strptime(date_value, '%Y-%m-%d').date()
                    print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Parsed R{date_row} string date: {actual_date}")
                    
                    # Ensure target_date is also a date object for comparison
                    if hasattr(target_date, 'date'):
                        target_date = target_date.date()
                    
                    # Calculate distance from target date with monthly matching (ignore day)
                    target_month_year = (target_date.year, target_date.month)
                    actual_month_year = (actual_date.year, actual_date.month)
                    
                    if target_month_year == actual_month_year:
                        # Perfect month match - use this value
                        closest_distance = 0
                        closest_value = value
                        print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Perfect month match found for {actual_month_year}")
                        break  # Perfect match found
                    else:
                        # Calculate month distance
                        month_distance = abs((target_date.year - actual_date.year) * 12 + (target_date.month - actual_date.month))
                        
                        # Apply tie-breaker logic: prefer closer months, and for equal distances prefer later dates
                        should_update = False
                        if month_distance < closest_distance:
                            should_update = True
                        elif month_distance == closest_distance and closest_date is not None:
                            # Tie-breaker: prefer later date when distances are equal
                            if actual_date > closest_date:
                                should_update = True
                        
                        if should_update:
                            closest_distance = month_distance
                            closest_value = value
                            closest_date = actual_date
                            print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Found closer match: R{value_row}={value} at distance {month_distance} months (date: {actual_date})")
                except Exception as e:
                    print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: Failed to parse date string {date_value}: {e}")
                    continue
        
        if closest_value is not None:
            # Return closest value regardless of whether it's zero - as per specifications
            return closest_value
        else:
            print(f"DEBUG - Monthly {'manpower' if value_row == 21 else 'equipment'}: No R{value_row} data found for period")
            return None
            
    except Exception as e:
        print(f"Error calculating monthly {'manpower' if value_row == 21 else 'equipment'} count: {e}")
        return None


def show_monthly_resources_table(selected_projects, monthly_periods):
    """Display monthly workforce and equipment tracking table with project code (E3) identification and closest date matching"""
    st.markdown("""
    <div class="progress-section">
        <h3>👷‍♂️ متابعة أعداد العمالة والمعدات شهرياً - Monthly Manpower & Equipment Tracking</h3>
        <p style="font-size: 12px; margin: 5px 0;">
            التاريخ (R20) | عدد العمالة المخططة شهرياً (R21) | عدد المعدات المخططة شهرياً (R22)
        </p>
    </div>
    <style>
    .monthly-resources-table-container {
        overflow-x: auto;
        position: relative;
        direction: rtl;
    }
    .monthly-resources-table {
        border-collapse: collapse;
        font-size: 11px;
        min-width: 100%;
        direction: rtl;
    }
    .monthly-resources-table .fixed-columns {
        position: sticky;
        right: 0;
        z-index: 10;
        background-color: white;
        border-left: 2px solid #ddd;
    }
    .monthly-resources-table th, .monthly-resources-table td {
        border: 1px solid #ddd;
        padding: 4px;
        text-align: center;
        white-space: normal;
    }
    .monthly-resources-table .project-name {
        writing-mode: horizontal-tb;
        text-align: right;
        min-width: 150px;
        max-width: 200px;
        font-weight: bold;
        background-color: #f8f9fa;
        padding: 8px;
        word-wrap: break-word;
        white-space: normal;
    }
    .monthly-resources-table .purchase-order {
        writing-mode: horizontal-tb;
        text-align: center;
        min-width: 80px;
        max-width: 100px;
        font-weight: bold;
        background-color: #e3f2fd;
        padding: 8px;
        color: #1976d2;
    }
    .monthly-resources-table .row-label {
        font-size: 9px;
        font-weight: bold;
        width: 80px;
        background-color: #f0f2f6;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if not monthly_periods:
        st.warning("لا توجد فترات شهرية في النطاق المحدد")
        return
    
    # Build HTML table with new structure (2 rows per project only)
    table_html = '<div class="monthly-resources-table-container">'
    table_html += '<table class="monthly-resources-table">'
    
    # Table header
    table_html += '<thead>'
    table_html += '<tr style="background-color: #2c3e50; color: white;">'
    table_html += '<th rowspan="2" class="fixed-columns purchase-order" style="min-width: 80px;">كود المشروع (E3)</th>'
    table_html += '<th rowspan="2" class="fixed-columns project-name" style="min-width: 150px;">اسم المشروع</th>'
    table_html += '<th rowspan="2" class="fixed-columns row-label" style="min-width: 80px;">النوع</th>'
    
    for period in monthly_periods:
        table_html += f'<th style="background-color: #34495e; min-width: 80px;">{period["display_name"]}</th>'
    
    table_html += '</tr>'
    table_html += '</thead>'
    table_html += '<tbody>'
    
    # Data rows - 2 rows per project (manpower + equipment only)
    for project in selected_projects:
        project_name = project['project_name']
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        
        # Get project code from E3 (stored in project data)
        project_code = get_project_code_from_e3(project)
        
        # Row 1: عدد العمالة المخططة شهرياً (R21)
        table_html += '<tr>'
        table_html += f'<td rowspan="2" class="fixed-columns purchase-order">{project_code}</td>'
        table_html += f'<td rowspan="2" class="fixed-columns project-name">{project_name}</td>'
        table_html += '<td class="fixed-columns row-label" style="color: #e67e22;">عدد العمالة</td>'
        
        for period in monthly_periods:
            # Get closest date match for monthly manpower (R21) with R20 date
            workforce_count = get_closest_monthly_value(
                progress_data, period['start_date'], period['end_date'], 21, 20
            )
            if workforce_count is not None:
                display_value = f"{int(round(workforce_count))}"
                if workforce_count > 0:
                    style = "color: #e67e22; font-weight: bold;"
                else:
                    style = "color: #666; font-weight: normal;"  # Show zeros with different styling
            else:
                display_value = "–"
                style = "color: #999;"
            table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 2: عدد المعدات المخططة شهرياً (R22)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #3498db;">عدد المعدات</td>'
        
        for period in monthly_periods:
            # Get closest date match for monthly equipment (R22) with R20 date
            equipment_count = get_closest_monthly_value(
                progress_data, period['start_date'], period['end_date'], 22, 20
            )
            if equipment_count is not None:
                display_value = f"{int(round(equipment_count))}"
                if equipment_count > 0:
                    style = "color: #3498db; font-weight: bold;"
                else:
                    style = "color: #666; font-weight: normal;"  # Show zeros with different styling
            else:
                display_value = "–"
                style = "color: #999;"
            table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
    
    table_html += '</tbody></table></div>'
    
    # Display table
    st.markdown(table_html, unsafe_allow_html=True)



@st.cache_data(ttl=300)  # Cache for 5 minutes to improve performance
def get_closest_weekly_value(progress_data, period_start, period_end, value_row, date_row):
    """Get closest weekly value using date matching logic with R17/R18/R19 rows
    
    Args:
        progress_data: DataFrame with progress data
        period_start: Start date of the period
        period_end: End date of the period
        value_row: Row number for the value (18 for manpower, 19 for equipment)
        date_row: Row number for the date (17 for weekly)
    """
    try:
        if progress_data.empty:
            return None
            
        # Convert dates to datetime for comparison
        from datetime import datetime
        import pandas as pd
        
        if isinstance(period_start, str):
            period_start = datetime.strptime(period_start, '%Y-%m-%d').date()
        if isinstance(period_end, str):
            period_end = datetime.strptime(period_end, '%Y-%m-%d').date()
        
        # Target date for matching (middle of period)
        target_date = period_start + (period_end - period_start) / 2
        # Ensure target_date is a date object
        if hasattr(target_date, 'date'):
            target_date = target_date.date()
        
        print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: Looking for R{value_row} values (weekly {'manpower' if value_row == 18 else 'equipment'} per requirements)")
        
        # Extract values and dates from notes field
        closest_value = None
        closest_distance = float('inf')
        closest_date = None
        
        for _, row in progress_data.iterrows():
            if pd.isna(row.get('notes')):
                continue
                
            # Extract value from the specified row (R18 or R19)
            value = extract_excel_row_data(row['notes'], value_row)
            if value is None:
                continue
            # Allow zero values as per specifications - zeros should be displayed
            value = float(value)
                
            # Extract date from R17 (weekly date row)
            date_value = extract_excel_row_data(row['notes'], date_row)
            
            # Process the date value using the comprehensive helper function
            actual_date = parse_excel_maybe_date(date_value)
            
            if actual_date is None:
                # Use entry_date as fallback
                print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: R{date_row} could not be parsed (value: {date_value}), using entry_date fallback")
                if 'entry_date' in row and pd.notna(row['entry_date']):
                    actual_date = parse_excel_maybe_date(row['entry_date'])
                    if actual_date is None:
                        print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: Could not parse entry_date either, skipping")
                        continue
                    else:
                        print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: Using entry_date {actual_date} as fallback")
                else:
                    continue
            else:
                print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: Parsed R{date_row} date: {actual_date} (from value: {date_value})")
            
            # Calculate distance if we have a valid date
            if actual_date:
                distance = abs((target_date - actual_date).days)
                
                # Apply tie-breaker logic: prefer closer dates, and for equal distances prefer later dates
                should_update = False
                if distance < closest_distance:
                    should_update = True
                elif distance == closest_distance and closest_date is not None:
                    # Tie-breaker: prefer later date when distances are equal
                    if actual_date > closest_date:
                        should_update = True
                
                if should_update:
                    closest_distance = distance
                    closest_value = value
                    closest_date = actual_date
                    print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: Found closer match: R{value_row}={value} at distance {distance} days (date: {actual_date})")
        
        if closest_value is not None:
            # Return closest value regardless of whether it's zero - as per specifications
            return closest_value
        else:
            print(f"DEBUG - Weekly {'manpower' if value_row == 18 else 'equipment'}: No R{value_row} data found for period")
            return None
            
    except Exception as e:
        print(f"Error calculating weekly {'manpower' if value_row == 18 else 'equipment'} count: {e}")
        return None


def show_weekly_resources_table(selected_projects, weekly_periods):
    """Display weekly workforce and equipment tracking table with project code (E3) identification and closest date matching"""
    st.markdown("""
    <div class="progress-section">
        <h3>👷‍♀️ متابعة أعداد العمالة والمعدات أسبوعياً - Weekly Manpower & Equipment Tracking</h3>
        <p style="font-size: 12px; margin: 5px 0;">
            التاريخ (R17) | عدد العمالة المخطط أسبوعياً (R18) | عدد المعدات المخطط أسبوعياً (R19)
        </p>
    </div>
    <style>
    .weekly-resources-table-container {
        overflow-x: auto;
        position: relative;
        direction: rtl;
        max-height: 70vh;
    }
    .weekly-resources-table {
        border-collapse: collapse;
        font-size: 10px;
        min-width: 100%;
        direction: rtl;
    }
    .weekly-resources-table .fixed-columns {
        position: sticky;
        right: 0;
        z-index: 10;
        background-color: white;
        border-left: 2px solid #ddd;
    }
    .weekly-resources-table th, .weekly-resources-table td {
        border: 1px solid #ddd;
        padding: 3px;
        text-align: center;
        white-space: nowrap;
    }
    .weekly-resources-table .project-name {
        writing-mode: horizontal-tb;
        text-align: right;
        min-width: 150px;
        max-width: 200px;
        font-weight: bold;
        background-color: #f8f9fa;
        padding: 8px;
        word-wrap: break-word;
        white-space: normal;
    }
    .weekly-resources-table .purchase-order {
        writing-mode: horizontal-tb;
        text-align: center;
        min-width: 60px;
        max-width: 80px;
        font-weight: bold;
        background-color: #e3f2fd;
        padding: 4px;
        color: #1976d2;
    }
    .weekly-resources-table .row-label {
        font-size: 8px;
        font-weight: bold;
        width: 60px;
        background-color: #f0f2f6;
    }
    .weekly-resources-table .month-header {
        background-color: #2c3e50;
        color: white;
        font-size: 11px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    if not weekly_periods:
        st.warning("لا توجد فترات أسبوعية في النطاق المحدد")
        return
    
    # Group weeks by month for better organization
    weeks_by_month = {}
    for week in weekly_periods:
        month_key = week['month_year']
        if month_key not in weeks_by_month:
            weeks_by_month[month_key] = []
        weeks_by_month[month_key].append(week)
    
    # Build HTML table with new structure
    table_html = '<div class="weekly-resources-table-container">'
    table_html += '<table class="weekly-resources-table">'
    
    # Table header with month groupings
    table_html += '<thead>'
    
    # Main header row
    table_html += '<tr class="month-header">'
    table_html += '<th rowspan="2" class="fixed-columns purchase-order" style="min-width: 60px;">كود المشروع (E3)</th>'
    table_html += '<th rowspan="2" class="fixed-columns project-name" style="min-width: 150px;">اسم المشروع</th>'
    table_html += '<th rowspan="2" class="fixed-columns row-label" style="min-width: 60px;">النوع</th>'
    
    for month_key, weeks in weeks_by_month.items():
        month_name = pd.to_datetime(month_key + '-01').strftime('%B %Y')
        # Convert to Arabic
        month_mapping = {
            'January': 'يناير', 'February': 'فبراير', 'March': 'مارس',
            'April': 'أبريل', 'May': 'مايو', 'June': 'يونيو',
            'July': 'يوليو', 'August': 'أغسطس', 'September': 'سبتمبر',
            'October': 'أكتوبر', 'November': 'نوفمبر', 'December': 'ديسمبر'
        }
        for en_month, ar_month in month_mapping.items():
            month_name = month_name.replace(en_month, ar_month)
        
        table_html += f'<th colspan="{len(weeks)}" style="background-color: #34495e;">{month_name}</th>'
    
    table_html += '</tr>'
    
    # Week headers
    table_html += '<tr class="month-header">'
    for month_key, weeks in weeks_by_month.items():
        for week in weeks:
            table_html += f'<th style="background-color: #34495e; min-width: 50px;">{week["display_name"]}</th>'
    table_html += '</tr>'
    table_html += '</thead>'
    table_html += '<tbody>'
    
    # Data rows - 2 rows per project (manpower + equipment only)
    for project in selected_projects:
        project_name = project['project_name']
        print(f"DEBUG - Processing project: {project_name}")
        progress_data = st.session_state.data_manager.get_progress_data(project_name)
        print(f"DEBUG - Got {len(progress_data)} progress records for {project_name}")
        
        # Get project code from E3 (stored in project data)
        project_code = get_project_code_from_e3(project)
        
        # Row 1: عدد العمالة المخطط أسبوعياً (R18)
        table_html += '<tr>'
        table_html += f'<td rowspan="2" class="fixed-columns purchase-order">{project_code}</td>'
        table_html += f'<td rowspan="2" class="fixed-columns project-name">{project_name}</td>'
        table_html += '<td class="fixed-columns row-label" style="color: #e67e22;">عدد العمالة</td>'
        
        for month_key, weeks in weeks_by_month.items():
            for week in weeks:
                # Get closest date match for weekly manpower (R18) with R17 date
                workforce_count = get_closest_weekly_value(
                    progress_data, week['week_start'], week['week_end'], 18, 17
                )
                if workforce_count is not None:
                    display_value = f"{int(round(workforce_count))}"
                    if workforce_count > 0:
                        style = "color: #e67e22; font-weight: bold;"
                    else:
                        style = "color: #666; font-weight: normal;"  # Show zeros with different styling
                else:
                    display_value = "–"
                    style = "color: #999;"
                table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
        
        # Row 2: عدد المعدات المخطط أسبوعياً (R19)
        table_html += '<tr>'
        table_html += '<td class="fixed-columns row-label" style="color: #3498db;">عدد المعدات</td>'
        
        for month_key, weeks in weeks_by_month.items():
            for week in weeks:
                # Get closest date match for weekly equipment (R19) with R17 date
                equipment_count = get_closest_weekly_value(
                    progress_data, week['week_start'], week['week_end'], 19, 17
                )
                if equipment_count is not None:
                    display_value = f"{int(round(equipment_count))}"
                    if equipment_count > 0:
                        style = "color: #3498db; font-weight: bold;"
                    else:
                        style = "color: #666; font-weight: normal;"  # Show zeros with different styling
                else:
                    display_value = "–"
                    style = "color: #999;"
                table_html += f'<td style="{style}">{display_value}</td>'
        table_html += '</tr>'
    
    table_html += '</tbody></table></div>'
    
    # Display table
    st.markdown(table_html, unsafe_allow_html=True)


def resources_tab():
    """Resources and team management tab - Manpower and Equipment Tracking"""
    st.markdown("""
    <div class="rtl">
        <h2>👥 الموارد والفرق - Resources & Teams</h2>
        <p>متابعة أعداد العمالة والمعدات الشهرية والأسبوعية</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Use projects selected from sidebar
    all_projects = st.session_state.data_manager.get_all_projects()
    
    if not all_projects:
        st.warning("لا توجد مشاريع متاحة. يرجى إضافة مشاريع أولاً.")
        return
        
    # Get selected projects from sidebar
    selected_project_names = st.session_state.get('selected_projects', [])
    
    if not selected_project_names:
        st.warning("يرجى اختيار مشروع واحد على الأقل من القائمة الجانبية.")
        return
        
    selected_projects = [proj for proj in all_projects if proj['project_name'] in selected_project_names]
    
    if not selected_projects:
        st.warning("المشاريع المختارة غير متاحة.")
        return
    
    st.markdown(f"**المشاريع المختارة:** {len(selected_projects)} مشروع")
    
    # Date range selection for resources tracking (simplified)
    date_cols = st.columns(2)
    
    with date_cols[0]:
        start_date = st.date_input(
            "📅 من تاريخ", 
            value=date(2023, 11, 20),  # Start from when actual resource data exists
            key="resources_date_start"
        )
    with date_cols[1]:
        end_date = st.date_input(
            "📅 إلى تاريخ", 
            value=date(2024, 2, 28),  # Cover the period with actual data
            key="resources_date_end"
        )
    
    # Generate time period columns
    monthly_periods = generate_monthly_columns(start_date, end_date)
    weekly_periods = generate_weekly_columns(start_date, end_date)
    
    # Display both sections
    show_monthly_resources_table(selected_projects, monthly_periods)
    st.divider()
    show_weekly_resources_table(selected_projects, weekly_periods)


def reports_tab():
    """Interactive reports with filters"""
    st.markdown('<div class="rtl"><h2>📋 التقارير التفاعلية</h2></div>', unsafe_allow_html=True)
    
    # Project list and filters
    all_projects = st.session_state.data_manager.get_all_projects()
    if not all_projects:
        st.warning("لا توجد مشاريع متاحة")
        return
    
    # Date filters
    col1, col2 = st.columns(2)
    
    with col1:
        start_date = st.date_input("تاريخ البداية", key="report_start_date")
    with col2:
        end_date = st.date_input("تاريخ النهاية", key="report_end_date")
    
    # Use projects selected from sidebar
    selected_for_report = st.session_state.get('selected_projects', [])
    
    if not selected_for_report:
        st.warning("يرجى اختيار مشروع واحد على الأقل من القائمة الجانبية.")
        return
        
    st.info(f"سيتم إنشاء التقرير للمشاريع المختارة: {len(selected_for_report)} مشروع")
    
    # Display interactive dashboard
    if selected_for_report:
        st.markdown('<div class="rtl"><h3>📊 لوحة التقرير التفاعلية</h3></div>', unsafe_allow_html=True)
        
        # Summary table
        summary_data = []
        for project_name in selected_for_report:
            project = st.session_state.data_manager.get_project_by_name(project_name)
            progress = st.session_state.data_manager.get_progress_data(project_name)
            
            if project and not progress.empty:
                latest = progress.iloc[-1]
                summary_data.append({
                    'اسم المشروع': project_name,
                    'الميزانية': f"{project.get('total_budget', 0):,.0f}",
                    'الإنجاز المخطط': f"{latest.get('planned_completion', 0):.1f}%",
                    'الإنجاز الفعلي': f"{latest.get('actual_completion', 0):.1f}%",
                    'التكلفة الفعلية': f"{latest.get('actual_cost', 0):,.0f}",
                    'الحالة': '🟢 متقدم' if latest.get('actual_completion', 0) > latest.get('planned_completion', 0) else '🔴 متأخر'
                })
        
        if summary_data:
            import pandas as pd
            df = pd.DataFrame(summary_data)
            st.dataframe(df, use_container_width=True)


def settings_tab():
    st.markdown('<div class="rtl"><h2>⚙️ الإعدادات</h2></div>', unsafe_allow_html=True)
    
    # Cash flow template section
    st.markdown('<div class="rtl"><h3>قالب التدفقات النقدية</h3></div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown('<div class="rtl"><h4>تصدير قالب Excel</h4></div>', unsafe_allow_html=True)
        st.markdown('<div class="rtl">تحميل قالب Excel للمشاريع مع جميع الأعمدة المطلوبة</div>', unsafe_allow_html=True)
        
        # حصر الخيار على قالب البيانات الحالية فقط
        st.markdown('<div class="rtl">تصدير قالب Excel مع البيانات الحالية</div>', unsafe_allow_html=True)
        
        if st.button("تحميل قالب المشاريع"):
            exporter = ExcelExporter(st.session_state.data_manager)
            
            # Export using the original imported file to preserve exact formatting
            template_data = exporter.export_project_template()
            
            if template_data:
                filename = "projects_with_data.xlsx"
                st.download_button(
                    label="📥 تحميل القالب",
                    data=template_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("تم تحميل الملف الأصلي المستورد بنفس التنسيق والبيانات!")
                st.info("الملف مطابق تماماً للملف الذي تم استيراده مع الحفاظ على جميع التنسيقات والبيانات")
            else:
                st.error("لا يوجد ملف أصلي محفوظ")
                st.warning("يجب استيراد ملف Excel أولاً من قسم إدخال البيانات لتتمكن من تصدير القالب")
                st.info("بعد استيراد ملف Excel، سيتم حفظه ويمكنك تصديره كما هو بنفس التنسيق")
    
    with col2:
        st.markdown('<div class="rtl"><h4>استيراد قالب Excel</h4></div>', unsafe_allow_html=True)
        st.markdown('<div class="rtl">ارفع ملف Excel المملوء بالمشاريع لاستيراده وحساب البيانات تلقائياً</div>', unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "اختر ملف Excel",
            type=['xlsx', 'xls'],
            key="project_template_upload"
        )
        
        if uploaded_file is not None:
            if st.button("استيراد المشاريع من القالب"):
                exporter = ExcelExporter(st.session_state.data_manager)
                result = exporter.import_project_template(uploaded_file)
                
                if result['success']:
                    total_processed = result.get('imported_count', 0) + result.get('updated_count', 0)
                    st.success(f"تم معالجة {total_processed} مشروع بنجاح! ({result.get('imported_count', 0)} جديد، {result.get('updated_count', 0)} محدث)")
                    
                    # Display imported projects summary
                    if result.get('projects'):
                        st.markdown('<div class="rtl"><h5>المشاريع المعالجة:</h5></div>', unsafe_allow_html=True)
                        for project in result['projects']:
                            status_icon = "🆕" if project['status'] == 'new' else "🔄"
                            status_text = "جديد" if project['status'] == 'new' else "محدث"
                            
                            with st.expander(f"{status_icon} {project['project_name']} ({project['project_id']}) - {status_text}"):
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.write(f"**من:** {project['start_date']}")
                                    st.write(f"**إلى:** {project['end_date']}")
                                    st.write(f"**الميزانية:** {project['total_budget']:,.2f}")
                                with col2:
                                    st.write(f"**الحالة:** {status_text}")
                                    if project['status'] == 'updated':
                                        st.write("**تم تحديث البيانات والتدفقات النقدية**")
                    
                    st.info("سيتم حساب منحنى التكلفة (S-Curve) والتقارير المالية تلقائياً من البيانات المستوردة")
                    st.rerun()
                else:
                    st.error(result.get('message', 'فشل في استيراد/تحديث المشاريع'))
                    
                    # Display detailed error messages
                    if result.get('error_details'):
                        st.markdown('<div class="rtl"><h6>تفاصيل الأخطاء:</h6></div>', unsafe_allow_html=True)
                        for error in result['error_details']:
                            st.error(f"❌ {error}")
                    
                    # Display warnings
                    if result.get('warnings'):
                        st.markdown('<div class="rtl"><h6>تحذيرات:</h6></div>', unsafe_allow_html=True)
                        for warning in result['warnings']:
                            st.warning(f"⚠️ {warning}")
                    
                    # Display helpful hints
                    with st.expander("💡 نصائح لحل المشاكل"):
                        st.markdown('''
                        <div class="rtl">
                        <p><strong>تأكد من الآتي:</strong></p>
                        <ul>
                            <li>الملف من نوع Excel (.xlsx أو .xls)</li>
                            <li>أسماء المشاريع موجودة في الخلية B3 من كل ورقة عمل</li>
                            <li>معرف المشروع موجود في الخلية E3</li>
                            <li>تواريخ البداية والانتهاء في التنسيق الصحيح</li>
                            <li>الميزانية مكتوبة كأرقام (بدون حروف)</li>
                            <li>البيانات المالية في الصفوف الصحيحة (من الصف 7 فما فوق)</li>
                        </ul>
                        <p><strong>للحصول على أفضل النتائج:</strong></p>
                        <ul>
                            <li>استخدم القالب المُحمَّل من النظام</li>
                            <li>لا تغير هيكل الجدول الأساسي</li>
                            <li>تأكد من أن كل مشروع له ورقة عمل منفصلة</li>
                        </ul>
                        </div>
                        ''', unsafe_allow_html=True)
    
    st.divider()
    
    # Backup and restore
    st.markdown('<div class="rtl"><h3>النسخ الاحتياطي والاستعادة</h3></div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("إنشاء نسخة احتياطية"):
            backup_success = st.session_state.data_manager.create_backup()
            if backup_success:
                st.success("تم إنشاء النسخة الاحتياطية بنجاح!")
            else:
                st.error("فشل في إنشاء النسخة الاحتياطية")
    
    with col2:
        uploaded_file = st.file_uploader("استعادة من نسخة احتياطية", type=['zip'])
        if uploaded_file and st.button("استعادة البيانات"):
            restore_success = st.session_state.data_manager.restore_backup(uploaded_file)
            if restore_success:
                st.success("تم استعادة البيانات بنجاح!")
                st.rerun()
            else:
                st.error("فشل في استعادة البيانات")
    
    # Data management
    st.markdown('<div class="rtl"><h3>إدارة البيانات</h3></div>', unsafe_allow_html=True)
    
    if st.button("مسح جميع البيانات", type="secondary"):
        if st.button("تأكيد المسح", type="secondary"):
            clear_success = st.session_state.data_manager.clear_all_data()
            if clear_success:
                st.success("تم مسح جميع البيانات")
                st.rerun()
            else:
                st.error("فشل في مسح البيانات")
    
    # Data statistics
    st.markdown('<div class="rtl"><h3>إحصائيات البيانات</h3></div>', unsafe_allow_html=True)
    stats = st.session_state.data_manager.get_data_statistics()
    
    if stats:
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("عدد المشاريع", stats.get('total_projects', 0))
        with col2:
            st.metric("إجمالي السجلات", stats.get('total_records', 0))
        with col3:
            st.metric("حجم البيانات", f"{stats.get('data_size', 0):.2f} MB")

if __name__ == "__main__":
    main()
